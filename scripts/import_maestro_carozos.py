"""Importador de datos del Excel MAESTRO CAROZOS FINAL COMPLETO CG.

Lee la hoja 'CAROZOS' del Excel y:
1. Crea/actualiza especies (Ciruela, Damasco, Durazno, Nectarin, Paraguayo)
2. Crea/actualiza PMGs (programas de mejoramiento genetico)
3. Crea/actualiza variedades, vinculandolas a especie y PMG
4. Crea una posicion testblock dummy para alojar mediciones sin posicion real
5. Importa mediciones de laboratorio (brix, acidez, firmeza, peso, etc.)
6. Auto-clasifica cada medicion usando el motor de clustering

Es idempotente: verifica existencia antes de insertar y trackea duplicados.

Uso:
    cd backend
    python -m scripts.import_maestro_carozos <ruta_al_excel>

    Si no se pasa ruta, usa la ubicacion por defecto del Excel.
"""

import sys
import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

# Agregar backend/ al path para poder importar app.*
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.models.maestras import Especie, Pmg, PmgEspecie, Campo, Cuartel
from app.models.variedades import Variedad
from app.models.testblock import TestBlock, PosicionTestBlock
from app.models.laboratorio import MedicionLaboratorio, ClasificacionCluster
from app.services.clustering_service import (
    clasificar_medicion as clasificar_bandas,
    determinar_regla,
    calcular_mejillas_promedio,
    calcular_punto_debil,
)

# ---------------------------------------------------------------------------
# Constantes del Excel
# ---------------------------------------------------------------------------
# Fila real de encabezados (1-indexed)
HEADER_ROW = 3
# Fila donde comienzan los datos
DATA_START_ROW = 4
# Nombre de la hoja principal
SHEET_NAME = "CAROZOS"

# Mapeo de indice de columna (0-indexed) a nombre logico
COL = {
    "especie": 0,
    "variedad": 1,
    "pmg": 2,
    "localidad": 3,
    "temporada": 4,
    "fecha_cosecha": 5,
    "fecha_evaluacion": 6,
    "periodo_almacenaje": 7,
    "raleo": 8,
    "planta": 9,
    "perimetro": 10,
    "rendimiento": 11,
    "repeticion": 12,
    # Cubrimiento de color (porcentaje de cobertura del fruto)
    "cubr_0_30": 13,
    "cubr_30_50": 14,
    "cubr_50_75": 15,
    "cubr_75_100": 16,
    "cubr_total": 17,
    # Color de fondo
    "color_verde": 18,
    "color_crema": 19,
    "color_amarillo": 20,
    "color_full": 21,
    "color_total": 22,
    # Pulpa y mediciones por fruto
    "color_pulpa": 23,
    "fruto_num": 24,
    "peso_g": 25,
    # Firmeza (lb)
    "firmeza_punta": 26,
    "firmeza_quilla": 27,
    "firmeza_hombro": 28,
    "firmeza_mejilla1": 29,
    "firmeza_mejilla2": 30,
    # Calidad
    "brix": 31,
    "acidez": 32,
    # Desordenes postcosecha
    "pardeamiento": 33,
    "pardeamiento_total": 34,
    "traslucidez": 35,
    "traslucidez_total": 36,
    "gelificacion": 37,
    "gelificacion_total": 38,
    "harinosidad": 39,
    "harinosidad_total": 40,
    # Notas
    "observaciones": 41,
}

# Ruta por defecto del Excel
DEFAULT_EXCEL_PATH = (
    r"C:\Users\Gonzalo\Desktop\APP_SEGMENTACION_ESPECIES\files"
    r"\MAESTRO CAROZOS FINAL COMPLETO CG.xlsx"
)

# Codigo del testblock/campo/cuartel dummy para mediciones de importacion masiva
IMPORT_TB_CODE = "TB-IMPORT-CAROZOS"
IMPORT_CAMPO_CODE = "CAM-IMPORT"
IMPORT_CUARTEL_CODE = "CUA-IMPORT"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_decimal(value, precision: int = 2) -> Optional[Decimal]:
    """Convierte un valor a Decimal de forma segura."""
    if value is None:
        return None
    try:
        d = Decimal(str(value))
        return round(d, precision)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _safe_float(value) -> Optional[float]:
    """Convierte un valor a float de forma segura."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_int(value) -> Optional[int]:
    """Convierte un valor a int de forma segura."""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _normalize(s: Optional[str]) -> str:
    """Normaliza un string para comparacion: strip + lower."""
    if not s:
        return ""
    return str(s).strip().lower()


def _to_date(value) -> Optional[date]:
    """Convierte un valor a date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _generate_var_code(nombre: str, especie_code: str) -> str:
    """Genera un codigo de variedad a partir del nombre y especie."""
    clean = nombre.strip().replace(" ", "-").upper()[:15]
    return f"{especie_code}-{clean}"


# ---------------------------------------------------------------------------
# Funciones de importacion
# ---------------------------------------------------------------------------
def get_or_create_especies(db: Session) -> dict[str, Especie]:
    """Carga especies existentes y crea las que falten del Excel.

    Retorna dict normalizado nombre_lower -> Especie.
    """
    # Mapeo de nombres del Excel a datos esperados
    excel_especies = {
        "ciruela": ("CRU", "Ciruela", "Prunus domestica"),
        "damasco": ("DAM", "Damasco", "Prunus armeniaca"),
        "durazno": ("DUR", "Durazno", "Prunus persica"),
        "nectarin": ("NEC", "Nectarin", "Prunus persica var. nucipersica"),
        "paraguayo": ("PAR", "Paraguayo", "Prunus persica var. platycarpa"),
    }

    # Cargar existentes
    existentes = db.query(Especie).filter(Especie.activo == True).all()
    cache: dict[str, Especie] = {}
    for e in existentes:
        cache[_normalize(e.nombre)] = e

    created = 0
    for nombre_lower, (codigo, nombre, cientifico) in excel_especies.items():
        if nombre_lower in cache:
            continue
        # Verificar tambien por codigo
        exists_by_code = db.query(Especie).filter(Especie.codigo == codigo).first()
        if exists_by_code:
            cache[nombre_lower] = exists_by_code
            continue
        esp = Especie(
            codigo=codigo,
            nombre=nombre,
            nombre_cientifico=cientifico,
            activo=True,
            usuario_creacion="import_maestro",
        )
        db.add(esp)
        db.flush()
        cache[nombre_lower] = esp
        created += 1

    print(f"  Especies: {created} creadas, {len(cache)} total en cache")
    return cache


def get_or_create_pmgs(
    db: Session, pmg_names: set[str], especie_cache: dict[str, Especie]
) -> dict[str, Pmg]:
    """Carga PMGs existentes y crea los que falten.

    Retorna dict normalizado nombre_lower -> Pmg.
    """
    existentes = db.query(Pmg).filter(Pmg.activo == True).all()
    cache: dict[str, Pmg] = {}
    for p in existentes:
        cache[_normalize(p.nombre)] = p

    created = 0
    for nombre_raw in sorted(pmg_names):
        key = _normalize(nombre_raw)
        if not key or key in cache:
            continue
        codigo = nombre_raw.strip().replace(" ", "_").upper()[:20]
        # Verificar por codigo
        exists_by_code = db.query(Pmg).filter(Pmg.codigo == codigo).first()
        if exists_by_code:
            cache[key] = exists_by_code
            continue
        pmg = Pmg(
            codigo=codigo,
            nombre=nombre_raw.strip(),
            activo=True,
            usuario_creacion="import_maestro",
        )
        db.add(pmg)
        db.flush()
        cache[key] = pmg
        created += 1

    print(f"  PMGs: {created} creados, {len(cache)} total en cache")
    return cache


def get_or_create_variedades(
    db: Session,
    var_set: set[tuple[str, str, str]],  # (especie, variedad, pmg)
    especie_cache: dict[str, Especie],
    pmg_cache: dict[str, Pmg],
) -> dict[str, Variedad]:
    """Crea variedades que no existan. Dedup por nombre case-insensitive.

    Retorna dict normalizado nombre_lower -> Variedad.
    """
    existentes = db.query(Variedad).all()
    cache: dict[str, Variedad] = {}
    for v in existentes:
        cache[_normalize(v.nombre)] = v

    created = 0
    skipped = 0
    for especie_raw, var_raw, pmg_raw in sorted(var_set):
        var_key = _normalize(var_raw)
        if not var_key:
            continue
        if var_key in cache:
            skipped += 1
            continue

        especie_obj = especie_cache.get(_normalize(especie_raw))
        pmg_obj = pmg_cache.get(_normalize(pmg_raw))

        esp_code = especie_obj.codigo if especie_obj else "XX"
        codigo = _generate_var_code(var_raw, esp_code)

        # Verificar si el codigo ya existe (podria haber colision)
        exists_by_code = db.query(Variedad).filter(Variedad.codigo == codigo).first()
        if exists_by_code:
            cache[var_key] = exists_by_code
            skipped += 1
            continue

        var = Variedad(
            codigo=codigo,
            nombre=var_raw.strip(),
            id_especie=especie_obj.id_especie if especie_obj else None,
            id_pmg=pmg_obj.id_pmg if pmg_obj else None,
            tipo="plantada",
            estado="en_evaluacion",
            activo=True,
            usuario_creacion="import_maestro",
        )
        db.add(var)
        db.flush()
        cache[var_key] = var
        created += 1

    print(f"  Variedades: {created} creadas, {skipped} ya existian, {len(cache)} total en cache")
    return cache


def ensure_import_infrastructure(db: Session) -> PosicionTestBlock:
    """Crea campo, cuartel, testblock y posicion dummy para mediciones importadas.

    Retorna la posicion dummy que se usara como id_posicion para todas las
    mediciones importadas que no tienen posicion real.
    """
    # Campo dummy
    campo = db.query(Campo).filter(Campo.codigo == IMPORT_CAMPO_CODE).first()
    if not campo:
        campo = Campo(
            codigo=IMPORT_CAMPO_CODE,
            nombre="Import Masivo Carozos",
            ubicacion="Importacion desde Excel maestro",
            activo=True,
            usuario_creacion="import_maestro",
        )
        db.add(campo)
        db.flush()
        print("  Campo dummy creado para importacion")

    # Cuartel dummy
    cuartel = db.query(Cuartel).filter(Cuartel.codigo == IMPORT_CUARTEL_CODE).first()
    if not cuartel:
        cuartel = Cuartel(
            id_campo=campo.id_campo,
            codigo=IMPORT_CUARTEL_CODE,
            nombre="Import Masivo",
            activo=True,
        )
        db.add(cuartel)
        db.flush()
        print("  Cuartel dummy creado para importacion")

    # Testblock dummy
    tb = db.query(TestBlock).filter(TestBlock.codigo == IMPORT_TB_CODE).first()
    if not tb:
        tb = TestBlock(
            codigo=IMPORT_TB_CODE,
            nombre="Import Maestro Carozos",
            id_campo=campo.id_campo,
            id_cuartel=cuartel.id_cuartel,
            num_hileras=1,
            posiciones_por_hilera=1,
            total_posiciones=1,
            estado="activo",
            notas="Testblock virtual para mediciones importadas desde Excel Maestro Carozos",
            activo=True,
        )
        db.add(tb)
        db.flush()
        print("  TestBlock dummy creado para importacion")

    # Posicion dummy
    pos_code = f"{IMPORT_TB_CODE}-H1-P1"
    pos = db.query(PosicionTestBlock).filter(
        PosicionTestBlock.codigo_unico == pos_code
    ).first()
    if not pos:
        pos = PosicionTestBlock(
            codigo_unico=pos_code,
            id_cuartel=cuartel.id_cuartel,
            id_testblock=tb.id_testblock,
            hilera=1,
            posicion=1,
            estado="ocupada",
            observaciones="Posicion virtual para importacion masiva de mediciones",
            usuario_alta="import_maestro",
        )
        db.add(pos)
        db.flush()
        print("  Posicion dummy creada para importacion")

    return pos


def import_mediciones(
    db: Session,
    ws,
    variedad_cache: dict[str, Variedad],
    especie_cache: dict[str, Especie],
    dummy_posicion: PosicionTestBlock,
) -> dict[str, int]:
    """Importa mediciones de laboratorio desde la hoja CAROZOS.

    Para cada fila:
    - Extrae firmeza mejillas (promedio mej1, mej2)
    - Extrae firmeza punto debil (min de punta, quilla, hombro)
    - Usa el promedio de mejillas como campo 'firmeza' del modelo
    - Inserta medicion con la posicion dummy
    - Auto-clasifica usando el motor de clustering

    Retorna contadores: {created, skipped_empty, skipped_no_date, errors}
    """
    counts = {"created": 0, "skipped_empty": 0, "skipped_no_date": 0, "errors": 0, "classified": 0}
    batch_size = 500
    batch_count = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=True), DATA_START_ROW
    ):
        # Saltar filas vacias
        if not any(v is not None for v in row[:32]):
            counts["skipped_empty"] += 1
            continue

        especie_raw = row[COL["especie"]]
        variedad_raw = row[COL["variedad"]]
        pmg_raw = row[COL["pmg"]]
        temporada = row[COL["temporada"]]
        fecha_cosecha = _to_date(row[COL["fecha_cosecha"]])
        fecha_eval = _to_date(row[COL["fecha_evaluacion"]])

        # Necesitamos al menos fecha de evaluacion o cosecha
        fecha_medicion = fecha_eval or fecha_cosecha
        if fecha_medicion is None:
            counts["skipped_no_date"] += 1
            continue

        # Valores de medicion
        brix = _safe_decimal(row[COL["brix"]], 2)
        acidez = _safe_decimal(row[COL["acidez"]], 3)
        peso = _safe_decimal(row[COL["peso_g"]], 2)

        # Firmeza: 5 puntos de medicion
        f_punta = _safe_float(row[COL["firmeza_punta"]])
        f_quilla = _safe_float(row[COL["firmeza_quilla"]])
        f_hombro = _safe_float(row[COL["firmeza_hombro"]])
        f_mej1 = _safe_float(row[COL["firmeza_mejilla1"]])
        f_mej2 = _safe_float(row[COL["firmeza_mejilla2"]])

        # Firmeza promedio mejillas (campo principal del modelo)
        firmeza_mejillas = calcular_mejillas_promedio(f_mej1, f_mej2)
        firmeza = _safe_decimal(firmeza_mejillas, 1) if firmeza_mejillas else None

        # Color cubrimiento
        cubr_total = _safe_int(row[COL["cubr_total"]])

        observaciones_parts = []
        obs_raw = row[COL["observaciones"]]
        if obs_raw:
            observaciones_parts.append(str(obs_raw))

        # Agregar info de periodo de almacenaje si existe
        periodo = row[COL["periodo_almacenaje"]]
        if periodo:
            observaciones_parts.append(f"Periodo: {periodo}")

        # Info de color de fondo
        color_pulpa_raw = row[COL["color_pulpa"]]
        if color_pulpa_raw:
            observaciones_parts.append(f"Color pulpa: {color_pulpa_raw}")

        # Desordenes postcosecha
        pardeamiento = row[COL["pardeamiento"]]
        traslucidez = row[COL["traslucidez"]]
        gelificacion = row[COL["gelificacion"]]
        harinosidad = row[COL["harinosidad"]]
        desordenes = []
        if pardeamiento:
            desordenes.append(f"Pard={pardeamiento}")
        if traslucidez:
            desordenes.append(f"Trasl={traslucidez}")
        if gelificacion:
            desordenes.append(f"Gelif={gelificacion}")
        if harinosidad:
            desordenes.append(f"Harin={harinosidad}")
        if desordenes:
            observaciones_parts.append("Desordenes: " + ", ".join(desordenes))

        # Firmeza detallada en observaciones
        firmeza_parts = []
        if f_punta is not None:
            firmeza_parts.append(f"Pta={f_punta}")
        if f_quilla is not None:
            firmeza_parts.append(f"Qui={f_quilla}")
        if f_hombro is not None:
            firmeza_parts.append(f"Hmb={f_hombro}")
        if f_mej1 is not None:
            firmeza_parts.append(f"Mj1={f_mej1}")
        if f_mej2 is not None:
            firmeza_parts.append(f"Mj2={f_mej2}")
        if firmeza_parts:
            observaciones_parts.append("Firmeza: " + ", ".join(firmeza_parts))

        # Info de localidad
        localidad = row[COL["localidad"]]
        if localidad:
            observaciones_parts.append(f"Loc: {localidad}")

        observaciones = " | ".join(observaciones_parts) if observaciones_parts else None

        # Truncar observaciones si es muy largo
        if observaciones and len(observaciones) > 2000:
            observaciones = observaciones[:1997] + "..."

        try:
            medicion = MedicionLaboratorio(
                id_posicion=dummy_posicion.id_posicion,
                id_planta=None,
                temporada=str(temporada).strip() if temporada else None,
                fecha_medicion=fecha_medicion,
                fecha_cosecha=fecha_cosecha,
                brix=brix,
                acidez=acidez,
                firmeza=firmeza,
                calibre=None,  # No hay dato de calibre en el Excel
                peso=peso,
                color_pct=cubr_total,
                cracking_pct=None,
                observaciones=observaciones,
                usuario_registro="import_maestro",
            )
            db.add(medicion)
            db.flush()
            counts["created"] += 1

            # Auto-clasificar usando el motor de clustering
            especie_nombre = str(especie_raw).strip() if especie_raw else "Ciruela"
            peso_float = _safe_float(row[COL["peso_g"]])
            color_pulpa_str = str(color_pulpa_raw).strip() if color_pulpa_raw else None

            # Resolver variedad para obtener color_pulpa si no viene en la fila
            if not color_pulpa_str:
                var_key = _normalize(str(variedad_raw)) if variedad_raw else ""
                var_obj = variedad_cache.get(var_key)
                if var_obj and var_obj.color_pulpa:
                    color_pulpa_str = var_obj.color_pulpa

            regla = determinar_regla(
                especie=especie_nombre,
                peso_promedio=peso_float,
                color_pulpa=color_pulpa_str,
                fecha_evaluacion=fecha_medicion,
            )

            punto_debil = calcular_punto_debil(f_punta, f_quilla, f_hombro)

            result = clasificar_bandas(
                brix=_safe_float(row[COL["brix"]]),
                acidez=_safe_float(row[COL["acidez"]]),
                firmeza_mejillas=firmeza_mejillas if firmeza_mejillas else None,
                firmeza_punto_debil=punto_debil if punto_debil else None,
                regla=regla,
            )

            clasif = ClasificacionCluster(
                id_medicion=medicion.id_medicion,
                cluster=result["cluster"],
                banda_brix=result["banda_brix"],
                banda_firmeza=result["banda_firmeza"],
                banda_acidez=result["banda_acidez"],
                banda_calibre=result["banda_firmeza_punto"],
                score_total=Decimal(str(result["suma_bandas"])),
                metodo="import_maestro_v1",
            )
            db.add(clasif)
            counts["classified"] += 1

        except Exception as exc:
            counts["errors"] += 1
            if counts["errors"] <= 10:
                print(f"  ERROR fila {row_idx}: {exc}")
            db.rollback()
            continue

        # Commit en batches para no saturar memoria
        batch_count += 1
        if batch_count >= batch_size:
            db.commit()
            batch_count = 0
            if counts["created"] % 2000 == 0:
                print(f"  ... {counts['created']} mediciones importadas")

    # Commit final
    if batch_count > 0:
        db.commit()

    return counts


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run_import(excel_path: str) -> dict:
    """Ejecuta la importacion completa del Excel Maestro Carozos.

    Args:
        excel_path: Ruta al archivo Excel.

    Returns:
        Diccionario con resumen de la importacion.
    """
    print(f"\n{'='*70}")
    print("IMPORTACION MAESTRO CAROZOS")
    print(f"{'='*70}")
    print(f"Archivo: {excel_path}")

    if not os.path.exists(excel_path):
        print(f"ERROR: No se encontro el archivo: {excel_path}")
        return {"error": f"Archivo no encontrado: {excel_path}"}

    # --- Paso 1: Leer Excel y extraer datos unicos ---
    print("\n[1/5] Leyendo Excel...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {wb.sheetnames}")
        wb.close()
        return {"error": f"Hoja '{SHEET_NAME}' no encontrada"}

    ws = wb[SHEET_NAME]

    # Recolectar datos unicos
    pmg_names: set[str] = set()
    var_set: set[tuple[str, str, str]] = set()  # (especie, variedad, pmg)
    total_rows = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not any(v is not None for v in row[:6]):
            continue
        total_rows += 1
        especie = row[COL["especie"]]
        variedad = row[COL["variedad"]]
        pmg = row[COL["pmg"]]
        if pmg:
            pmg_names.add(str(pmg).strip())
        if variedad and especie:
            var_set.add((
                str(especie).strip(),
                str(variedad).strip(),
                str(pmg).strip() if pmg else "",
            ))

    print(f"  Filas con datos: {total_rows}")
    print(f"  PMGs unicos: {len(pmg_names)}")
    print(f"  Combinaciones especie-variedad-pmg unicas: {len(var_set)}")

    # --- Paso 2: Crear entidades maestras ---
    print("\n[2/5] Creando/verificando especies...")
    db = SessionLocal()
    try:
        especie_cache = get_or_create_especies(db)
        db.commit()

        print("\n[3/5] Creando/verificando PMGs...")
        pmg_cache = get_or_create_pmgs(db, pmg_names, especie_cache)
        db.commit()

        print("\n[4/5] Creando/verificando variedades...")
        variedad_cache = get_or_create_variedades(db, var_set, especie_cache, pmg_cache)
        db.commit()

        # Crear infraestructura dummy para mediciones
        print("\n  Verificando infraestructura de importacion (campo/testblock/posicion dummy)...")
        dummy_pos = ensure_import_infrastructure(db)
        db.commit()

        # --- Paso 3: Importar mediciones ---
        # Re-abrir el workbook sin read_only para poder iterar de nuevo
        # (read_only mode solo permite una iteracion)
        wb.close()
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]

        print(f"\n[5/5] Importando mediciones de laboratorio ({total_rows} filas)...")
        counts = import_mediciones(db, ws, variedad_cache, especie_cache, dummy_pos)

        wb.close()

        # --- Resumen ---
        print(f"\n{'='*70}")
        print("RESUMEN DE IMPORTACION")
        print(f"{'='*70}")
        print(f"  Mediciones creadas:      {counts['created']}")
        print(f"  Mediciones clasificadas: {counts['classified']}")
        print(f"  Filas sin fecha (skip):  {counts['skipped_no_date']}")
        print(f"  Filas vacias (skip):     {counts['skipped_empty']}")
        print(f"  Errores:                 {counts['errors']}")
        print(f"{'='*70}\n")

        return {
            "especies": len(especie_cache),
            "pmgs": len(pmg_cache),
            "variedades": len(variedad_cache),
            "mediciones_creadas": counts["created"],
            "mediciones_clasificadas": counts["classified"],
            "filas_sin_fecha": counts["skipped_no_date"],
            "filas_vacias": counts["skipped_empty"],
            "errores": counts["errors"],
        }

    except Exception as exc:
        db.rollback()
        print(f"\nERROR CRITICO: {exc}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL_PATH
    result = run_import(path)
    if "error" not in result:
        print("Importacion completada exitosamente.")
    else:
        print(f"Importacion fallida: {result['error']}")
        sys.exit(1)
