"""Carga masiva de datos desde archivos Excel en data/ hacia la base de datos.

Lee los siguientes archivos (en orden de dependencias FK):
  1. BBDD Programas.xlsx       -> pmg
  2. Seed de Especies           -> especies (7 especies fijas)
  3. BBDD Variedades.xlsx      -> variedades (por hoja = 1 PMG)
  4. BBDD Portainjertos.xlsx   -> portainjertos
  5. BASE MAESTROS GLOBAL.xlsx -> campos, testblocks, posiciones_testblock, plantas
  6. MAESTRO CEREZAS.xlsx      -> mediciones_laboratorio

El script es idempotente: verifica existencia antes de insertar (get_or_create).
Re-ejecutar no duplica datos.

Uso:
    cd backend
    python -m scripts.load_data
"""

import sys
import os
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Path setup: agregar backend/ al sys.path y establecer cwd
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.chdir(os.path.join(os.path.dirname(__file__), ".."))

from dotenv import load_dotenv

load_dotenv(".env")

import openpyxl
from sqlalchemy.orm import Session

from app.core.database import SessionLocal, engine
from sqlmodel import SQLModel

from app.models.maestras import Especie, Portainjerto, Pmg, PmgEspecie, Campo, Cuartel
from app.models.variedades import Variedad
from app.models.inventario import InventarioVivero
from app.models.testblock import TestBlock, PosicionTestBlock, Planta
from app.models.laboratorio import MedicionLaboratorio

# ---------------------------------------------------------------------------
# Rutas a archivos Excel
# ---------------------------------------------------------------------------
DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"

FILE_PROGRAMAS = DATA_DIR / "BBDD Programas.xlsx"
FILE_VARIEDADES = DATA_DIR / "BBDD Variedades.xlsx"
FILE_PORTAINJERTOS = DATA_DIR / "BBDD Portainjertos.xlsx"
FILE_MAESTROS_GLOBAL = DATA_DIR / "BASE MAESTROS GLOBAL.xlsx"
FILE_MAESTRO_CEREZAS = DATA_DIR / "MAESTRO CEREZAS.xlsx"

USUARIO_CARGA = "load_data"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_str(value) -> Optional[str]:
    """Convierte a str limpio o None."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "None", "nan", "#VALUE!", "N/A", "-"):
        return None
    # Limpiar caracteres no-breaking space
    s = s.replace("\xa0", "").strip()
    return s if s else None


def _safe_decimal(value, precision: int = 2) -> Optional[Decimal]:
    """Convierte un valor a Decimal de forma segura."""
    if value is None:
        return None
    try:
        d = Decimal(str(value))
        return round(d, precision)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _safe_int(value) -> Optional[int]:
    """Convierte un valor a int de forma segura."""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _safe_float(value) -> Optional[float]:
    """Convierte un valor a float de forma segura."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _to_date(value) -> Optional[date]:
    """Convierte un valor a date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _normalize(s) -> str:
    """Normaliza a lowercase para comparaciones."""
    if s is None:
        return ""
    return str(s).strip().lower()


def _make_code(prefix: str, name: str, max_len: int = 20) -> str:
    """Genera un codigo a partir de un nombre, con prefijo."""
    clean = name.strip().replace(" ", "-").upper()
    code = f"{prefix}-{clean}"
    return code[:max_len]


# ---------------------------------------------------------------------------
# Step 1: PMGs desde BBDD Programas.xlsx
# ---------------------------------------------------------------------------
def load_pmgs(db: Session) -> dict[str, Pmg]:
    """Carga PMGs desde BBDD Programas.xlsx (hoja Hoja1).

    Columnas: Nombre PMG, Pais, Ciudad, Nombre contacto Chile, Email, Telefono,
              Vivero 1..3, Especie 1..5
    """
    print("\n[Step 1] Cargando PMGs desde BBDD Programas.xlsx...")

    # Cache existentes
    existentes = db.query(Pmg).all()
    cache: dict[str, Pmg] = {}
    for p in existentes:
        cache[_normalize(p.nombre)] = p

    wb = openpyxl.load_workbook(str(FILE_PROGRAMAS), read_only=True, data_only=True)
    ws = wb["Hoja1"]

    created = 0
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre_raw = _safe_str(row[0])
        if not nombre_raw:
            continue

        key = _normalize(nombre_raw)
        pais = _safe_str(row[1])
        ciudad = _safe_str(row[2])
        contacto_nombre = _safe_str(row[3])
        email = _safe_str(row[4])
        telefono = _safe_str(row[5])

        # Viveros como texto concatenado
        viveros = []
        for i in range(6, 9):
            v = _safe_str(row[i]) if len(row) > i else None
            if v:
                viveros.append(v)
        viveros_str = ", ".join(viveros) if viveros else None

        if key in cache:
            # Upsert: actualizar campos vacios
            pmg = cache[key]
            changed = False
            if not pmg.pais and pais:
                pmg.pais = pais
                changed = True
            if not pmg.ciudad and ciudad:
                pmg.ciudad = ciudad
                changed = True
            if not pmg.contacto_nombre and contacto_nombre:
                pmg.contacto_nombre = contacto_nombre
                changed = True
            if not pmg.contacto_email and email:
                pmg.contacto_email = email
                changed = True
            if not pmg.contacto_telefono and telefono:
                pmg.contacto_telefono = telefono
                changed = True
            if not pmg.viveros_chile and viveros_str:
                pmg.viveros_chile = viveros_str
                changed = True
            if changed:
                pmg.fecha_modificacion = datetime.utcnow()
                pmg.usuario_modificacion = USUARIO_CARGA
                updated += 1
            continue

        # Generar codigo unico
        codigo = nombre_raw.strip().replace(" ", "_").upper()[:20]
        # Verificar colision de codigo
        exists_by_code = db.query(Pmg).filter(Pmg.codigo == codigo).first()
        if exists_by_code:
            cache[key] = exists_by_code
            continue

        pmg = Pmg(
            codigo=codigo,
            nombre=nombre_raw,
            pais=pais,
            ciudad=ciudad,
            contacto_nombre=contacto_nombre,
            contacto_email=email,
            contacto_telefono=telefono,
            viveros_chile=viveros_str,
            activo=True,
            usuario_creacion=USUARIO_CARGA,
        )
        db.add(pmg)
        db.flush()
        cache[key] = pmg
        created += 1

    wb.close()
    db.commit()
    print(f"  Step 1: Loaded {created} new PMGs, updated {updated} ({len(cache)} total)")
    return cache


# ---------------------------------------------------------------------------
# Step 2: Seed Especies
# ---------------------------------------------------------------------------
SEED_ESPECIES = [
    ("CER", "Cerezo", "Prunus avium"),
    ("CIR", "Ciruela", "Prunus domestica"),
    ("DUR", "Durazno", "Prunus persica"),
    ("NEC", "Nectarin", "Prunus persica var. nucipersica"),
    ("PAR", "Paraguayo", "Prunus persica var. platycarpa"),
    ("PLA", "Platerina", "Prunus persica var. platycarpa"),
    ("KIW", "Kiwi", "Actinidia deliciosa"),
]

# Mapeo de nombres alternativos del Excel a nombre canonico
ESPECIE_ALIASES = {
    "cereza": "cerezo",
    "cherry": "cerezo",
}


def load_especies(db: Session) -> dict[str, Especie]:
    """Inserta las 7 especies seed si no existen.

    Retorna cache normalizado nombre_lower -> Especie.
    """
    print("\n[Step 2] Seeding Especies...")

    existentes = db.query(Especie).all()
    cache: dict[str, Especie] = {}
    for e in existentes:
        cache[_normalize(e.nombre)] = e

    created = 0
    for codigo, nombre, cientifico in SEED_ESPECIES:
        key = _normalize(nombre)
        if key in cache:
            continue
        # Check por codigo
        exists_by_code = db.query(Especie).filter(Especie.codigo == codigo).first()
        if exists_by_code:
            cache[key] = exists_by_code
            continue

        esp = Especie(
            codigo=codigo,
            nombre=nombre,
            nombre_cientifico=cientifico,
            activo=True,
            usuario_creacion=USUARIO_CARGA,
        )
        db.add(esp)
        db.flush()
        cache[key] = esp
        created += 1

    db.commit()

    # Agregar aliases al cache
    for alias, canonical in ESPECIE_ALIASES.items():
        if canonical in cache and alias not in cache:
            cache[alias] = cache[canonical]

    print(f"  Step 2: Loaded {created} new Especies ({len(cache)} total in cache)")
    return cache


def _resolve_especie(
    especie_raw: str, cache: dict[str, Especie]
) -> Optional[Especie]:
    """Resuelve una especie por nombre, con fallback a aliases."""
    key = _normalize(especie_raw)
    if key in cache:
        return cache[key]
    # Intentar alias
    alias_key = ESPECIE_ALIASES.get(key)
    if alias_key and alias_key in cache:
        return cache[alias_key]
    # Busqueda parcial
    for k, v in cache.items():
        if key in k or k in key:
            return v
    return None


# ---------------------------------------------------------------------------
# Step 3: Variedades desde BBDD Variedades.xlsx
# ---------------------------------------------------------------------------
def load_variedades(
    db: Session,
    pmg_cache: dict[str, Pmg],
    especie_cache: dict[str, Especie],
) -> dict[str, Variedad]:
    """Carga variedades desde BBDD Variedades.xlsx.

    Cada hoja = 1 PMG. Columnas por hoja:
    ESPECIE, ORIGEN, NOMBRE, CODIGO COMERCIAL|CODIGO TEST, FECHA COSECHA,
    IMAGEN, COLOR, FERTILIDAD, ALELOS, POLINIZANTES, PORTAINJERTOS,
    FAMILIA GENETICA INFO, RECOMENDACIONES PROGRAMA
    """
    print("\n[Step 3] Cargando Variedades desde BBDD Variedades.xlsx...")

    # Cache existentes
    existentes = db.query(Variedad).all()
    cache: dict[str, Variedad] = {}
    for v in existentes:
        # Key = nombre_lower + especie_id para desambiguar variedades con mismo nombre
        cache[f"{_normalize(v.nombre)}|{v.id_especie}"] = v
        # Tambien cacheamos solo por nombre para busquedas rapidas
        cache[_normalize(v.nombre)] = v

    wb = openpyxl.load_workbook(str(FILE_VARIEDADES), read_only=True, data_only=True)

    created = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Leer encabezados para encontrar columna de codigo
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [_safe_str(c) for c in row]
            break

        if not headers:
            continue

        # Encontrar indice de columna de codigo (puede ser CODIGO COMERCIAL o CODIGO TEST)
        codigo_col = None
        for i, h in enumerate(headers):
            if h and "CODIGO" in h.upper():
                codigo_col = i
                break

        # Resolver PMG por nombre de hoja
        pmg_key = _normalize(sheet_name)
        pmg_obj = pmg_cache.get(pmg_key)
        if not pmg_obj:
            # Busqueda parcial
            for k, v in pmg_cache.items():
                if pmg_key in k or k in pmg_key:
                    pmg_obj = v
                    break

        sheet_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            especie_raw = _safe_str(row[0])
            nombre_raw = _safe_str(row[2]) if len(row) > 2 else None

            if not nombre_raw:
                continue

            # Convertir nombre a str (puede ser numerico, ej: 2219)
            nombre = str(nombre_raw).strip()
            if not nombre:
                continue

            # Resolver especie
            especie_obj = _resolve_especie(especie_raw, especie_cache) if especie_raw else None
            especie_id = especie_obj.id_especie if especie_obj else None

            # Check si ya existe por nombre + especie
            compound_key = f"{_normalize(nombre)}|{especie_id}"
            if compound_key in cache:
                skipped += 1
                continue

            # Check solo por nombre (sin especie)
            name_key = _normalize(nombre)
            if name_key in cache:
                skipped += 1
                continue

            # Extraer codigo del Excel
            codigo_raw = None
            if codigo_col is not None and len(row) > codigo_col:
                codigo_raw = _safe_str(row[codigo_col])

            if codigo_raw:
                codigo = str(codigo_raw).strip()[:30]
            else:
                esp_code = especie_obj.codigo if especie_obj else "XX"
                codigo = _make_code(esp_code, nombre, 30)

            # Verificar colision de codigo
            exists_by_code = db.query(Variedad).filter(Variedad.codigo == codigo).first()
            if exists_by_code:
                cache[compound_key] = exists_by_code
                cache[name_key] = exists_by_code
                skipped += 1
                continue

            # Extraer campos adicionales
            origen = _safe_str(row[1]) if len(row) > 1 else None
            color = _safe_str(row[6]) if len(row) > 6 else None
            fertilidad = _safe_str(row[7]) if len(row) > 7 else None
            alelos = _safe_str(row[8]) if len(row) > 8 else None
            polinizantes = _safe_str(row[9]) if len(row) > 9 else None
            portainjertos_rec = _safe_str(row[10]) if len(row) > 10 else None
            familia_genetica = _safe_str(row[11]) if len(row) > 11 else None
            recomendaciones = _safe_str(row[12]) if len(row) > 12 else None
            fecha_cosecha_ref = _safe_str(row[4]) if len(row) > 4 else None

            # Autofertil?
            auto_fertil = None
            if fertilidad:
                fert_lower = fertilidad.lower()
                if "autofertil" in fert_lower or "autof" in fert_lower:
                    auto_fertil = "incompatible" not in fert_lower

            var = Variedad(
                codigo=codigo,
                nombre=nombre,
                id_especie=especie_id,
                id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                origen=origen,
                color_fruto=color,
                fertilidad=fertilidad,
                alelos=alelos,
                polinizantes=polinizantes,
                portainjertos_recomendados=portainjertos_rec,
                familia_genetica=familia_genetica,
                recomendaciones=recomendaciones,
                fecha_cosecha_ref=fecha_cosecha_ref,
                auto_fertil=auto_fertil,
                tipo="evaluacion",
                estado="en_evaluacion",
                activo=True,
                usuario_creacion=USUARIO_CARGA,
            )
            db.add(var)
            db.flush()
            cache[compound_key] = var
            cache[name_key] = var
            created += 1
            sheet_created += 1

        if sheet_created > 0:
            print(f"    Sheet '{sheet_name}': {sheet_created} variedades creadas")

    wb.close()
    db.commit()
    print(f"  Step 3: Loaded {created} new Variedades, skipped {skipped}")
    return cache


# ---------------------------------------------------------------------------
# Step 4: Portainjertos desde BBDD Portainjertos.xlsx
# ---------------------------------------------------------------------------
def load_portainjertos(db: Session) -> dict[str, Portainjerto]:
    """Carga portainjertos desde BBDD Portainjertos.xlsx (hoja V1).

    Columnas: PORTAINJERTO, ORIGEN, CULTIVO INTERES, Compatibilidad,
              Caracteristica, Sensibilidad/Dificultades, ...
    """
    print("\n[Step 4] Cargando Portainjertos desde BBDD Portainjertos.xlsx...")

    existentes = db.query(Portainjerto).all()
    cache: dict[str, Portainjerto] = {}
    for p in existentes:
        cache[_normalize(p.nombre)] = p

    wb = openpyxl.load_workbook(str(FILE_PORTAINJERTOS), read_only=True, data_only=True)
    ws = wb["V1"]

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre_raw = _safe_str(row[0])
        if not nombre_raw:
            continue

        key = _normalize(nombre_raw)
        if key in cache:
            continue

        origen = _safe_str(row[1]) if len(row) > 1 else None
        especie_interes = _safe_str(row[2]) if len(row) > 2 else None
        compatibilidad = _safe_str(row[3]) if len(row) > 3 else None
        caracteristicas = _safe_str(row[4]) if len(row) > 4 else None
        sensibilidad = _safe_str(row[5]) if len(row) > 5 else None
        propagacion = _safe_str(row[7]) if len(row) > 7 else None
        obtentor = _safe_str(row[9]) if len(row) > 9 else None

        codigo = nombre_raw.strip().replace(" ", "-").upper()[:20]

        # Verificar colision de codigo
        exists_by_code = db.query(Portainjerto).filter(Portainjerto.codigo == codigo).first()
        if exists_by_code:
            cache[key] = exists_by_code
            continue

        pi = Portainjerto(
            codigo=codigo,
            nombre=nombre_raw.strip(),
            origen=origen,
            especie=especie_interes,
            compatibilidad=compatibilidad,
            caracteristicas=caracteristicas,
            sensibilidad=sensibilidad,
            propagacion=propagacion,
            obtentor=obtentor,
            activo=True,
            usuario_creacion=USUARIO_CARGA,
        )
        db.add(pi)
        db.flush()
        cache[key] = pi
        created += 1

    wb.close()
    db.commit()
    print(f"  Step 4: Loaded {created} new Portainjertos ({len(cache)} total)")
    return cache


def _get_or_create_portainjerto(
    db: Session, nombre: str, cache: dict[str, Portainjerto]
) -> Portainjerto:
    """Obtiene o crea un portainjerto por nombre."""
    key = _normalize(nombre)
    if key in cache:
        return cache[key]

    codigo = nombre.strip().replace(" ", "-").upper()[:20]
    exists = db.query(Portainjerto).filter(Portainjerto.codigo == codigo).first()
    if exists:
        cache[key] = exists
        return exists

    pi = Portainjerto(
        codigo=codigo,
        nombre=nombre.strip(),
        activo=True,
        usuario_creacion=USUARIO_CARGA,
    )
    db.add(pi)
    db.flush()
    cache[key] = pi
    return pi


# ---------------------------------------------------------------------------
# Step 5: Campos desde BASE MAESTROS GLOBAL + MAESTRO CEREZAS
# ---------------------------------------------------------------------------
def load_campos(db: Session) -> dict[str, Campo]:
    """Extrae campos unicos de ambos archivos maestros e inserta los que no existan."""
    print("\n[Step 5] Cargando Campos desde BASE MAESTROS GLOBAL + MAESTRO CEREZAS...")

    existentes = db.query(Campo).all()
    cache: dict[str, Campo] = {}
    for c in existentes:
        cache[_normalize(c.nombre)] = c

    campos_set: set[str] = set()

    # Desde BASE MAESTROS GLOBAL (col 0 = Campo)
    wb1 = openpyxl.load_workbook(str(FILE_MAESTROS_GLOBAL), read_only=True, data_only=True)
    ws1 = wb1["BBDD"]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        campo = _safe_str(row[0])
        if campo:
            campos_set.add(campo.strip())
    wb1.close()

    # Desde MAESTRO CEREZAS (col 2 = Campo)
    wb2 = openpyxl.load_workbook(str(FILE_MAESTRO_CEREZAS), read_only=True, data_only=True)
    ws2 = wb2["Ev. Cosecha Extenso"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        campo = _safe_str(row[2])
        if campo:
            campos_set.add(campo.strip())
    wb2.close()

    created = 0
    for nombre in sorted(campos_set):
        key = _normalize(nombre)
        if key in cache:
            continue

        codigo = nombre.strip().replace(" ", "-").upper()[:20]
        exists = db.query(Campo).filter(Campo.codigo == codigo).first()
        if exists:
            cache[key] = exists
            continue

        campo = Campo(
            codigo=codigo,
            nombre=nombre,
            activo=True,
            usuario_creacion=USUARIO_CARGA,
        )
        db.add(campo)
        db.flush()
        cache[key] = campo
        created += 1

    db.commit()
    print(f"  Step 5: Loaded {created} new Campos ({len(cache)} total)")
    return cache


def _get_or_create_campo(db: Session, nombre: str, cache: dict[str, Campo]) -> Campo:
    """Obtiene o crea un campo por nombre."""
    key = _normalize(nombre)
    if key in cache:
        return cache[key]

    codigo = nombre.strip().replace(" ", "-").upper()[:20]
    exists = db.query(Campo).filter(Campo.codigo == codigo).first()
    if exists:
        cache[key] = exists
        return exists

    campo = Campo(
        codigo=codigo,
        nombre=nombre.strip(),
        activo=True,
        usuario_creacion=USUARIO_CARGA,
    )
    db.add(campo)
    db.flush()
    cache[key] = campo
    return campo


# ---------------------------------------------------------------------------
# Step 6: TestBlocks + Posiciones desde BASE MAESTROS GLOBAL.xlsx
# ---------------------------------------------------------------------------
def load_testblocks_global(
    db: Session,
    campo_cache: dict[str, Campo],
    especie_cache: dict[str, Especie],
    pmg_cache: dict[str, Pmg],
    pi_cache: dict[str, Portainjerto],
    var_cache: dict[str, Variedad],
) -> None:
    """Carga testblocks, cuarteles, posiciones y plantas desde BASE MAESTROS GLOBAL.xlsx.

    Columnas:
    0=Campo, 1=Cuartel, 2=Especie, 3=PMG, 4=Codigo Variedad, 5=Variedad Comercial,
    6=Hilera, 7=Posicion, 8=N Planta Variedad, 9=Color cubrimiento,
    10=Color pulpa, 11=Ano plantacion, 12=Portainjerto, 13=Tipo Patron,
    14=Marco plantacion, 15=Ano injertacion, 16=Injertacion, 17=Conduccion,
    18=Condicion, 19=Observacion
    """
    print("\n[Step 6] Cargando TestBlocks + Posiciones desde BASE MAESTROS GLOBAL.xlsx...")

    wb = openpyxl.load_workbook(str(FILE_MAESTROS_GLOBAL), read_only=True, data_only=True)
    ws = wb["BBDD"]

    # Cache de cuarteles y testblocks
    cuartel_cache: dict[str, Cuartel] = {}
    for c in db.query(Cuartel).all():
        cuartel_cache[_normalize(c.codigo)] = c

    tb_cache: dict[str, TestBlock] = {}
    for tb in db.query(TestBlock).all():
        tb_cache[_normalize(tb.codigo)] = tb

    pos_cache: set[str] = set()
    for p in db.query(PosicionTestBlock).all():
        pos_cache.add(p.codigo_unico)

    created_tb = 0
    created_cuartel = 0
    created_pos = 0
    created_planta = 0
    errors = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            campo_raw = _safe_str(row[0])
            cuartel_raw = _safe_str(row[1])
            especie_raw = _safe_str(row[2])
            pmg_raw = _safe_str(row[3])
            codigo_var = _safe_str(row[4])
            nombre_var = _safe_str(row[5])
            hilera = _safe_int(row[6])
            posicion = _safe_int(row[7])
            color_cubrimiento = _safe_str(row[9])
            color_pulpa = _safe_str(row[10])
            ano_plantacion = _safe_int(row[11])
            pi_raw = _safe_str(row[12])
            tipo_patron = _safe_str(row[13])
            marco_plantacion = _safe_str(row[14])
            ano_injertacion = _safe_int(row[15])
            injertacion = _safe_str(row[16])
            conduccion = _safe_str(row[17])
            condicion = _safe_str(row[18])
            observacion = _safe_str(row[19])

            if not campo_raw or hilera is None or posicion is None:
                continue

            # Resolver campo
            campo_obj = _get_or_create_campo(db, campo_raw, campo_cache)

            # Resolver cuartel (usando el valor original como codigo)
            cuartel_code = str(cuartel_raw).strip() if cuartel_raw else "SN"
            cuartel_key = _normalize(cuartel_code)
            if cuartel_key not in cuartel_cache:
                exists = db.query(Cuartel).filter(Cuartel.codigo == cuartel_code).first()
                if exists:
                    cuartel_cache[cuartel_key] = exists
                else:
                    cuartel_obj = Cuartel(
                        id_campo=campo_obj.id_campo,
                        codigo=cuartel_code,
                        nombre=f"Cuartel {cuartel_code}",
                        activo=True,
                        usuario_creacion=USUARIO_CARGA,
                    )
                    db.add(cuartel_obj)
                    db.flush()
                    cuartel_cache[cuartel_key] = cuartel_obj
                    created_cuartel += 1

            cuartel_obj = cuartel_cache[cuartel_key]

            # Resolver testblock (1 por campo+cuartel)
            tb_code = f"TB-{_normalize(campo_raw).upper()[:8]}-{cuartel_code}"[:20]
            tb_key = _normalize(tb_code)
            if tb_key not in tb_cache:
                exists = db.query(TestBlock).filter(TestBlock.codigo == tb_code).first()
                if exists:
                    tb_cache[tb_key] = exists
                else:
                    tb_obj = TestBlock(
                        codigo=tb_code,
                        nombre=f"TestBlock {campo_raw} - {cuartel_code}",
                        id_campo=campo_obj.id_campo,
                        id_cuartel=cuartel_obj.id_cuartel,
                        estado="activo",
                        activo=True,
                    )
                    db.add(tb_obj)
                    db.flush()
                    tb_cache[tb_key] = tb_obj
                    created_tb += 1

            tb_obj = tb_cache[tb_key]

            # Resolver especie
            especie_obj = _resolve_especie(especie_raw, especie_cache) if especie_raw else None

            # Resolver PMG
            pmg_obj = pmg_cache.get(_normalize(pmg_raw)) if pmg_raw else None
            if pmg_raw and not pmg_obj:
                # Busqueda parcial
                pmg_key_search = _normalize(pmg_raw)
                for k, v in pmg_cache.items():
                    if pmg_key_search in k or k in pmg_key_search:
                        pmg_obj = v
                        break

            # Resolver variedad
            var_obj = None
            if codigo_var:
                var_key = _normalize(codigo_var)
                var_obj = var_cache.get(var_key)
            if not var_obj and nombre_var:
                var_key = _normalize(nombre_var)
                var_obj = var_cache.get(var_key)

            # Si la variedad no existe, crearla
            if not var_obj and (codigo_var or nombre_var):
                var_nombre = (nombre_var or codigo_var).strip()
                var_codigo = (codigo_var or nombre_var).strip()[:30]
                exists = db.query(Variedad).filter(Variedad.codigo == var_codigo).first()
                if exists:
                    var_obj = exists
                else:
                    var_obj = Variedad(
                        codigo=var_codigo,
                        nombre=var_nombre,
                        id_especie=especie_obj.id_especie if especie_obj else None,
                        id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                        tipo="plantada",
                        estado="en_evaluacion",
                        activo=True,
                        usuario_creacion=USUARIO_CARGA,
                    )
                    db.add(var_obj)
                    db.flush()

                var_cache[_normalize(var_nombre)] = var_obj
                if codigo_var:
                    var_cache[_normalize(codigo_var)] = var_obj

            # Resolver portainjerto
            pi_obj = None
            if pi_raw:
                pi_obj = _get_or_create_portainjerto(db, pi_raw, pi_cache)

            # Crear posicion
            pos_code = f"{tb_code}-H{hilera}-P{posicion}"
            if pos_code in pos_cache:
                continue

            pos_obj = PosicionTestBlock(
                codigo_unico=pos_code,
                id_cuartel=cuartel_obj.id_cuartel,
                id_testblock=tb_obj.id_testblock,
                id_variedad=var_obj.id_variedad if var_obj else None,
                id_portainjerto=pi_obj.id_portainjerto if pi_obj else None,
                id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                hilera=hilera,
                posicion=posicion,
                estado="alta" if condicion and "EVALUACI" in condicion.upper() else "alta",
                conduccion=conduccion,
                marco_plantacion=marco_plantacion,
                observaciones=observacion,
                usuario_alta=USUARIO_CARGA,
            )
            db.add(pos_obj)
            db.flush()
            pos_cache.add(pos_code)
            created_pos += 1

            # Crear planta asociada
            planta_code = f"{pos_code}-PL"
            exists_planta = db.query(Planta).filter(Planta.codigo == planta_code).first()
            if not exists_planta:
                planta = Planta(
                    codigo=planta_code,
                    id_posicion=pos_obj.id_posicion,
                    id_variedad=var_obj.id_variedad if var_obj else None,
                    id_portainjerto=pi_obj.id_portainjerto if pi_obj else None,
                    id_especie=especie_obj.id_especie if especie_obj else None,
                    id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                    condicion=condicion if condicion else "EN_EVALUACION",
                    activa=True,
                    ano_plantacion=ano_plantacion,
                    ano_injertacion=ano_injertacion,
                    metodo_injertacion=injertacion,
                    tipo_patron=tipo_patron,
                    conduccion=conduccion,
                    marco_plantacion=marco_plantacion,
                    color_cubrimiento=color_cubrimiento,
                    color_pulpa=color_pulpa,
                    usuario_creacion=USUARIO_CARGA,
                )
                db.add(planta)
                created_planta += 1

        except Exception as exc:
            errors += 1
            if errors <= 10:
                print(f"    ERROR row: {exc}")
            continue

    db.flush()
    wb.close()
    db.commit()

    print(f"  Step 6: Loaded {created_cuartel} Cuarteles, {created_tb} TestBlocks, "
          f"{created_pos} Posiciones, {created_planta} Plantas ({errors} errors)")


# ---------------------------------------------------------------------------
# Step 7: Mediciones desde MAESTRO CEREZAS.xlsx
# ---------------------------------------------------------------------------
def load_mediciones_cerezas(
    db: Session,
    campo_cache: dict[str, Campo],
    especie_cache: dict[str, Especie],
    pmg_cache: dict[str, Pmg],
    pi_cache: dict[str, Portainjerto],
    var_cache: dict[str, Variedad],
) -> None:
    """Carga mediciones de laboratorio desde MAESTRO CEREZAS.xlsx (Ev. Cosecha Extenso).

    Columnas clave (0-indexed):
      0=Especie, 1=Localidad, 2=Campo, 3=Cuartel, 4=Origen, 5=Temporada,
      6=PMG, 7=Variedad, 8=Portainjerto, 9=Color,
      18=Fecha cosecha, 19=N fruto, 20=Peso (g), 21=Solidos solubles (%),
      22=Largo pedicelo, 23=Arraigamiento pedicelar,
      24-31=Calibre (PC,L,XL,J,2J,3J,4J,5J), 33-39=Color dist, 41-44=Color cubr %,
      46=Firmeza (UD), 47=SS (%), 48=Acidez (%),
      49=Rendimiento (g), 50=Observaciones cosecha,
      53=Periodo Almacenaje (dias)
    """
    print("\n[Step 7] Cargando Mediciones desde MAESTRO CEREZAS.xlsx...")

    wb = openpyxl.load_workbook(str(FILE_MAESTRO_CEREZAS), read_only=True, data_only=True)
    ws = wb["Ev. Cosecha Extenso"]

    # Para idempotencia, necesitamos un criterio de unicidad.
    # Usaremos una posicion dummy por campo+cuartel+variedad+pi+temporada
    cuartel_cache: dict[str, Cuartel] = {}
    for c in db.query(Cuartel).all():
        cuartel_cache[_normalize(c.codigo)] = c

    tb_cache: dict[str, TestBlock] = {}
    for tb in db.query(TestBlock).all():
        tb_cache[_normalize(tb.codigo)] = tb

    # Verificar si ya hay mediciones importadas por este script
    existing_count = db.query(MedicionLaboratorio).filter(
        MedicionLaboratorio.usuario_registro == USUARIO_CARGA
    ).count()

    if existing_count > 0:
        print(f"  WARNING: Ya existen {existing_count} mediciones creadas por {USUARIO_CARGA}")
        print(f"  Saltando Step 7 para evitar duplicados. Elimine las mediciones existentes")
        print(f"  con usuario_registro='{USUARIO_CARGA}' si desea re-importar.")
        wb.close()
        return

    created = 0
    skipped_no_date = 0
    skipped_empty = 0
    errors = 0
    batch_count = 0
    BATCH_SIZE = 500

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Saltar filas completamente vacias
            if not any(v is not None for v in row[:22]):
                skipped_empty += 1
                continue

            especie_raw = _safe_str(row[0])
            campo_raw = _safe_str(row[2])
            cuartel_raw = _safe_str(row[3])
            temporada = _safe_str(row[5])
            pmg_raw = _safe_str(row[6])
            variedad_raw = _safe_str(row[7])
            pi_raw = _safe_str(row[8])
            color_raw = _safe_str(row[9])

            fecha_cosecha = _to_date(row[18])
            n_fruto = _safe_int(row[19])
            peso = _safe_decimal(row[20], 2)
            brix = _safe_decimal(row[21], 2)

            # Firmeza y acidez en columnas de resumen
            firmeza = _safe_decimal(row[46], 1)
            acidez = _safe_decimal(row[48], 3)

            # Rendimiento
            rendimiento = _safe_decimal(row[49], 2)

            # Observaciones
            observaciones = _safe_str(row[50]) if len(row) > 50 else None

            # Periodo almacenaje
            periodo_almacenaje = _safe_int(row[53]) if len(row) > 53 else None

            # Color cubrimiento (porcentajes)
            color_0_25 = _safe_int(row[41]) if len(row) > 41 else None
            color_25_50 = _safe_int(row[42]) if len(row) > 42 else None
            color_50_75 = _safe_int(row[43]) if len(row) > 43 else None
            color_75_100 = _safe_int(row[44]) if len(row) > 44 else None

            # Necesitamos al menos una fecha
            fecha_medicion = fecha_cosecha
            if fecha_medicion is None:
                skipped_no_date += 1
                continue

            # Resolver especie
            especie_obj = _resolve_especie(especie_raw, especie_cache) if especie_raw else None

            # Resolver campo
            campo_obj = None
            if campo_raw:
                campo_obj = _get_or_create_campo(db, campo_raw, campo_cache)

            # Resolver PMG
            pmg_obj = pmg_cache.get(_normalize(pmg_raw)) if pmg_raw else None
            if pmg_raw and not pmg_obj:
                # Crear PMG on the fly
                pmg_key = _normalize(pmg_raw)
                pmg_codigo = pmg_raw.strip().replace(" ", "_").upper()[:20]
                exists = db.query(Pmg).filter(Pmg.codigo == pmg_codigo).first()
                if exists:
                    pmg_obj = exists
                else:
                    pmg_obj = Pmg(
                        codigo=pmg_codigo,
                        nombre=pmg_raw.strip(),
                        activo=True,
                        usuario_creacion=USUARIO_CARGA,
                    )
                    db.add(pmg_obj)
                    db.flush()
                pmg_cache[pmg_key] = pmg_obj

            # Resolver variedad
            var_obj = None
            if variedad_raw:
                var_key = _normalize(str(variedad_raw))
                var_obj = var_cache.get(var_key)
                if not var_obj:
                    # Crear variedad on the fly
                    var_nombre = str(variedad_raw).strip()
                    esp_code = especie_obj.codigo if especie_obj else "XX"
                    var_codigo = _make_code(esp_code, var_nombre, 30)
                    exists = db.query(Variedad).filter(Variedad.codigo == var_codigo).first()
                    if exists:
                        var_obj = exists
                    else:
                        var_obj = Variedad(
                            codigo=var_codigo,
                            nombre=var_nombre,
                            id_especie=especie_obj.id_especie if especie_obj else None,
                            id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                            tipo="plantada",
                            estado="en_evaluacion",
                            activo=True,
                            usuario_creacion=USUARIO_CARGA,
                        )
                        db.add(var_obj)
                        db.flush()
                    var_cache[var_key] = var_obj

            # Resolver portainjerto
            pi_obj = None
            if pi_raw:
                pi_obj = _get_or_create_portainjerto(db, pi_raw, pi_cache)

            # Color cubrimiento total (pct)
            color_pct = None
            if color_75_100 is not None:
                color_pct = color_75_100

            # Crear medicion
            medicion = MedicionLaboratorio(
                id_posicion=None,
                id_planta=None,
                temporada=temporada,
                fecha_medicion=fecha_medicion,
                fecha_cosecha=fecha_cosecha,
                brix=brix,
                acidez=acidez,
                firmeza=firmeza,
                calibre=None,  # Cerezas usan categorias (PC,L,XL,J...), no mm directo
                peso=peso,
                color_pct=color_pct,
                n_muestra=n_fruto,
                periodo_almacenaje=periodo_almacenaje,
                rendimiento=rendimiento,
                observaciones=observaciones,
                usuario_registro=USUARIO_CARGA,
                # FKs de contexto
                id_campo=campo_obj.id_campo if campo_obj else None,
                id_variedad=var_obj.id_variedad if var_obj else None,
                id_especie=especie_obj.id_especie if especie_obj else None,
                id_portainjerto=pi_obj.id_portainjerto if pi_obj else None,
                # Color cubrimiento detallado
                color_0_30=color_0_25,  # Mapeo aproximado 0-25 -> 0-30
                color_50_75=color_50_75,
                color_75_100=color_75_100,
            )
            db.add(medicion)
            created += 1

            # Commit en batches
            batch_count += 1
            if batch_count >= BATCH_SIZE:
                db.flush()
                db.commit()
                batch_count = 0
                if created % 2000 == 0:
                    print(f"    ... {created} mediciones importadas")

        except Exception as exc:
            errors += 1
            if errors <= 10:
                print(f"    ERROR row {row_idx}: {exc}")
            db.rollback()
            batch_count = 0
            continue

    # Commit final
    if batch_count > 0:
        db.flush()
        db.commit()

    wb.close()
    print(f"  Step 7: Loaded {created} mediciones, "
          f"skipped {skipped_no_date} (no date) + {skipped_empty} (empty), "
          f"{errors} errors")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run():
    """Ejecuta la carga completa de datos en orden de dependencias FK."""
    print("=" * 70)
    print("CARGA MASIVA DE DATOS DESDE EXCEL")
    print(f"Data directory: {DATA_DIR}")
    print("=" * 70)

    # Verificar que existen los archivos
    files = [
        FILE_PROGRAMAS,
        FILE_VARIEDADES,
        FILE_PORTAINJERTOS,
        FILE_MAESTROS_GLOBAL,
        FILE_MAESTRO_CEREZAS,
    ]
    for f in files:
        if not f.exists():
            print(f"ERROR: Archivo no encontrado: {f}")
            sys.exit(1)
        print(f"  OK: {f.name}")

    db = SessionLocal()
    try:
        # Step 1: PMGs
        pmg_cache = load_pmgs(db)

        # Step 2: Especies
        especie_cache = load_especies(db)

        # Step 3: Variedades
        var_cache = load_variedades(db, pmg_cache, especie_cache)

        # Step 4: Portainjertos
        pi_cache = load_portainjertos(db)

        # Step 5: Campos
        campo_cache = load_campos(db)

        # Step 6: TestBlocks + Posiciones
        load_testblocks_global(db, campo_cache, especie_cache, pmg_cache, pi_cache, var_cache)

        # Step 7: Mediciones cerezas
        load_mediciones_cerezas(db, campo_cache, especie_cache, pmg_cache, pi_cache, var_cache)

        print("\n" + "=" * 70)
        print("CARGA COMPLETADA EXITOSAMENTE")
        print("=" * 70)

    except Exception as exc:
        db.rollback()
        print(f"\nERROR CRITICO: {exc}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    run()
