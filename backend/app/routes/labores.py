"""Labores routes: planificacion, ejecucion, ordenes de trabajo, dashboard, evidencias, QR, fenologia."""

import json
import random
from datetime import date, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import qrcode
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user, require_role, require_non_production
from app.models.sistema import Usuario
from app.models.laboratorio import EjecucionLabor, RegistroFenologico, DetalleLabor
from app.models.maestras import TipoLabor, EstadoFenologico, Especie
from app.models.testblock import PosicionTestBlock
from app.models.evidencia import EvidenciaLabor
from app.schemas.laboratorio import LaborPlanificacion, LaborPlanificacionTestblock, LaborEjecucion
from app.services import crud

router = APIRouter(prefix="/labores", tags=["Labores"])


# ---------------------------------------------------------------------------
# Tipos de labor
# ---------------------------------------------------------------------------

@router.get("/tipos-labor")
def list_tipos_labor(
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """List all active labor types."""
    return db.query(TipoLabor).filter(TipoLabor.activo == True).all()


@router.post("/seed-tipos-labor", status_code=201)
def seed_tipos_labor(
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
    _guard: bool = Depends(require_non_production),
):
    """Seed basic tipos_labor if the table is empty. Admin only. EF-4: non-prod only."""
    existing = db.query(TipoLabor).count()
    if existing > 0:
        return {"message": f"Ya existen {existing} tipos de labor. No se inserto nada.", "created": 0}

    seed_data = [
        # === Labores reales de Garces Fruit (Excel Planificacion) ===
        {"codigo": "FORMACION", "nombre": "Formacion", "categoria": "manejo",
         "descripcion": "Formacion del arbol: definir eje, amarrar, eliminar sierpes", "aplica_a": "planta",
         "aplica_especies": "General"},
        {"codigo": "PLANTACION", "nombre": "Plantacion", "categoria": "manejo",
         "descripcion": "Plantacion de nuevos ejemplares en testblock", "aplica_a": "planta",
         "aplica_especies": "General"},
        {"codigo": "INJERTACION", "nombre": "Injertacion", "categoria": "manejo",
         "descripcion": "Injertacion de yema, chequeo cicatrizacion, manejo de cintas", "aplica_a": "planta",
         "aplica_especies": "General"},
        {"codigo": "INCISIONES", "nombre": "Incisiones", "categoria": "manejo",
         "descripcion": "Incisiones en yemas para estimular brotacion", "aplica_a": "planta",
         "aplica_especies": "Cerezo"},
        {"codigo": "ORTOPEDIA", "nombre": "Ortopedia", "categoria": "manejo",
         "descripcion": "Ortopedia de laterales, entrepisos y punteros", "aplica_a": "planta",
         "aplica_especies": "Cerezo,Ciruela"},
        {"codigo": "RALEO", "nombre": "Raleo", "categoria": "manejo",
         "descripcion": "Eliminacion de frutos excedentes para mejorar calibre", "aplica_a": "planta",
         "aplica_especies": "Carozo,Ciruela"},
        {"codigo": "COSECHA", "nombre": "Cosecha", "categoria": "cosecha",
         "descripcion": "Cosecha de frutos: numero de pasada, rendimiento por planta", "aplica_a": "planta",
         "aplica_especies": "Cerezo,General"},
        {"codigo": "RIEGO", "nombre": "Riego", "categoria": "riego",
         "descripcion": "Riego programado, revision de calicata", "aplica_a": "testblock",
         "aplica_especies": "General"},
        {"codigo": "MALEZAS", "nombre": "Malezas", "categoria": "fitosanidad",
         "descripcion": "Control de malezas con herbicida de contacto", "aplica_a": "testblock",
         "aplica_especies": "General"},
        {"codigo": "CARTELES", "nombre": "Carteles", "categoria": "manejo",
         "descripcion": "Realizar carteles post injertacion para identificacion", "aplica_a": "testblock",
         "aplica_especies": "General"},
        {"codigo": "REG_FENOL", "nombre": "Registro fenologico", "categoria": "fenologia",
         "descripcion": "Registro del estado fenologico actual de la planta", "aplica_a": "planta",
         "aplica_especies": "Cerezo,Ciruela,Carozo"},
    ]

    created = 0
    for item in seed_data:
        tl = TipoLabor(**item)
        db.add(tl)
        created += 1

    db.commit()
    return {"message": f"Se crearon {created} tipos de labor.", "created": created}


# ---------------------------------------------------------------------------
# Detalles labor: sub-items/instructions per labor type and species
# ---------------------------------------------------------------------------

SEED_DETALLES_LABOR: dict[str, dict[str, list[str]]] = {
    "FORMACION": {
        "General": [
            "Definir eje",
            "Definir dos ejes",
            "Amarrar eje al colihue, sin estrangular",
            "Amarrar los dos ejes a los colihues, sin estrangular",
            "Definir laterales (dejar uno por cordel y un entrepiso desde el tercer cordel hacia arriba) eliminar el resto con corte apegado, priorizar brotes con vigor y activos",
            "Eliminar brotes restantes de la variedad",
            "Eliminar sierpes",
            "Limpiar sierpes",
        ],
    },
    "PLANTACION": {
        "General": [
            "Dimension hoyo de plantacion 50 x 50 x 50 cm de profundidad (dependera del tamano del pan de raices)",
            "El suelo debe estar a capacidad de campo",
            "Sacar planta de la bolsa plastica / maceta",
            "Escarmenar raices (lateralmente y en la base)",
            "Aplicar 10 gr por hoyo de nematicida RUGBY alrededor de las raices (sin tocarlas)",
            "Presentar planta y acomodar al suelo del hoyo de plantacion al mismo nivel que viene el sustrato",
            "Tapar raices con tierra mullida, libre de bloques o piedras",
            "Terminar de plantar con tierra mullida realizando un lomo de 10 a 20 cm sobre el nivel de suelo (jamas pisar el hoyo)",
            "Poner protector para evitar dano de animales y del control de malezas",
            "Riego post plantacion",
        ],
    },
    "INJERTACION": {
        "General": [
            "Limpiar portainjertos los primeros 50 cm",
            "Injertar de yema",
            "Cubrir firmemente con cinta",
            "Chequear cicatrizacion",
            "Eliminar cinta",
            "Quebrar brotes sobre el injerto",
            "Quebrar eje 40 cm sobre el injerto",
            "Despuntar apices activos de brotes",
            "Rebajar ejes 10 cm sobre la ultima yema injertada",
        ],
    },
    "INCISIONES": {
        "Cerezo": [
            "Realizar corte sobre yemas en estado puntas verdes",
            "Realizar incisiones cada 10 cm o un puno en yemas verdes dejando los ultimos 30 cm libres",
            "Realizar un corte (que salga aserrin) con sierra sobre las yemas ubicadas en los cordeles (2 yemas por cordel)",
            "Realizar un corte (que salga aserrin) con sierra sobre las yemas ubicadas en los entrepisos (2 yemas)",
            "Dejar 30 cm libre de incisiones cuando el eje no ha llegado a su altura",
            "Pintar completamente el corte",
            "Dosis producto bien mezclado: 50 cc de promalina/perlan + 5 gr de streptoplus/agrygent plus + 1 Litro de pintura latex negro",
            "Repase incisiones",
        ],
    },
    "ORTOPEDIA": {
        "Cerezo": [
            "Ortopediar todos los laterales de manera que queden horizontales",
            "Ortopediar laterales que tengan desde 60 cm de largo en portainjertos vigorosos",
            "Ortopediar laterales que tengan desde 80 cm de largo en portainjertos debiles",
            "Ortopediar punteros",
            "Ortopediar sobre piso (sobre el alambre)",
            "Ortopediar un entre piso por lado desde el tercer cordel hacia arriba",
        ],
        "Ciruela": [
            "Ortopediar laterales",
            "Ortopediar entrepisos",
        ],
    },
    "RALEO": {
        "Carozo": [
            "Dejar 1 fruto en la parte baja",
            "Dejar 3 frutos en la parte aerea",
            "Dejar 2 frutos en la parte baja",
            "Dejar 1 fruto en ramillas de 20 o menos cm",
            "Dejar 2 frutos en ramillas entre 20 y 40 cm",
            "Dejar 3 frutos en ramillas sobre 40 cm",
        ],
        "Ciruela": [
            "Dejar X frutos por planta",
            "Dejar frutos distribuidos a X cm",
            "Dejar frutos mas grandes",
        ],
    },
    "COSECHA": {
        "Cerezo": [
            "Numero de pasada",
            "Rendimiento por planta",
        ],
        "General": [
            "Observaciones",
        ],
    },
    "RIEGO": {
        "General": [
            "Revision de calicata si / no",
            "Comentarios",
        ],
    },
    "MALEZAS": {
        "General": [
            "Aplicar herbicida de contacto",
        ],
    },
    "CARTELES": {
        "General": [
            "Realizar carteles post injertacion",
        ],
    },
}


@router.post("/seed-detalles-labor", status_code=201)
def seed_detalles_labor(
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
    _guard: bool = Depends(require_non_production),
):
    """Seed detalles_labor (sub-items/instructions per labor type). Admin only. EF-4: non-prod only."""
    existing = db.query(DetalleLabor).count()
    if existing > 0:
        return {"message": f"Ya existen {existing} detalles de labor. No se inserto nada.", "created": 0}

    created = 0
    for labor_codigo, especies_dict in SEED_DETALLES_LABOR.items():
        tipo_labor = db.query(TipoLabor).filter(TipoLabor.codigo == labor_codigo).first()
        if not tipo_labor:
            continue
        for especie_nombre, detalles in especies_dict.items():
            for orden, desc in enumerate(detalles, 1):
                dl = DetalleLabor(
                    id_labor=tipo_labor.id_labor,
                    descripcion=desc,
                    aplica_especie=especie_nombre,
                    orden=orden,
                    usuario_creacion=user.username,
                )
                db.add(dl)
                created += 1

    db.commit()
    return {"message": f"Se crearon {created} detalles de labor.", "created": created}


# ---------------------------------------------------------------------------
# CRUD Detalles labor
# ---------------------------------------------------------------------------

@router.get("/tipos-labor/{id_labor}/detalles")
def list_detalles_labor(
    id_labor: int,
    especie: str | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """List detail items for a labor type, optionally filtered by species."""
    q = db.query(DetalleLabor).filter(
        DetalleLabor.id_labor == id_labor,
        DetalleLabor.activo == True,
    )
    if especie:
        q = q.filter(DetalleLabor.aplica_especie.in_([especie, "General"]))
    return q.order_by(DetalleLabor.aplica_especie, DetalleLabor.orden).all()


@router.post("/tipos-labor/{id_labor}/detalles", status_code=201)
def create_detalle_labor(
    id_labor: int,
    data: dict,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo")),
):
    """Add a new detail item to a labor type."""
    if "descripcion" not in data or not data["descripcion"]:
        raise HTTPException(status_code=422, detail="descripcion es requerida")
    dl = DetalleLabor(
        id_labor=id_labor,
        descripcion=data["descripcion"],
        aplica_especie=data.get("aplica_especie", "General"),
        orden=data.get("orden", 0),
        es_checklist=data.get("es_checklist", True),
        usuario_creacion=user.username,
    )
    db.add(dl)
    db.commit()
    db.refresh(dl)
    return dl


@router.put("/detalles-labor/{id_detalle}")
def update_detalle_labor(
    id_detalle: int,
    data: dict,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo")),
):
    """Update a detail item."""
    dl = db.get(DetalleLabor, id_detalle)
    if not dl:
        raise HTTPException(status_code=404, detail="Detalle no encontrado")
    for field in ["descripcion", "aplica_especie", "orden", "es_checklist", "activo"]:
        if field in data:
            setattr(dl, field, data[field])
    dl.usuario_modificacion = user.username
    from app.core.utils import utcnow
    dl.fecha_modificacion = utcnow()
    db.commit()
    db.refresh(dl)
    return dl


@router.delete("/detalles-labor/{id_detalle}")
def delete_detalle_labor(
    id_detalle: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Soft-delete a detail item."""
    dl = db.get(DetalleLabor, id_detalle)
    if not dl:
        raise HTTPException(status_code=404, detail="Detalle no encontrado")
    dl.activo = False
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Seed estados fenologicos
# ---------------------------------------------------------------------------

_MES_MAP = {
    "Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12,
}

def _parse_mes_range(mes_orientativo: str | None) -> tuple[int | None, int | None]:
    """Parse 'Jul-Sep' -> (7, 9), 'Oct' -> (10, 10), None -> (None, None)."""
    if not mes_orientativo:
        return None, None
    parts = [p.strip() for p in mes_orientativo.replace("/", "-").split("-")]
    m_ini = _MES_MAP.get(parts[0])
    m_fin = _MES_MAP.get(parts[-1]) if len(parts) > 1 else m_ini
    return m_ini, m_fin

SEED_ESTADOS_FENOLOGICOS: dict[str, list[dict]] = {
    # Rangos ampliados: variedades tempranas pueden diferir 1-2 meses de las tardias
    "Cerezo": [
        {"codigo": "CER_YEMA_HINCH", "nombre": "Yema hinchada", "orden": 1, "mes_orientativo": "Jul-Sep", "color_hex": "#9ACD32"},
        {"codigo": "CER_PUNTA_VERDE", "nombre": "Puntas verdes", "orden": 2, "mes_orientativo": "Ago-Oct", "color_hex": "#6B8E23"},
        {"codigo": "CER_FLOR_INI", "nombre": "Inicio de flor (5-10%)", "orden": 3, "mes_orientativo": "Ago-Oct", "color_hex": "#FFB6C1"},
        {"codigo": "CER_FLOR_PLENA", "nombre": "Plena flor (100%)", "orden": 4, "mes_orientativo": "Sep-Nov", "color_hex": "#FF69B4"},
        {"codigo": "CER_CAIDA_PET", "nombre": "Caida de petalos", "orden": 5, "mes_orientativo": "Sep-Nov", "color_hex": "#DDA0DD"},
        {"codigo": "CER_CAIDA_CHQ", "nombre": "Caida de chaqueta", "orden": 6, "mes_orientativo": "Oct-Nov", "color_hex": "#C9A0DC"},
        {"codigo": "CER_CUAJA", "nombre": "Fruto cuajado", "orden": 7, "mes_orientativo": "Oct-Dic", "color_hex": "#90EE90"},
        {"codigo": "CER_ARVEJADO", "nombre": "Fruto arvejado (10 mm)", "orden": 8, "mes_orientativo": "Oct-Dic", "color_hex": "#32CD32"},
        {"codigo": "CER_PAJIZO", "nombre": "Fruto color pajizo", "orden": 9, "mes_orientativo": "Nov-Dic", "color_hex": "#DAA520"},
        {"codigo": "CER_VIRAJE", "nombre": "Viraje de color", "orden": 10, "mes_orientativo": "Nov-Ene", "color_hex": "#DC143C"},
        {"codigo": "CER_PRE_COS", "nombre": "Color pre cosecha", "orden": 11, "mes_orientativo": "Nov-Ene", "color_hex": "#8B0000"},
        {"codigo": "CER_REG_FOTO", "nombre": "Registro fotografico", "orden": 12, "mes_orientativo": None, "color_hex": "#4682B4"},
        {"codigo": "CER_HOJ_CAI_INI", "nombre": "Inicio caida de hoja", "orden": 13, "mes_orientativo": "Mar-Jun", "color_hex": "#A0522D"},
        {"codigo": "CER_HOJ_CAI_50", "nombre": "50% caida de hoja", "orden": 14, "mes_orientativo": "Abr-Jun", "color_hex": "#8B4513"},
        {"codigo": "CER_HOJ_CAI_100", "nombre": "100% caida de hoja", "orden": 15, "mes_orientativo": "May-Jul", "color_hex": "#6B3A2A"},
    ],
    "Ciruela": [
        {"codigo": "CIR_YEMA_INV", "nombre": "Yema de invierno", "orden": 1, "mes_orientativo": "May-Jul", "color_hex": "#708090"},
        {"codigo": "CIR_YEMA_HINCH", "nombre": "Yema hinchada", "orden": 2, "mes_orientativo": "Jul-Sep", "color_hex": "#9ACD32"},
        {"codigo": "CIR_BOT_SEP", "nombre": "Botones separados", "orden": 3, "mes_orientativo": "Ago-Oct", "color_hex": "#BDB76B"},
        {"codigo": "CIR_FLOR_INI", "nombre": "Inicio de flor", "orden": 4, "mes_orientativo": "Ago-Oct", "color_hex": "#DDA0DD"},
        {"codigo": "CIR_FLOR_PLENA", "nombre": "Plena flor (100%)", "orden": 5, "mes_orientativo": "Sep-Oct", "color_hex": "#9932CC"},
        {"codigo": "CIR_CAIDA_PET", "nombre": "Caida de petalos", "orden": 6, "mes_orientativo": "Sep-Nov", "color_hex": "#E6C2DC"},
        {"codigo": "CIR_CUAJA", "nombre": "Fruto cuajado", "orden": 7, "mes_orientativo": "Oct-Nov", "color_hex": "#90EE90"},
        {"codigo": "CIR_FRUTO_8MM", "nombre": "Fruto (8mm)", "orden": 8, "mes_orientativo": "Oct-Dic", "color_hex": "#32CD32"},
        {"codigo": "CIR_END_CAROZO", "nombre": "Endurecimiento carozo", "orden": 9, "mes_orientativo": "Nov-Dic", "color_hex": "#556B2F"},
        {"codigo": "CIR_VIRAJE", "nombre": "Viraje color (pinta)", "orden": 10, "mes_orientativo": "Nov-Ene", "color_hex": "#8B008B"},
        {"codigo": "CIR_COSECHA", "nombre": "Madurez de cosecha", "orden": 11, "mes_orientativo": "Dic-Feb", "color_hex": "#4B0082"},
        {"codigo": "CIR_REG_FOTO", "nombre": "Registro fotografico", "orden": 12, "mes_orientativo": None, "color_hex": "#4682B4"},
        {"codigo": "CIR_HOJ_CAI_INI", "nombre": "Inicio caida de hoja", "orden": 13, "mes_orientativo": "Mar-Jun", "color_hex": "#A0522D"},
        {"codigo": "CIR_HOJ_CAI_50", "nombre": "50% caida de hoja", "orden": 14, "mes_orientativo": "Abr-Jun", "color_hex": "#8B4513"},
        {"codigo": "CIR_HOJ_CAI_100", "nombre": "100% caida de hoja", "orden": 15, "mes_orientativo": "May-Jul", "color_hex": "#6B3A2A"},
    ],
    "Carozo": [
        {"codigo": "CAR_YEMA_HINCH", "nombre": "Yema hinchada", "orden": 1, "mes_orientativo": "Jul-Sep", "color_hex": "#9ACD32"},
        {"codigo": "CAR_PUNTA_VERDE", "nombre": "Puntas verdes", "orden": 2, "mes_orientativo": "Ago-Oct", "color_hex": "#6B8E23"},
        {"codigo": "CAR_FLOR_INI", "nombre": "Inicio flor", "orden": 3, "mes_orientativo": "Ago-Oct", "color_hex": "#FFB6C1"},
        {"codigo": "CAR_FLOR_PLENA", "nombre": "Plena flor (100%)", "orden": 4, "mes_orientativo": "Sep-Oct", "color_hex": "#FF69B4"},
        {"codigo": "CAR_CAIDA_PET", "nombre": "Caida de petalos", "orden": 5, "mes_orientativo": "Sep-Nov", "color_hex": "#DDA0DD"},
        {"codigo": "CAR_CAIDA_CHQ", "nombre": "Caida de chaqueta", "orden": 6, "mes_orientativo": "Oct-Nov", "color_hex": "#C9A0DC"},
        {"codigo": "CAR_CUAJA", "nombre": "Fruto cuajado", "orden": 7, "mes_orientativo": "Oct-Nov", "color_hex": "#90EE90"},
        {"codigo": "CAR_END_CAROZO", "nombre": "Endurecimiento carozo", "orden": 8, "mes_orientativo": "Nov-Dic", "color_hex": "#556B2F"},
        {"codigo": "CAR_VIRAJE", "nombre": "Viraje color (pinta)", "orden": 9, "mes_orientativo": "Nov-Ene", "color_hex": "#DC143C"},
        {"codigo": "CAR_COSECHA", "nombre": "Madurez de cosecha", "orden": 10, "mes_orientativo": "Dic-Feb", "color_hex": "#8B0000"},
        {"codigo": "CAR_REG_FOTO", "nombre": "Registro fotografico", "orden": 11, "mes_orientativo": None, "color_hex": "#4682B4"},
        {"codigo": "CAR_HOJ_CAI_INI", "nombre": "Inicio caida de hoja", "orden": 12, "mes_orientativo": "Mar-Jun", "color_hex": "#A0522D"},
        {"codigo": "CAR_HOJ_CAI_50", "nombre": "50% caida de hoja", "orden": 13, "mes_orientativo": "Abr-Jun", "color_hex": "#8B4513"},
        {"codigo": "CAR_HOJ_CAI_100", "nombre": "100% caida de hoja", "orden": 14, "mes_orientativo": "May-Jul", "color_hex": "#6B3A2A"},
    ],
    "Nectarina": [
        {"codigo": "NEC_HOJ_CAI", "nombre": "Caida de hoja", "orden": 1, "mes_orientativo": "Abr-May", "color_hex": "#A0522D"},
        {"codigo": "NEC_YEMA_DORM", "nombre": "Yema dormante", "orden": 2, "mes_orientativo": "Jun-Jul", "color_hex": "#708090"},
        {"codigo": "NEC_YEMA_HINCH", "nombre": "Yema hinchada", "orden": 3, "mes_orientativo": "Ago", "color_hex": "#9ACD32"},
        {"codigo": "NEC_FLOR_INI", "nombre": "Inicio floracion", "orden": 4, "mes_orientativo": "Sep", "color_hex": "#FFB6C1"},
        {"codigo": "NEC_FLOR_PLENA", "nombre": "Plena floracion", "orden": 5, "mes_orientativo": "Sep", "color_hex": "#FF1493"},
        {"codigo": "NEC_CUAJA", "nombre": "Cuaja", "orden": 6, "mes_orientativo": "Oct", "color_hex": "#90EE90"},
        {"codigo": "NEC_RALEO", "nombre": "Estado raleo", "orden": 7, "mes_orientativo": "Oct-Nov", "color_hex": "#228B22"},
        {"codigo": "NEC_CRECIM", "nombre": "Crecimiento rapido", "orden": 8, "mes_orientativo": "Nov", "color_hex": "#00FF7F"},
        {"codigo": "NEC_MADURACION", "nombre": "Maduracion", "orden": 9, "mes_orientativo": "Dic-Ene", "color_hex": "#FF8C00"},
    ],
    "Durazno": [
        {"codigo": "DUR_HOJ_CAI", "nombre": "Caida de hoja", "orden": 1, "mes_orientativo": "Abr-May", "color_hex": "#A0522D"},
        {"codigo": "DUR_YEMA_DORM", "nombre": "Yema dormante", "orden": 2, "mes_orientativo": "Jun-Jul", "color_hex": "#708090"},
        {"codigo": "DUR_YEMA_HINCH", "nombre": "Yema hinchada", "orden": 3, "mes_orientativo": "Jul-Ago", "color_hex": "#9ACD32"},
        {"codigo": "DUR_FLOR_INI", "nombre": "Inicio floracion", "orden": 4, "mes_orientativo": "Ago-Sep", "color_hex": "#FFB6C1"},
        {"codigo": "DUR_FLOR_PLENA", "nombre": "Plena floracion", "orden": 5, "mes_orientativo": "Sep", "color_hex": "#FF69B4"},
        {"codigo": "DUR_CUAJA", "nombre": "Cuaja", "orden": 6, "mes_orientativo": "Sep-Oct", "color_hex": "#90EE90"},
        {"codigo": "DUR_CRECIM", "nombre": "Crecimiento de fruto", "orden": 7, "mes_orientativo": "Oct-Nov", "color_hex": "#228B22"},
        {"codigo": "DUR_PINTADO", "nombre": "Pintado / Color", "orden": 8, "mes_orientativo": "Nov-Dic", "color_hex": "#FF4500"},
        {"codigo": "DUR_MADURACION", "nombre": "Maduracion cosecha", "orden": 9, "mes_orientativo": "Dic-Ene", "color_hex": "#FF6347"},
    ],
}


@router.post("/seed-estados-fenologicos", status_code=201)
def seed_estados_fenologicos(
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
    _guard: bool = Depends(require_non_production),
):
    """Seed estados_fenologicos for 4 main species. Admin only.

    EF-4: blocked in production — this is the endpoint QA flagged as
    "expuesto a admin sin confirmación". Re-running could corrupt the 82
    live states that labores/fenología comparativa/AI depend on.

    - Updates existing records for species that already have data (fills missing fields).
    - Inserts new records for species that have no data yet.
    """
    created = 0
    updated = 0
    skipped_species = []

    for especie_nombre, estados in SEED_ESTADOS_FENOLOGICOS.items():
        especie = db.query(Especie).filter(Especie.nombre == especie_nombre).first()
        if not especie:
            skipped_species.append(especie_nombre)
            continue

        existing = (
            db.query(EstadoFenologico)
            .filter(EstadoFenologico.id_especie == especie.id_especie)
            .all()
        )

        if existing:
            # Update existing records: fill in missing fields (mes_orientativo, color_hex, activo, mes_inicio, mes_fin)
            existing_by_orden = {e.orden: e for e in existing}
            for seed_item in estados:
                ef = existing_by_orden.get(seed_item["orden"])
                if ef:
                    if not ef.color_hex:
                        ef.color_hex = seed_item.get("color_hex")
                    if not getattr(ef, "mes_orientativo", None):
                        ef.mes_orientativo = seed_item.get("mes_orientativo")
                    if getattr(ef, "activo", None) is None:
                        ef.activo = True
                    # Fill mes_inicio / mes_fin from mes_orientativo if not set
                    if not getattr(ef, "mes_inicio", None):
                        m_ini, m_fin = _parse_mes_range(ef.mes_orientativo or seed_item.get("mes_orientativo"))
                        ef.mes_inicio = m_ini
                        ef.mes_fin = m_fin
                    db.add(ef)
                    updated += 1
        else:
            # Insert new records for this species
            for estado in estados:
                m_ini, m_fin = _parse_mes_range(estado.get("mes_orientativo"))
                ef = EstadoFenologico(
                    id_especie=especie.id_especie,
                    mes_inicio=m_ini,
                    mes_fin=m_fin,
                    **estado,
                )
                db.add(ef)
                created += 1

    db.commit()
    msg = f"Seed completo: {created} creados, {updated} actualizados."
    if skipped_species:
        msg += f" Especies no encontradas: {', '.join(skipped_species)}"
    return {"message": msg, "created": created, "updated": updated, "skipped_species": skipped_species}


# ---------------------------------------------------------------------------
# Dashboard KPIs
# ---------------------------------------------------------------------------

@router.get("/dashboard")
def labores_dashboard(
    testblock: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Dashboard KPIs and stats for labores."""
    q = db.query(EjecucionLabor)
    if testblock:
        pos_ids = [
            p.id_posicion
            for p in db.query(PosicionTestBlock.id_posicion).filter(
                PosicionTestBlock.id_testblock == testblock
            ).all()
        ]
        q = q.filter(EjecucionLabor.id_posicion.in_(pos_ids))

    all_labores = q.all()
    today = date.today()

    planificadas = [l for l in all_labores if l.estado == "planificada"]
    ejecutadas = [l for l in all_labores if l.estado == "ejecutada"]

    # Delayed: planificada and fecha_programada < today
    atrasadas = [
        l for l in planificadas
        if l.fecha_programada and l.fecha_programada < today
    ]

    # This week
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    esta_semana = [
        l for l in planificadas
        if l.fecha_programada and week_start <= l.fecha_programada <= week_end
    ]

    # Pre-load all labor types (avoid N+1)
    labor_names: dict[int, str] = {}
    for t in db.query(TipoLabor).all():
        labor_names[t.id_labor] = t.nombre

    # Group by tipo_labor
    por_tipo: dict[str, dict[str, int]] = {}
    for l in all_labores:
        nombre = labor_names.get(l.id_labor, f"Labor #{l.id_labor}")
        if nombre not in por_tipo:
            por_tipo[nombre] = {"planificadas": 0, "ejecutadas": 0, "atrasadas": 0}
        if l.estado == "planificada":
            por_tipo[nombre]["planificadas"] += 1
            if l.fecha_programada and l.fecha_programada < today:
                por_tipo[nombre]["atrasadas"] += 1
        elif l.estado == "ejecutada":
            por_tipo[nombre]["ejecutadas"] += 1

    # Group by month
    por_mes: dict[str, dict[str, int]] = {}
    for l in all_labores:
        if l.fecha_programada:
            mes = l.fecha_programada.strftime("%Y-%m")
            if mes not in por_mes:
                por_mes[mes] = {"planificadas": 0, "ejecutadas": 0}
            if l.estado == "planificada":
                por_mes[mes]["planificadas"] += 1
            else:
                por_mes[mes]["ejecutadas"] += 1

    return {
        "total": len(all_labores),
        "planificadas": len(planificadas),
        "ejecutadas": len(ejecutadas),
        "atrasadas": len(atrasadas),
        "esta_semana": len(esta_semana),
        "por_tipo": por_tipo,
        "por_mes": dict(sorted(por_mes.items())),
        "pct_cumplimiento": round(len(ejecutadas) / max(len(all_labores), 1) * 100, 1),
    }


# ---------------------------------------------------------------------------
# Count (lightweight badge endpoint)
# ---------------------------------------------------------------------------

@router.get("/count")
def labores_count(
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Lightweight count for sidebar badges — avoids loading all records."""
    from sqlalchemy import func
    total = db.query(func.count(EjecucionLabor.id_ejecucion)).scalar()
    pendientes = db.query(func.count(EjecucionLabor.id_ejecucion)).filter(
        EjecucionLabor.estado == "planificada"
    ).scalar()
    atrasadas = db.query(func.count(EjecucionLabor.id_ejecucion)).filter(
        EjecucionLabor.estado == "planificada",
        EjecucionLabor.fecha_programada < date.today(),
    ).scalar()
    return {"total": total, "pendientes": pendientes, "atrasadas": atrasadas or 0}


# ---------------------------------------------------------------------------
# Planificacion
# ---------------------------------------------------------------------------

@router.get("/planificacion")
def list_planificacion(
    testblock: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Lista ejecuciones labor.

    L-20: enriquece id_posicion con posicion_display ("H3P5") y id_testblock
    para que el frontend no muestre IDs numéricos crudos.
    """
    q = db.query(EjecucionLabor)
    if testblock:
        pos_ids = [p.id_posicion for p in db.query(PosicionTestBlock.id_posicion).filter(
            PosicionTestBlock.id_testblock == testblock
        ).all()]
        q = q.filter(EjecucionLabor.id_posicion.in_(pos_ids))
    labores = q.order_by(EjecucionLabor.fecha_programada).all()

    # Batch lookup de posiciones para evitar N+1
    pos_ids_in_use = list({l.id_posicion for l in labores if l.id_posicion})
    pos_map: dict[int, tuple[int, int, int | None]] = {}
    if pos_ids_in_use:
        rows = db.query(
            PosicionTestBlock.id_posicion,
            PosicionTestBlock.hilera,
            PosicionTestBlock.posicion,
            PosicionTestBlock.id_testblock,
        ).filter(PosicionTestBlock.id_posicion.in_(pos_ids_in_use)).all()
        pos_map = {r[0]: (r[1], r[2], r[3]) for r in rows}

    result = []
    for l in labores:
        d = {c: getattr(l, c) for c in l.__class__.model_fields}
        if l.id_posicion and l.id_posicion in pos_map:
            h, p, tb = pos_map[l.id_posicion]
            d["posicion_display"] = f"H{h}P{p}"
            d["id_testblock"] = tb
        else:
            d["posicion_display"] = None
            d["id_testblock"] = None
        result.append(d)
    return result


@router.post("/planificacion", status_code=201)
def create_planificacion(
    data: LaborPlanificacion,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo", "operador")),
):
    # Resolve id_lote from position if available
    pos_lote = None
    if data.id_posicion:
        pos = db.query(PosicionTestBlock).filter(PosicionTestBlock.id_posicion == data.id_posicion).first()
        if pos:
            pos_lote = pos.id_lote

    labor = EjecucionLabor(
        id_posicion=data.id_posicion,
        id_planta=data.id_planta,
        id_labor=data.id_labor,
        temporada=data.temporada,
        fecha_programada=data.fecha_programada,
        fecha_ejecucion=data.fecha_programada,
        estado="planificada",
        observaciones=data.observaciones,
        usuario_registro=user.username,
        id_lote=pos_lote,
    )
    db.add(labor)
    db.commit()
    db.refresh(labor)

    # S-8: audit
    try:
        from app.services.audit_service import log_audit
        log_audit(
            db,
            tabla="ejecucion_labores",
            registro_id=labor.id_ejecucion,
            accion="CREATE",
            datos_nuevos=json.dumps({
                "id_posicion": data.id_posicion,
                "id_labor": data.id_labor,
                "temporada": data.temporada,
                "fecha_programada": str(data.fecha_programada),
            }, default=str, ensure_ascii=False),
            usuario=user.username,
        )
        db.commit()
    except Exception:
        pass
    return labor


@router.post("/planificacion-testblock", status_code=201)
def create_planificacion_testblock(
    data: LaborPlanificacionTestblock,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo", "operador")),
):
    """Plan a labor for all active positions in a testblock."""
    posiciones = (
        db.query(PosicionTestBlock)
        .filter(
            PosicionTestBlock.id_testblock == data.id_testblock,
            PosicionTestBlock.estado == "alta",
        )
        .all()
    )
    if not posiciones:
        raise HTTPException(status_code=400, detail="No hay posiciones activas en este testblock")

    created = 0
    for pos in posiciones:
        labor = EjecucionLabor(
            id_posicion=pos.id_posicion,
            id_labor=data.id_labor,
            temporada=data.temporada,
            fecha_programada=data.fecha_programada,
            fecha_ejecucion=data.fecha_programada,
            estado="planificada",
            observaciones=data.observaciones,
            usuario_registro=user.username,
            id_lote=pos.id_lote,
        )
        db.add(labor)
        created += 1

    db.commit()

    # S-8: audit — una sola entrada resumen del bulk
    try:
        from app.services.audit_service import log_audit
        log_audit(
            db,
            tabla="ejecucion_labores",
            registro_id=None,
            accion="BULK_CREATE_TESTBLOCK",
            datos_nuevos=json.dumps({
                "id_testblock": data.id_testblock,
                "id_labor": data.id_labor,
                "temporada": data.temporada,
                "fecha_programada": str(data.fecha_programada),
                "created": created,
            }, default=str, ensure_ascii=False),
            usuario=user.username,
        )
        db.commit()
    except Exception:
        pass
    return {"created": created, "testblock": data.id_testblock}


# ---------------------------------------------------------------------------
# Labores de hoy / atrasadas
# ---------------------------------------------------------------------------

@router.get("/hoy")
def labores_hoy(
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Return labores due today or overdue (fecha_programada <= today, estado planificada), grouped by testblock."""
    today = date.today()
    rows = (
        db.query(EjecucionLabor)
        .filter(
            EjecucionLabor.fecha_programada <= today,
            EjecucionLabor.estado.in_(["planificada"]),
        )
        .order_by(EjecucionLabor.fecha_programada)
        .all()
    )

    # Build a position → testblock lookup for all positions referenced
    pos_ids = list({r.id_posicion for r in rows if r.id_posicion is not None})
    pos_tb_map: dict[int, int | None] = {}
    if pos_ids:
        for p in (
            db.query(PosicionTestBlock.id_posicion, PosicionTestBlock.id_testblock)
            .filter(PosicionTestBlock.id_posicion.in_(pos_ids))
            .all()
        ):
            pos_tb_map[p.id_posicion] = p.id_testblock

    # Group labores by testblock id
    grouped: dict[int | None, list] = {}
    for r in rows:
        tb_id = pos_tb_map.get(r.id_posicion)
        grouped.setdefault(tb_id, []).append(r)

    return {
        "total": len(rows),
        "por_testblock": {
            str(tb_id): items for tb_id, items in grouped.items()
        },
    }


# ---------------------------------------------------------------------------
# Ejecucion masiva
# ---------------------------------------------------------------------------

@router.post("/ejecutar-masivo")
def ejecutar_masivo(
    data: dict,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo", "operador")),
):
    """Mark multiple labores as executed in one call."""
    from datetime import datetime

    ids = data.get("ids", [])
    fecha = data.get("fecha_ejecucion", datetime.utcnow().date().isoformat())
    ejecutor = data.get("ejecutor", user.username)
    updated = 0
    for lid in ids:
        labor = db.get(EjecucionLabor, lid)
        if labor and labor.estado in ("planificada",):
            labor.estado = "ejecutada"
            labor.fecha_ejecucion = fecha
            labor.ejecutor = ejecutor
            db.add(labor)
            updated += 1
    db.commit()
    return {"updated": updated}


# ---------------------------------------------------------------------------
# Ejecucion
# ---------------------------------------------------------------------------

@router.put("/ejecucion/{id}")
def ejecutar_labor(
    id: int,
    data: LaborEjecucion,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo", "operador")),
):
    labor = crud.get_by_id(db, EjecucionLabor, id)
    labor.fecha_ejecucion = data.fecha_ejecucion
    labor.ejecutor = data.ejecutor
    labor.duracion_min = data.duracion_min
    labor.estado = "ejecutada"
    if data.observaciones:
        labor.observaciones = data.observaciones
    labor.usuario_registro = user.username
    db.commit()
    db.refresh(labor)

    # S-8: audit
    try:
        from app.services.audit_service import log_audit
        log_audit(
            db,
            tabla="ejecucion_labores",
            registro_id=id,
            accion="EJECUTAR",
            datos_nuevos=json.dumps({
                "fecha_ejecucion": str(data.fecha_ejecucion),
                "ejecutor": data.ejecutor,
                "duracion_min": data.duracion_min,
            }, default=str, ensure_ascii=False),
            usuario=user.username,
        )
        db.commit()
    except Exception:
        pass
    return labor


# ---------------------------------------------------------------------------
# Ordenes de trabajo
# ---------------------------------------------------------------------------

@router.get("/ordenes-trabajo")
def ordenes_trabajo(
    testblock: int | None = Query(None),
    fecha: date | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    q = db.query(EjecucionLabor).filter(EjecucionLabor.estado == "planificada")
    if testblock:
        pos_ids = [p.id_posicion for p in db.query(PosicionTestBlock.id_posicion).filter(
            PosicionTestBlock.id_testblock == testblock
        ).all()]
        q = q.filter(EjecucionLabor.id_posicion.in_(pos_ids))
    if fecha:
        q = q.filter(EjecucionLabor.fecha_programada == fecha)
    return q.order_by(EjecucionLabor.fecha_programada).all()


# ---------------------------------------------------------------------------
# Evidencias
# ---------------------------------------------------------------------------

@router.get("/ejecucion/{id}/evidencias")
def get_evidencias(
    id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """List all evidence items for a labor execution."""
    return (
        db.query(EvidenciaLabor)
        .filter(EvidenciaLabor.id_ejecucion == id)
        .order_by(EvidenciaLabor.fecha_creacion.desc())
        .all()
    )


@router.post("/ejecucion/{id}/evidencias", status_code=201)
def add_evidencia(
    id: int,
    data: dict,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo", "operador")),
):
    """Add photo/note evidence to a labor execution."""
    # Verify the execution exists
    crud.get_by_id(db, EjecucionLabor, id)
    ev = EvidenciaLabor(
        id_ejecucion=id,
        tipo=data.get("tipo", "foto"),
        descripcion=data.get("descripcion"),
        imagen_base64=data.get("imagen_base64"),
        url=data.get("url"),
        lat=data.get("lat"),
        lng=data.get("lng"),
        usuario=user.username,
    )
    db.add(ev)
    db.commit()
    db.refresh(ev)
    return ev


# ---------------------------------------------------------------------------
# Registro fenologico
# ---------------------------------------------------------------------------

@router.post("/registro-fenologico", status_code=201)
def registrar_fenologico(
    data: dict,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "agronomo", "laboratorio", "operador")),
):
    """Register fenological observation for one or more positions.

    Expects:
        id_estado_fenol: int  (FK to estados_fenologicos)
        posiciones_ids: list[int]
        porcentaje: int | None
        fecha: str (YYYY-MM-DD)
        observaciones: str
        temporada: str
    """
    from datetime import date as date_type

    posiciones_ids = data.get("posiciones_ids", [])
    id_estado_fenol = data.get("id_estado_fenol")
    porcentaje = data.get("porcentaje")
    fecha = data.get("fecha", date_type.today().isoformat())
    observaciones = data.get("observaciones", "")
    temporada = data.get("temporada", "2025-2026")

    if not posiciones_ids:
        raise HTTPException(status_code=400, detail="Debe seleccionar al menos una posicion")
    if not id_estado_fenol:
        raise HTTPException(status_code=400, detail="Debe indicar el estado fenologico (id_estado_fenol)")

    # Validate the estado_fenologico exists
    estado = db.get(EstadoFenologico, id_estado_fenol)
    if not estado:
        raise HTTPException(status_code=404, detail=f"Estado fenologico {id_estado_fenol} no encontrado")

    # Use the generic REG_FENOL tipo_labor (not a dynamic one)
    tipo_labor = db.query(TipoLabor).filter(TipoLabor.codigo == "REG_FENOL").first()
    if not tipo_labor:
        raise HTTPException(status_code=400, detail="Tipo de labor REG_FENOL no existe. Ejecute seed-tipos-labor primero.")

    created = 0
    for pos_id in posiciones_ids:
        # Resolve planta, lote, and testblock from position
        pos = db.get(PosicionTestBlock, pos_id)
        id_planta = getattr(pos, "id_planta", None) if pos else None
        pos_lote = pos.id_lote if pos else None
        pos_tb = pos.id_testblock if pos else None

        # Create EjecucionLabor entry (for labor tracking)
        ej = EjecucionLabor(
            id_labor=tipo_labor.id_labor,
            id_posicion=pos_id,
            id_planta=id_planta,
            temporada=temporada,
            fecha_programada=fecha,
            fecha_ejecucion=fecha,
            estado="ejecutada",
            ejecutor=user.username,
            observaciones=f"{estado.nombre}: {porcentaje}% - {observaciones}" if porcentaje else f"{estado.nombre} - {observaciones}",
            usuario_registro=user.username,
            id_lote=pos_lote,
        )
        db.add(ej)

        # Create RegistroFenologico entry with proper FK
        reg = RegistroFenologico(
            id_posicion=pos_id,
            id_planta=id_planta,
            id_estado_fenol=id_estado_fenol,
            temporada=temporada,
            fecha_registro=fecha,
            porcentaje=porcentaje,
            observaciones=observaciones,
            usuario_registro=user.username,
            id_testblock=pos_tb,
            id_lote=pos_lote,
        )
        db.add(reg)
        created += 1

    db.commit()
    return {"created": created, "estado": estado.nombre, "id_estado_fenol": id_estado_fenol}


@router.get("/historial-fenologico/{testblock_id}")
def historial_fenologico(
    testblock_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Get fenologia history for a testblock from registros_fenologicos."""
    from sqlalchemy import literal_column

    posiciones = db.query(PosicionTestBlock.id_posicion).filter(
        PosicionTestBlock.id_testblock == testblock_id
    ).all()
    pos_ids = [p.id_posicion for p in posiciones]

    if not pos_ids:
        return []

    rows = (
        db.query(RegistroFenologico)
        .filter(RegistroFenologico.id_posicion.in_(pos_ids))
        .order_by(RegistroFenologico.fecha_registro.desc())
        .limit(200)
        .all()
    )

    # Enrich with estado name/color for the response
    estado_ids = list({r.id_estado_fenol for r in rows if r.id_estado_fenol})
    estado_map: dict[int, dict] = {}
    if estado_ids:
        for ef in db.query(EstadoFenologico).filter(EstadoFenologico.id_estado.in_(estado_ids)).all():
            estado_map[ef.id_estado] = {
                "nombre": ef.nombre,
                "color_hex": ef.color_hex,
                "codigo": ef.codigo,
                "mes_orientativo": ef.mes_orientativo,
            }

    results = []
    for r in rows:
        item = {
            "id_registro": r.id_registro,
            "id_posicion": r.id_posicion,
            "id_planta": r.id_planta,
            "id_estado_fenol": r.id_estado_fenol,
            "temporada": r.temporada,
            "fecha_registro": str(r.fecha_registro) if r.fecha_registro else None,
            "porcentaje": r.porcentaje,
            "observaciones": r.observaciones,
            "usuario_registro": r.usuario_registro,
            "estado": estado_map.get(r.id_estado_fenol, {}) if r.id_estado_fenol else None,
        }
        results.append(item)

    return results


@router.get("/fenologia-comparativa")
def fenologia_comparativa(
    temporada_actual: str = Query(..., description="Ej: 2025-2026"),
    temporada_anterior: str = Query(..., description="Ej: 2024-2025"),
    id_especie: int | None = Query(None),
    id_variedad: int | None = Query(None),
    id_testblock: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Compare fenological records between two seasons.

    Returns per-estado: { estado, fecha_actual, fecha_anterior, diferencia_dias }
    grouped by variedad, so the user can see how each variety is tracking vs last year.
    """
    from sqlalchemy import func, case, and_

    # Base query: registros joined with posiciones and estados
    q_base = (
        db.query(
            RegistroFenologico.id_estado_fenol,
            RegistroFenologico.temporada,
            func.min(RegistroFenologico.fecha_registro).label("primera_fecha"),
            PosicionTestBlock.id_variedad,
        )
        .join(PosicionTestBlock, PosicionTestBlock.id_posicion == RegistroFenologico.id_posicion)
        .filter(
            RegistroFenologico.temporada.in_([temporada_actual, temporada_anterior]),
            PosicionTestBlock.id_variedad.isnot(None),
        )
    )

    if id_testblock:
        q_base = q_base.filter(PosicionTestBlock.id_testblock == id_testblock)
    if id_variedad:
        q_base = q_base.filter(PosicionTestBlock.id_variedad == id_variedad)
    if id_especie:
        from app.models.variedades import Variedad
        var_ids = [v.id_variedad for v in db.query(Variedad.id_variedad).filter(Variedad.id_especie == id_especie).all()]
        if var_ids:
            q_base = q_base.filter(PosicionTestBlock.id_variedad.in_(var_ids))
        else:
            return []

    rows = (
        q_base
        .group_by(RegistroFenologico.id_estado_fenol, RegistroFenologico.temporada, PosicionTestBlock.id_variedad)
        .all()
    )

    # Build lookup maps
    estado_ids = list({r.id_estado_fenol for r in rows})
    variedad_ids = list({r.id_variedad for r in rows if r.id_variedad})

    estado_map = {}
    if estado_ids:
        for ef in db.query(EstadoFenologico).filter(EstadoFenologico.id_estado.in_(estado_ids)).all():
            estado_map[ef.id_estado] = {
                "nombre": ef.nombre, "codigo": ef.codigo,
                "color_hex": ef.color_hex, "orden": ef.orden,
            }

    variedad_map = {}
    if variedad_ids:
        from app.models.variedades import Variedad
        for v in db.query(Variedad).filter(Variedad.id_variedad.in_(variedad_ids)).all():
            variedad_map[v.id_variedad] = {"nombre": v.nombre, "codigo": v.codigo}

    # Pivot: (variedad, estado) -> { actual: date, anterior: date }
    pivot: dict[tuple, dict] = {}
    for r in rows:
        key = (r.id_variedad, r.id_estado_fenol)
        if key not in pivot:
            pivot[key] = {"actual": None, "anterior": None}
        if r.temporada == temporada_actual:
            pivot[key]["actual"] = r.primera_fecha
        else:
            pivot[key]["anterior"] = r.primera_fecha

    # Build response grouped by variedad
    by_variedad: dict[int, list] = {}
    for (vid, eid), dates in pivot.items():
        if vid not in by_variedad:
            by_variedad[vid] = []
        diff_dias = None
        if dates["actual"] and dates["anterior"]:
            diff_dias = (dates["actual"] - dates["anterior"]).days
        by_variedad[vid].append({
            "estado": estado_map.get(eid, {}),
            "id_estado_fenol": eid,
            "fecha_actual": str(dates["actual"]) if dates["actual"] else None,
            "fecha_anterior": str(dates["anterior"]) if dates["anterior"] else None,
            "diferencia_dias": diff_dias,
        })

    # Sort estados by orden within each variedad
    result = []
    for vid, estados in by_variedad.items():
        estados.sort(key=lambda x: x.get("estado", {}).get("orden", 999))
        result.append({
            "variedad": variedad_map.get(vid, {"nombre": f"#{vid}", "codigo": ""}),
            "id_variedad": vid,
            "estados": estados,
        })
    result.sort(key=lambda x: x["variedad"].get("nombre", ""))

    return {
        "temporada_actual": temporada_actual,
        "temporada_anterior": temporada_anterior,
        "variedades": result,
    }


# ---------------------------------------------------------------------------
# Fenologia comparativa — Export Excel
# ---------------------------------------------------------------------------

@router.get("/fenologia-comparativa/export")
def fenologia_comparativa_export(
    temporada_actual: str = Query(..., description="Ej: 2025-2026"),
    temporada_anterior: str = Query(..., description="Ej: 2024-2025"),
    id_especie: int | None = Query(None),
    id_variedad: int | None = Query(None),
    id_testblock: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Export fenologia comparativa data as Excel file."""
    from sqlalchemy import func

    # Reuse the same query logic from fenologia_comparativa
    q_base = (
        db.query(
            RegistroFenologico.id_estado_fenol,
            RegistroFenologico.temporada,
            func.min(RegistroFenologico.fecha_registro).label("primera_fecha"),
            PosicionTestBlock.id_variedad,
        )
        .join(PosicionTestBlock, PosicionTestBlock.id_posicion == RegistroFenologico.id_posicion)
        .filter(
            RegistroFenologico.temporada.in_([temporada_actual, temporada_anterior]),
            PosicionTestBlock.id_variedad.isnot(None),
        )
    )

    if id_testblock:
        q_base = q_base.filter(PosicionTestBlock.id_testblock == id_testblock)
    if id_variedad:
        q_base = q_base.filter(PosicionTestBlock.id_variedad == id_variedad)
    if id_especie:
        from app.models.variedades import Variedad
        var_ids = [v.id_variedad for v in db.query(Variedad.id_variedad).filter(Variedad.id_especie == id_especie).all()]
        if var_ids:
            q_base = q_base.filter(PosicionTestBlock.id_variedad.in_(var_ids))
        else:
            # No varieties: return empty workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparativa"
            ws.append(["Sin datos para esta especie"])
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=comparativa_fenologica.xlsx"},
            )

    rows = (
        q_base
        .group_by(RegistroFenologico.id_estado_fenol, RegistroFenologico.temporada, PosicionTestBlock.id_variedad)
        .all()
    )

    # Build lookup maps
    estado_ids = list({r.id_estado_fenol for r in rows})
    variedad_ids = list({r.id_variedad for r in rows if r.id_variedad})

    estado_map = {}
    if estado_ids:
        for ef in db.query(EstadoFenologico).filter(EstadoFenologico.id_estado.in_(estado_ids)).all():
            estado_map[ef.id_estado] = {"nombre": ef.nombre, "codigo": ef.codigo, "orden": ef.orden}

    variedad_map = {}
    if variedad_ids:
        from app.models.variedades import Variedad
        for v in db.query(Variedad).filter(Variedad.id_variedad.in_(variedad_ids)).all():
            variedad_map[v.id_variedad] = {"nombre": v.nombre, "codigo": v.codigo}

    # Pivot data
    pivot: dict[tuple, dict] = {}
    for r in rows:
        key = (r.id_variedad, r.id_estado_fenol)
        if key not in pivot:
            pivot[key] = {"actual": None, "anterior": None}
        if r.temporada == temporada_actual:
            pivot[key]["actual"] = r.primera_fecha
        else:
            pivot[key]["anterior"] = r.primera_fecha

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparativa Fenologica"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Variedad", "Codigo", "Estado Fenologico", temporada_actual, temporada_anterior, "Diferencia (dias)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row_num = 2
    # Sort by variedad name, then estado order
    sorted_keys = sorted(
        pivot.keys(),
        key=lambda k: (variedad_map.get(k[0], {}).get("nombre", ""), estado_map.get(k[1], {}).get("orden", 999)),
    )

    for vid, eid in sorted_keys:
        var_info = variedad_map.get(vid, {"nombre": f"#{vid}", "codigo": ""})
        est_info = estado_map.get(eid, {"nombre": f"#{eid}"})
        dates = pivot[(vid, eid)]
        diff = None
        if dates["actual"] and dates["anterior"]:
            diff = (dates["actual"] - dates["anterior"]).days

        ws.cell(row=row_num, column=1, value=var_info["nombre"])
        ws.cell(row=row_num, column=2, value=var_info.get("codigo", ""))
        ws.cell(row=row_num, column=3, value=est_info["nombre"])
        ws.cell(row=row_num, column=4, value=str(dates["actual"]) if dates["actual"] else "-")
        ws.cell(row=row_num, column=5, value=str(dates["anterior"]) if dates["anterior"] else "-")
        ws.cell(row=row_num, column=6, value=diff if diff is not None else "-")
        row_num += 1

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=comparativa_fenologica.xlsx"},
    )


# ---------------------------------------------------------------------------
# Backfill — Propagate id_lote from positions to existing labores
# ---------------------------------------------------------------------------

@router.post("/backfill-lote")
def backfill_lote(
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Backfill id_lote on existing EjecucionLabor and RegistroFenologico
    records by resolving from PosicionTestBlock.id_lote.
    """
    # Backfill EjecucionLabor
    labores = (
        db.query(EjecucionLabor)
        .filter(
            EjecucionLabor.id_lote == None,
            EjecucionLabor.id_posicion != None,
        )
        .all()
    )
    updated_labores = 0
    for l in labores:
        pos = db.get(PosicionTestBlock, l.id_posicion)
        if pos and pos.id_lote:
            l.id_lote = pos.id_lote
            updated_labores += 1

    # Backfill RegistroFenologico
    registros = (
        db.query(RegistroFenologico)
        .filter(
            RegistroFenologico.id_lote == None,
            RegistroFenologico.id_posicion != None,
        )
        .all()
    )
    updated_fenol = 0
    for r in registros:
        pos = db.get(PosicionTestBlock, r.id_posicion)
        if pos:
            if pos.id_lote:
                r.id_lote = pos.id_lote
                updated_fenol += 1
            if pos.id_testblock and not r.id_testblock:
                r.id_testblock = pos.id_testblock

    db.commit()
    return {
        "labores_updated": updated_labores,
        "labores_total": len(labores),
        "fenologia_updated": updated_fenol,
        "fenologia_total": len(registros),
    }


# ---------------------------------------------------------------------------
# Fenologia — Seed demo data
# ---------------------------------------------------------------------------

@router.post("/fenologia/seed-demo", status_code=201)
def seed_fenologia_demo(
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
    _guard: bool = Depends(require_non_production),
):
    """Seed synthetic registros_fenologicos for 2 seasons (2024-2025 and 2025-2026).

    Creates realistic dates based on each estado's mes_orientativo for a sample
    of positions from existing testblocks. Admin only. EF-4: non-prod only.
    """
    from app.models.variedades import Variedad

    temporadas = ["2024-2025", "2025-2026"]

    # Get all active estados fenologicos
    estados = db.query(EstadoFenologico).filter(EstadoFenologico.activo == True).all()
    if not estados:
        raise HTTPException(status_code=400, detail="No hay estados fenologicos. Ejecute seed-estados-fenologicos primero.")

    # Group estados by especie
    estados_by_especie: dict[int, list] = {}
    for ef in estados:
        estados_by_especie.setdefault(ef.id_especie, []).append(ef)

    # Get variedades for each especie
    variedades_by_especie: dict[int, list[int]] = {}
    for vid, eid in db.query(Variedad.id_variedad, Variedad.id_especie).filter(Variedad.activo == True).all():
        variedades_by_especie.setdefault(eid, []).append(vid)

    # Get sample positions (up to 5 per variedad from active positions in testblocks)
    all_positions = (
        db.query(PosicionTestBlock)
        .filter(
            PosicionTestBlock.estado == "alta",
            PosicionTestBlock.id_variedad.isnot(None),
        )
        .all()
    )

    pos_by_variedad: dict[int, list] = {}
    for p in all_positions:
        pos_by_variedad.setdefault(p.id_variedad, []).append(p)

    created = 0

    for temporada in temporadas:
        # Parse temporada year: "2024-2025" -> season starts Jul 2024
        year_start = int(temporada.split("-")[0])

        for especie_id, especie_estados in estados_by_especie.items():
            especie_estados_sorted = sorted(especie_estados, key=lambda e: e.orden)
            variedad_ids = variedades_by_especie.get(especie_id, [])
            if not variedad_ids:
                continue

            for var_id in variedad_ids[:3]:  # Max 3 variedades per especie
                positions = pos_by_variedad.get(var_id, [])
                if not positions:
                    continue
                sample_positions = positions[:3]  # Max 3 positions per variedad

                for ef in especie_estados_sorted:
                    # Skip photo-only estados (no real date)
                    if "foto" in (ef.nombre or "").lower():
                        continue

                    # Compute a realistic date from mes_inicio or mes_orientativo
                    m_start = ef.mes_inicio
                    if m_start is None:
                        continue

                    # In southern hemisphere seasons: Jul-Dec belong to year_start,
                    # Jan-Jun belong to year_start+1
                    if m_start >= 7:
                        year = year_start
                    else:
                        year = year_start + 1

                    # Pick a day in the first half of the month + small random offset
                    day = random.randint(5, 20)
                    try:
                        base_date = date(year, m_start, day)
                    except ValueError:
                        base_date = date(year, m_start, 15)

                    # Add random jitter per temporada so seasons differ a bit
                    jitter = random.randint(-5, 10)
                    if temporada == "2025-2026":
                        jitter += random.randint(-3, 7)  # slightly different from anterior
                    registro_date = base_date + timedelta(days=jitter)

                    for pos in sample_positions:
                        # Check for duplicates
                        existing = (
                            db.query(RegistroFenologico)
                            .filter(
                                RegistroFenologico.id_posicion == pos.id_posicion,
                                RegistroFenologico.id_estado_fenol == ef.id_estado,
                                RegistroFenologico.temporada == temporada,
                            )
                            .first()
                        )
                        if existing:
                            continue

                        reg = RegistroFenologico(
                            id_posicion=pos.id_posicion,
                            id_planta=getattr(pos, "id_planta", None),
                            id_estado_fenol=ef.id_estado,
                            temporada=temporada,
                            fecha_registro=registro_date,
                            porcentaje=random.choice([50, 75, 100]),
                            observaciones=f"Dato demo - {ef.nombre}",
                            usuario_registro=user.username,
                        )
                        db.add(reg)
                        created += 1

    db.commit()
    return {"message": f"Se crearon {created} registros fenologicos demo.", "created": created}


# ---------------------------------------------------------------------------
# QR Code generation
# ---------------------------------------------------------------------------

@router.get("/ejecucion/{id}/qr")
def get_labor_qr(
    id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Generate a QR code PNG for a specific labor execution."""
    labor = crud.get_by_id(db, EjecucionLabor, id)
    tipo = db.query(TipoLabor).filter(TipoLabor.id_labor == labor.id_labor).first()

    qr_data = json.dumps({
        "type": "labor",
        "id": labor.id_ejecucion,
        "labor": tipo.nombre if tipo else str(labor.id_labor),
        "pos": labor.id_posicion,
        "fecha": str(labor.fecha_programada),
        "estado": labor.estado,
    })

    img = qrcode.make(qr_data)
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")
