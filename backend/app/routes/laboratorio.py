"""Laboratorio routes: mediciones, KPIs, clustering, bulk import."""

from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from app.core.utils import utcnow

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user, require_role
from app.models.sistema import Usuario
from app.models.laboratorio import MedicionLaboratorio, ClasificacionCluster
from app.models.testblock import TestBlock, PosicionTestBlock, Planta
from app.schemas.laboratorio import (
    MedicionCreate,
    MedicionBatchRequest,
    MedicionBatchResponse,
    MedicionBatchRowResult,
)
from app.services.laboratorio_service import crear_medicion, get_kpis

router = APIRouter(prefix="/laboratorio", tags=["Laboratorio"])


# ---------------------------------------------------------------------------
# Schemas para el endpoint de clasificacion manual
# ---------------------------------------------------------------------------

class ClasificarRequest(BaseModel):
    """Datos para clasificar una medicion manualmente o re-clasificar una existente."""
    # Metricas de calidad
    brix: Optional[float] = None
    acidez: Optional[float] = None
    # Firmeza de mejillas (2 lecturas)
    firmeza_mejilla_1: Optional[float] = None
    firmeza_mejilla_2: Optional[float] = None
    mejilla_1: Optional[float] = None  # alias
    mejilla_2: Optional[float] = None  # alias
    # Firmeza puntos debiles
    firmeza_punta: Optional[float] = None
    firmeza_quilla: Optional[float] = None
    firmeza_hombro: Optional[float] = None
    punta: Optional[float] = None  # alias
    quilla: Optional[float] = None  # alias
    hombro: Optional[float] = None  # alias
    # Firmeza generica (si no se tienen mejillas/puntos separados)
    firmeza: Optional[float] = None
    # Contexto para determinar regla
    especie: Optional[str] = None
    peso: Optional[float] = None
    color_pulpa: Optional[str] = None
    fecha_evaluacion: Optional[date] = None
    # Opcional: si se proporciona, persiste la clasificacion en BD
    id_medicion: Optional[int] = None


# ---------------------------------------------------------------------------
# Reglas de Cluster (configuracion de umbrales)
# ---------------------------------------------------------------------------

from app.models.regla_cluster import ReglaCluster


class ReglaClusterUpdate(BaseModel):
    """Campos editables de una regla de cluster."""
    nombre: Optional[str] = None
    brix_b1: Optional[float] = None
    brix_b2: Optional[float] = None
    brix_b3: Optional[float] = None
    mejillas_b1: Optional[float] = None
    mejillas_b2: Optional[float] = None
    mejillas_b3: Optional[float] = None
    punto_b1: Optional[float] = None
    punto_b2: Optional[float] = None
    punto_b3: Optional[float] = None
    acidez_b1: Optional[float] = None
    acidez_b2: Optional[float] = None
    acidez_b3: Optional[float] = None
    cluster1_max: Optional[int] = None
    cluster2_max: Optional[int] = None
    cluster3_max: Optional[int] = None
    notas: Optional[str] = None


@router.get("/reglas-cluster")
def list_reglas_cluster(
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Lista todas las reglas de clustering activas."""
    return (
        db.query(ReglaCluster)
        .filter(ReglaCluster.activo == True)
        .order_by(ReglaCluster.codigo_regla)
        .all()
    )


@router.get("/reglas-cluster/as-dict")
def reglas_as_dict(
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Retorna reglas en el mismo formato que el dict RULES hardcodeado.

    Util para que el motor de clustering pueda usar reglas de BD.
    """
    reglas = db.query(ReglaCluster).filter(ReglaCluster.activo == True).all()
    result = {}
    for r in reglas:
        result[r.codigo_regla] = {
            "brix": [float(r.brix_b1 or 0), float(r.brix_b2 or 0), float(r.brix_b3 or 0)],
            "mejillas": [float(r.mejillas_b1 or 0), float(r.mejillas_b2 or 0), float(r.mejillas_b3 or 0)],
            "punto": [float(r.punto_b1 or 0), float(r.punto_b2 or 0), float(r.punto_b3 or 0)],
            "acidez": [float(r.acidez_b1 or 0), float(r.acidez_b2 or 0), float(r.acidez_b3 or 0)],
        }
    return result


@router.post("/reglas-cluster/seed-from-hardcoded")
def seed_reglas_from_hardcoded(
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Pobla la tabla reglas_cluster desde el dict RULES hardcodeado.

    Solo crea reglas que no existan aun (por codigo_regla).
    """
    from app.services.clustering_service import RULES

    NAMES = {
        "ciruela_candy": "Ciruela Candy (>60g)",
        "ciruela_cherry": "Ciruela Cherry (<=60g)",
        "nectarina_amarilla_muy_temprana": "Nectarina Amarilla Muy Temprana",
        "nectarina_amarilla_temprana": "Nectarina Amarilla Temprana",
        "nectarina_amarilla_tardia": u"Nectarina Amarilla Tard\u00eda",
        "nectarina_blanca_muy_temprana": "Nectarina Blanca Muy Temprana",
        "nectarina_blanca_temprana": "Nectarina Blanca Temprana",
        "nectarina_blanca_tardia": u"Nectarina Blanca Tard\u00eda",
        "durazno_amarillo_muy_temprana": "Durazno Amarillo Muy Temprana",
        "durazno_amarillo_temprana": "Durazno Amarillo Temprana",
        "durazno_amarillo_tardia": u"Durazno Amarillo Tard\u00eda",
        "durazno_blanco_muy_temprana": "Durazno Blanco Muy Temprana",
        "durazno_blanco_temprana": "Durazno Blanco Temprana",
        "durazno_blanco_tardia": u"Durazno Blanco Tard\u00eda",
        "paraguayo_default": "Paraguayo",
        "platerina_default": "Platerina",
        "damasco_default": "Damasco",
    }

    created = 0
    for code, thresholds in RULES.items():
        existing = db.query(ReglaCluster).filter(ReglaCluster.codigo_regla == code).first()
        if existing:
            continue
        regla = ReglaCluster(
            codigo_regla=code,
            nombre=NAMES.get(code, code),
            brix_b1=thresholds["brix"][0],
            brix_b2=thresholds["brix"][1],
            brix_b3=thresholds["brix"][2],
            mejillas_b1=thresholds["mejillas"][0],
            mejillas_b2=thresholds["mejillas"][1],
            mejillas_b3=thresholds["mejillas"][2],
            punto_b1=thresholds["punto"][0],
            punto_b2=thresholds["punto"][1],
            punto_b3=thresholds["punto"][2],
            acidez_b1=thresholds["acidez"][0],
            acidez_b2=thresholds["acidez"][1],
            acidez_b3=thresholds["acidez"][2],
        )
        db.add(regla)
        created += 1
    db.commit()
    return {"created": created}


@router.get("/reglas-cluster/{regla_id}")
def get_regla_cluster(
    regla_id: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Obtiene una regla de cluster por ID."""
    regla = db.get(ReglaCluster, regla_id)
    if not regla:
        raise HTTPException(status_code=404, detail="Regla de cluster no encontrada")
    return regla


@router.put("/reglas-cluster/{regla_id}")
def update_regla_cluster(
    regla_id: int,
    data: ReglaClusterUpdate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Actualiza los umbrales de una regla de cluster."""
    regla = db.get(ReglaCluster, regla_id)
    if not regla:
        raise HTTPException(status_code=404, detail="Regla de cluster no encontrada")

    update_data = data.model_dump(exclude_none=True)
    for field, value in update_data.items():
        setattr(regla, field, value)

    regla.fecha_modificacion = datetime.utcnow()
    regla.usuario_modificacion = user.username
    db.commit()
    db.refresh(regla)

    # Invalidate clustering cache so new thresholds take effect immediately
    from app.services.clustering_service import invalidate_rules_cache
    invalidate_rules_cache()

    return regla


# ---------------------------------------------------------------------------
# Plantas para seleccion en laboratorio
# ---------------------------------------------------------------------------

@router.get("/plantas")
def list_plantas(
    testblock: int | None = Query(None),
    especie: int | None = Query(None),
    campo: int | None = Query(None),
    lote: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    q = db.query(Planta).filter(Planta.activa == True)
    if testblock:
        pos_ids = [p.id_posicion for p in db.query(PosicionTestBlock.id_posicion).filter(
            PosicionTestBlock.id_testblock == testblock
        ).all()]
        q = q.filter(Planta.id_posicion.in_(pos_ids))
    if campo:
        from app.models.maestras import Cuartel
        tb_ids = [t.id_testblock for t in db.query(TestBlock.id_testblock).filter(TestBlock.id_campo == campo).all()]
        if tb_ids:
            pos_ids_campo = [p.id_posicion for p in db.query(PosicionTestBlock.id_posicion).filter(
                PosicionTestBlock.id_testblock.in_(tb_ids)
            ).all()]
            q = q.filter(Planta.id_posicion.in_(pos_ids_campo))
        else:
            q = q.filter(False)
    if especie:
        q = q.filter(Planta.id_especie == especie)
    if lote:
        q = q.filter(Planta.id_lote_origen == lote)
    return q.all()


# ---------------------------------------------------------------------------
# CRUD mediciones
# ---------------------------------------------------------------------------

@router.post("/mediciones", status_code=201)
def create_medicion(
    data: MedicionCreate,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "laboratorio")),
):
    return crear_medicion(db, data, usuario=user.username)


@router.post("/mediciones/batch", status_code=201, response_model=MedicionBatchResponse)
def create_mediciones_batch(
    data: MedicionBatchRequest,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "laboratorio")),
):
    """Create multiple mediciones in a single request.

    Uses savepoints so individual row failures do NOT abort the batch
    and do NOT corrupt the session for subsequent rows.
    """
    resultados: list[MedicionBatchRowResult] = []
    total_creadas = 0

    for idx, medicion_data in enumerate(data.mediciones):
        sp = db.begin_nested()  # SAVEPOINT
        try:
            result = crear_medicion(db, medicion_data, usuario=user.username, auto_commit=False)

            clasificacion_dict = None
            if result.get("clasificacion"):
                c = result["clasificacion"]
                clasificacion_dict = {
                    "cluster": c.get("cluster"),
                    "cluster_label": c.get("cluster_label", ""),
                    "banda_brix": c.get("banda_brix"),
                    "banda_firmeza": c.get("banda_firmeza"),
                    "banda_acidez": c.get("banda_acidez"),
                    "banda_firmeza_punto": c.get("banda_firmeza_punto"),
                    "score_total": c.get("score_total"),
                }

            sp.commit()  # release savepoint

            # Serialize the medicion ORM object to a plain dict
            medicion_obj = result["medicion"]
            medicion_dict = {
                "id_medicion": medicion_obj.id_medicion,
                "id_planta": medicion_obj.id_planta,
                "id_posicion": medicion_obj.id_posicion,
                "temporada": medicion_obj.temporada,
                "fecha_medicion": str(medicion_obj.fecha_medicion) if medicion_obj.fecha_medicion else None,
                "brix": float(medicion_obj.brix) if medicion_obj.brix is not None else None,
                "acidez": float(medicion_obj.acidez) if medicion_obj.acidez is not None else None,
                "firmeza": float(medicion_obj.firmeza) if medicion_obj.firmeza is not None else None,
                "peso": float(medicion_obj.peso) if medicion_obj.peso is not None else None,
            }

            resultados.append(MedicionBatchRowResult(
                index=idx,
                success=True,
                medicion=medicion_dict,
                clasificacion=clasificacion_dict,
            ))
            total_creadas += 1
        except Exception as e:
            sp.rollback()  # rollback only this row's savepoint
            resultados.append(MedicionBatchRowResult(
                index=idx,
                success=False,
                error=str(e),
            ))

    # Commit all successful rows in one transaction
    db.commit()

    return MedicionBatchResponse(
        total_enviadas=len(data.mediciones),
        total_creadas=total_creadas,
        total_errores=len(data.mediciones) - total_creadas,
        resultados=resultados,
    )


@router.get("/planta/{id_planta}/mediciones")
def mediciones_por_planta(
    id_planta: int,
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Return all mediciones for a specific plant, ordered by fecha_medicion DESC.

    Includes classification cluster data when available.
    """
    mediciones = (
        db.query(MedicionLaboratorio)
        .filter(MedicionLaboratorio.id_planta == id_planta)
        .order_by(MedicionLaboratorio.fecha_medicion.desc())
        .all()
    )

    # Enrich with cluster data
    med_ids = [m.id_medicion for m in mediciones]
    cluster_map: dict[int, ClasificacionCluster] = {}
    if med_ids:
        clusters = (
            db.query(ClasificacionCluster)
            .filter(ClasificacionCluster.id_medicion.in_(med_ids))
            .all()
        )
        for c in clusters:
            cluster_map[c.id_medicion] = c

    result = []
    for m in mediciones:
        data = _serialize_medicion(m)
        cc = cluster_map.get(m.id_medicion)
        data["cluster"] = cc.cluster if cc else None
        data["cluster_label"] = f"C{cc.cluster}" if cc and cc.cluster else None
        result.append(data)

    return result


def _serialize_medicion(m: MedicionLaboratorio) -> dict:
    """Convert a MedicionLaboratorio ORM object to a plain dict with Decimal->float."""
    data = {c: getattr(m, c) for c in m.__class__.model_fields}
    # Convert Decimal fields to float for JSON serialization
    _decimal_fields = (
        "brix", "acidez", "firmeza", "calibre", "peso", "perimetro",
        "firmeza_punta", "firmeza_quilla", "firmeza_hombro",
        "firmeza_mejilla_1", "firmeza_mejilla_2",
        "pardeamiento", "traslucidez", "gelificacion", "harinosidad",
        "rendimiento",
    )
    for key in _decimal_fields:
        if data.get(key) is not None:
            data[key] = float(data[key])
    return data


def _apply_medicion_filters(
    q,
    *,
    testblock: int | None = None,
    temporada: str | None = None,
    especie: int | None = None,
    campo: int | None = None,
    variedad: int | None = None,
    pmg: int | None = None,
    fecha_cosecha_desde: str | None = None,
    fecha_cosecha_hasta: str | None = None,
    tipo_evaluacion: str | None = None,
    periodo_almacenaje_min: int | None = None,
    periodo_almacenaje_max: int | None = None,
    db: Session | None = None,
):
    """Apply all medicion filters to a query. Returns the filtered query."""
    if testblock:
        pos_ids = [p.id_posicion for p in db.query(PosicionTestBlock.id_posicion).filter(
            PosicionTestBlock.id_testblock == testblock
        ).all()]
        q = q.filter(MedicionLaboratorio.id_posicion.in_(pos_ids))
    if temporada:
        q = q.filter(MedicionLaboratorio.temporada == temporada)
    if especie:
        q = q.filter(MedicionLaboratorio.id_especie == especie)
    if campo:
        q = q.filter(MedicionLaboratorio.id_campo == campo)
    if variedad:
        q = q.filter(MedicionLaboratorio.id_variedad == variedad)
    if pmg:
        from app.models.variedades import Variedad
        var_ids = [v.id_variedad for v in db.query(Variedad.id_variedad).filter(Variedad.id_pmg == pmg).all()]
        if var_ids:
            q = q.filter(MedicionLaboratorio.id_variedad.in_(var_ids))
        else:
            q = q.filter(False)
    if fecha_cosecha_desde:
        from datetime import date as date_type
        q = q.filter(MedicionLaboratorio.fecha_cosecha >= date_type.fromisoformat(fecha_cosecha_desde))
    if fecha_cosecha_hasta:
        from datetime import date as date_type
        q = q.filter(MedicionLaboratorio.fecha_cosecha <= date_type.fromisoformat(fecha_cosecha_hasta))
    if tipo_evaluacion:
        if tipo_evaluacion == "poscosecha":
            q = q.filter(MedicionLaboratorio.periodo_almacenaje != None)
        else:
            q = q.filter(
                (MedicionLaboratorio.periodo_almacenaje == None) | (MedicionLaboratorio.periodo_almacenaje == 0)
            )
    if periodo_almacenaje_min is not None:
        q = q.filter(MedicionLaboratorio.periodo_almacenaje >= periodo_almacenaje_min)
    if periodo_almacenaje_max is not None:
        q = q.filter(MedicionLaboratorio.periodo_almacenaje <= periodo_almacenaje_max)
    return q


@router.get("/mediciones")
def list_mediciones(
    testblock: int | None = Query(None),
    temporada: str | None = Query(None),
    especie: int | None = Query(None),
    campo: int | None = Query(None),
    variedad: int | None = Query(None),
    pmg: int | None = Query(None),
    fecha_cosecha_desde: str | None = Query(None, description="Fecha cosecha desde (YYYY-MM-DD)"),
    fecha_cosecha_hasta: str | None = Query(None, description="Fecha cosecha hasta (YYYY-MM-DD)"),
    tipo_evaluacion: str | None = Query(None, description="'laboratorio' o 'poscosecha'"),
    periodo_almacenaje_min: int | None = Query(None, description="Periodo almacenaje minimo (dias)"),
    periodo_almacenaje_max: int | None = Query(None, description="Periodo almacenaje maximo (dias)"),
    skip: int = Query(0, ge=0, description="Offset for pagination"),
    limit: int = Query(5000, ge=1, le=50000, description="Max records to return"),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """List mediciones with filters and pagination.

    Returns a dict with ``data`` (list of mediciones) and ``total`` (count
    before pagination) so the frontend can implement server-side pagination.
    """
    q = db.query(MedicionLaboratorio)
    q = _apply_medicion_filters(
        q,
        testblock=testblock, temporada=temporada, especie=especie,
        campo=campo, variedad=variedad, pmg=pmg,
        fecha_cosecha_desde=fecha_cosecha_desde,
        fecha_cosecha_hasta=fecha_cosecha_hasta,
        tipo_evaluacion=tipo_evaluacion,
        periodo_almacenaje_min=periodo_almacenaje_min,
        periodo_almacenaje_max=periodo_almacenaje_max,
        db=db,
    )
    total = q.count()
    rows = q.order_by(MedicionLaboratorio.fecha_medicion.desc()).offset(skip).limit(limit).all()
    return {
        "data": [_serialize_medicion(m) for m in rows],
        "total": total,
        "skip": skip,
        "limit": limit,
    }


# ---------------------------------------------------------------------------
# KPIs
# ---------------------------------------------------------------------------

@router.get("/kpis")
def api_kpis(
    testblock: int | None = Query(None),
    temporada: str | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    return get_kpis(db, testblock_id=testblock, temporada=temporada)


# ---------------------------------------------------------------------------
# Clustering endpoints
# ---------------------------------------------------------------------------

@router.post("/clasificar")
def clasificar_medicion_endpoint(
    data: ClasificarRequest,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "laboratorio")),
):
    """Clasifica una medicion manualmente o re-clasifica una existente.

    Aplica el algoritmo Band-Sum del motor de clustering:
    1. Calcula mejillas promedio y punto debil
    2. Determina la regla de umbrales segun especie/peso/color/fecha
    3. Clasifica cada metrica en banda 1-4
    4. Suma bandas y asigna cluster (C1-C4)

    Si se proporciona id_medicion, persiste la clasificacion en la tabla
    clasificacion_cluster. Si ya existe una clasificacion para esa medicion,
    la actualiza.
    """
    from app.services.clustering_service import (
        clasificar_medicion as clasificar_bandas,
        determinar_regla,
        calcular_mejillas_promedio,
        calcular_punto_debil,
    )

    brix = data.brix or 0.0
    acidez = data.acidez or 0.0

    # Calcular promedio de firmeza de mejillas
    mejilla1 = data.firmeza_mejilla_1 or data.mejilla_1 or 0.0
    mejilla2 = data.firmeza_mejilla_2 or data.mejilla_2 or 0.0
    mejillas = calcular_mejillas_promedio(mejilla1, mejilla2)

    # Si no se proporcionaron mejillas separadas, usar firmeza generica
    if mejillas == 0.0 and data.firmeza:
        mejillas = data.firmeza

    # Calcular punto debil (minimo de punta, quilla, hombro)
    punta = data.firmeza_punta or data.punta or 0.0
    quilla = data.firmeza_quilla or data.quilla or 0.0
    hombro = data.firmeza_hombro or data.hombro or 0.0
    punto_debil = calcular_punto_debil(punta, quilla, hombro)

    # Si no se proporcionaron puntos separados, usar firmeza generica
    if punto_debil == 0.0 and data.firmeza:
        punto_debil = data.firmeza

    # Determinar regla de umbrales
    regla = determinar_regla(
        especie=data.especie or "",
        peso_promedio=data.peso if data.peso else None,
        color_pulpa=data.color_pulpa,
        fecha_evaluacion=data.fecha_evaluacion,
    )

    # Clasificar
    result = clasificar_bandas(
        brix=brix if brix > 0 else None,
        acidez=acidez if acidez > 0 else None,
        firmeza_mejillas=mejillas if mejillas > 0 else None,
        firmeza_punto_debil=punto_debil if punto_debil > 0 else None,
        regla=regla,
    )

    # Si se proporciono id_medicion, persistir en BD
    if data.id_medicion:
        existing = db.query(ClasificacionCluster).filter(
            ClasificacionCluster.id_medicion == data.id_medicion
        ).first()

        if existing:
            existing.cluster = result["cluster"]
            existing.banda_brix = result["banda_brix"]
            existing.banda_firmeza = result["banda_firmeza"]
            existing.banda_acidez = result["banda_acidez"]
            existing.banda_calibre = result["banda_firmeza_punto"]
            existing.score_total = Decimal(str(result["suma_bandas"]))
            existing.metodo = "reglas_v2"
            existing.fecha_calculo = utcnow()
        else:
            cc = ClasificacionCluster(
                id_medicion=data.id_medicion,
                cluster=result["cluster"],
                banda_brix=result["banda_brix"],
                banda_firmeza=result["banda_firmeza"],
                banda_acidez=result["banda_acidez"],
                banda_calibre=result["banda_firmeza_punto"],
                score_total=Decimal(str(result["suma_bandas"])),
                metodo="reglas_v2",
            )
            db.add(cc)

        db.commit()

    return result


@router.get("/clustering-rules")
def get_clustering_rules(
    user: Usuario = Depends(get_current_user),
):
    """Retorna las reglas de clustering disponibles y los rangos de cada cluster.

    Util para el frontend para mostrar los umbrales y permitir
    al usuario entender la clasificacion.
    """
    from app.services.clustering_service import RULES, CLUSTER_RANGES

    return {
        "rules": {k: v for k, v in RULES.items()},
        "cluster_ranges": CLUSTER_RANGES,
    }


@router.get("/clustering-rule-preview")
def clustering_rule_preview(
    especie: str = Query(..., description="Nombre de la especie"),
    peso: float | None = Query(None),
    color_pulpa: str | None = Query(None),
    fecha: str | None = Query(None),
    user: Usuario = Depends(get_current_user),
):
    """Preview: que regla de clustering se aplicaria y sus umbrales."""
    from app.services.clustering_service import determinar_regla, RULES, CLUSTER_RANGES
    from datetime import date as date_type

    fecha_eval = None
    if fecha:
        try:
            fecha_eval = date_type.fromisoformat(fecha)
        except ValueError:
            pass

    rule_key = determinar_regla(especie, peso, color_pulpa, fecha_eval)
    thresholds = RULES.get(rule_key, {})

    # Format thresholds as readable bands
    bands_display = {}
    for metric, vals in thresholds.items():
        if metric == "acidez":
            bands_display[metric] = {
                "B1": f"<= {vals[0]}",
                "B2": f"<= {vals[1]}",
                "B3": f"<= {vals[2]}",
                "B4": f"> {vals[2]}",
            }
        else:
            bands_display[metric] = {
                "B1": f">= {vals[0]}",
                "B2": f">= {vals[1]}",
                "B3": f">= {vals[2]}",
                "B4": f"< {vals[2]}",
            }

    return {
        "regla": rule_key,
        "regla_label": rule_key.replace("_", " ").title(),
        "umbrales": thresholds,
        "bandas": bands_display,
        "cluster_ranges": CLUSTER_RANGES,
    }


# ---------------------------------------------------------------------------
# Analysis endpoints
# ---------------------------------------------------------------------------

@router.get("/analisis/resumen-variedades")
def analisis_resumen_variedades(
    especie: int | None = Query(None),
    temporada: str | None = Query(None),
    portainjerto: int | None = Query(None),
    pmg: int | None = Query(None),
    campo: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Returns per unique combination (variedad+portainjerto+PMG+campo) summary.

    Each row represents a unique evaluation unit: the same variety grafted on
    different rootstocks or from different breeding programs behaves differently.
    The campo dimension is included because field conditions affect performance.
    """
    from sqlalchemy import func, case, cast, Float
    from app.models.variedades import Variedad
    from app.models.maestras import Especie, Portainjerto, Pmg, Campo

    q = (
        db.query(
            MedicionLaboratorio.id_variedad,
            MedicionLaboratorio.id_portainjerto,
            MedicionLaboratorio.id_campo,
            Variedad.nombre.label("variedad"),
            Variedad.id_pmg,
            Especie.nombre.label("especie"),
            Portainjerto.nombre.label("portainjerto"),
            Pmg.nombre.label("pmg"),
            Campo.nombre.label("campo"),
            func.count().label("total_mediciones"),
            func.count(func.distinct(MedicionLaboratorio.temporada)).label("n_temporadas"),
            func.avg(cast(MedicionLaboratorio.brix, Float)).label("brix_avg"),
            func.min(cast(MedicionLaboratorio.brix, Float)).label("brix_min"),
            func.max(cast(MedicionLaboratorio.brix, Float)).label("brix_max"),
            func.avg(cast(MedicionLaboratorio.firmeza, Float)).label("firmeza_avg"),
            func.avg(cast(MedicionLaboratorio.acidez, Float)).label("acidez_avg"),
            func.avg(cast(MedicionLaboratorio.peso, Float)).label("peso_avg"),
            func.avg(cast(MedicionLaboratorio.calibre, Float)).label("calibre_avg"),
            # Cluster distribution
            func.sum(case((ClasificacionCluster.cluster == 1, 1), else_=0)).label("c1"),
            func.sum(case((ClasificacionCluster.cluster == 2, 1), else_=0)).label("c2"),
            func.sum(case((ClasificacionCluster.cluster == 3, 1), else_=0)).label("c3"),
            func.sum(case((ClasificacionCluster.cluster == 4, 1), else_=0)).label("c4"),
        )
        .outerjoin(ClasificacionCluster, ClasificacionCluster.id_medicion == MedicionLaboratorio.id_medicion)
        .outerjoin(Variedad, Variedad.id_variedad == MedicionLaboratorio.id_variedad)
        .outerjoin(Especie, Especie.id_especie == MedicionLaboratorio.id_especie)
        .outerjoin(Portainjerto, Portainjerto.id_portainjerto == MedicionLaboratorio.id_portainjerto)
        .outerjoin(Pmg, Pmg.id_pmg == Variedad.id_pmg)
        .outerjoin(Campo, Campo.id_campo == MedicionLaboratorio.id_campo)
        .filter(MedicionLaboratorio.id_variedad.isnot(None))
    )

    if especie:
        q = q.filter(MedicionLaboratorio.id_especie == especie)
    if temporada:
        q = q.filter(MedicionLaboratorio.temporada == temporada)
    if portainjerto:
        q = q.filter(MedicionLaboratorio.id_portainjerto == portainjerto)
    if pmg:
        q = q.filter(Variedad.id_pmg == pmg)
    if campo:
        q = q.filter(MedicionLaboratorio.id_campo == campo)

    q = q.group_by(
        MedicionLaboratorio.id_variedad, MedicionLaboratorio.id_portainjerto,
        MedicionLaboratorio.id_campo,
        Variedad.nombre, Variedad.id_pmg, Especie.nombre,
        Portainjerto.nombre, Pmg.nombre, Campo.nombre,
    ).order_by(func.count().desc())

    results = q.all()

    def safe_round(val, decimals=2):
        return round(float(val), decimals) if val is not None else None

    return [
        {
            "id_variedad": r.id_variedad,
            "id_portainjerto": r.id_portainjerto,
            "id_campo": r.id_campo,
            "id_pmg": r.id_pmg,
            "variedad": r.variedad,
            "especie": r.especie,
            "portainjerto": r.portainjerto or "-",
            "pmg": r.pmg or "-",
            "campo": r.campo or "-",
            "total_mediciones": r.total_mediciones,
            "n_temporadas": r.n_temporadas,
            "brix_avg": safe_round(r.brix_avg, 1),
            "brix_min": safe_round(r.brix_min, 1),
            "brix_max": safe_round(r.brix_max, 1),
            "firmeza_avg": safe_round(r.firmeza_avg, 1),
            "acidez_avg": safe_round(r.acidez_avg, 2),
            "peso_avg": safe_round(r.peso_avg, 1),
            "calibre_avg": safe_round(r.calibre_avg, 1),
            "cluster_dist": {"c1": r.c1 or 0, "c2": r.c2 or 0, "c3": r.c3 or 0, "c4": r.c4 or 0},
            "cluster_predominante": max(
                [(r.c1 or 0, 1), (r.c2 or 0, 2), (r.c3 or 0, 3), (r.c4 or 0, 4)],
                key=lambda x: x[0]
            )[1] if (r.c1 or r.c2 or r.c3 or r.c4) else None,
        }
        for r in results
    ]


@router.get("/analisis/evolucion")
def analisis_evolucion(
    id_variedad: int = Query(...),
    id_portainjerto: int | None = Query(None),
    campo: int | None = Query(None),
    db: Session = Depends(get_db),
    user: Usuario = Depends(get_current_user),
):
    """Returns per-temporada AND per-fecha evolution for a variedad+portainjerto combo.

    Returns two datasets:
    - por_temporada: aggregated per season (for line charts)
    - por_fecha: individual measurements ordered by date (for scatter/timeline)
      Each individual measurement includes campo context.
    """
    from sqlalchemy import func, case, cast, Float
    from app.models.variedades import Variedad
    from app.models.maestras import Especie, Portainjerto, Pmg, Campo

    # Base filter
    base_filter = MedicionLaboratorio.id_variedad == id_variedad
    if id_portainjerto:
        base_filter = base_filter & (MedicionLaboratorio.id_portainjerto == id_portainjerto)
    if campo:
        base_filter = base_filter & (MedicionLaboratorio.id_campo == campo)

    # --- Per temporada ---
    q_temp = (
        db.query(
            MedicionLaboratorio.temporada,
            func.count().label("total"),
            func.avg(cast(MedicionLaboratorio.brix, Float)).label("brix_avg"),
            func.avg(cast(MedicionLaboratorio.firmeza, Float)).label("firmeza_avg"),
            func.avg(cast(MedicionLaboratorio.acidez, Float)).label("acidez_avg"),
            func.avg(cast(MedicionLaboratorio.peso, Float)).label("peso_avg"),
            func.sum(case((ClasificacionCluster.cluster == 1, 1), else_=0)).label("c1"),
            func.sum(case((ClasificacionCluster.cluster == 2, 1), else_=0)).label("c2"),
            func.sum(case((ClasificacionCluster.cluster == 3, 1), else_=0)).label("c3"),
            func.sum(case((ClasificacionCluster.cluster == 4, 1), else_=0)).label("c4"),
        )
        .outerjoin(ClasificacionCluster, ClasificacionCluster.id_medicion == MedicionLaboratorio.id_medicion)
        .filter(base_filter)
        .group_by(MedicionLaboratorio.temporada)
        .order_by(MedicionLaboratorio.temporada)
    )

    # --- Individual measurements (latest 500) with campo name ---
    q_ind = (
        db.query(
            MedicionLaboratorio.id_medicion,
            MedicionLaboratorio.fecha_medicion,
            MedicionLaboratorio.temporada,
            MedicionLaboratorio.brix,
            MedicionLaboratorio.firmeza,
            MedicionLaboratorio.acidez,
            MedicionLaboratorio.peso,
            MedicionLaboratorio.calibre,
            MedicionLaboratorio.n_muestra,
            MedicionLaboratorio.id_campo,
            Campo.nombre.label("campo"),
            ClasificacionCluster.cluster,
        )
        .outerjoin(ClasificacionCluster, ClasificacionCluster.id_medicion == MedicionLaboratorio.id_medicion)
        .outerjoin(Campo, Campo.id_campo == MedicionLaboratorio.id_campo)
        .filter(base_filter)
        .order_by(MedicionLaboratorio.fecha_medicion, MedicionLaboratorio.id_medicion)
        .limit(500)
    )

    # --- Variety context ---
    variedad = db.query(Variedad).filter(Variedad.id_variedad == id_variedad).first()
    especie = db.query(Especie).filter(Especie.id_especie == variedad.id_especie).first() if variedad else None
    pi = db.query(Portainjerto).filter(Portainjerto.id_portainjerto == id_portainjerto).first() if id_portainjerto else None
    pmg_obj = db.query(Pmg).filter(Pmg.id_pmg == variedad.id_pmg).first() if variedad and variedad.id_pmg else None

    # Campo context when filtering by a specific campo
    campo_obj = db.query(Campo).filter(Campo.id_campo == campo).first() if campo else None

    def safe_round(val, decimals=2):
        return round(float(val), decimals) if val is not None else None

    return {
        "contexto": {
            "variedad": variedad.nombre if variedad else "?",
            "especie": especie.nombre if especie else "?",
            "portainjerto": pi.nombre if pi else "-",
            "pmg": pmg_obj.nombre if pmg_obj else "-",
            "campo": campo_obj.nombre if campo_obj else None,
        },
        "por_temporada": [
            {
                "temporada": r.temporada,
                "total": r.total,
                "brix_avg": safe_round(r.brix_avg, 1),
                "firmeza_avg": safe_round(r.firmeza_avg, 1),
                "acidez_avg": safe_round(r.acidez_avg, 2),
                "peso_avg": safe_round(r.peso_avg, 1),
                "cluster_dist": {"c1": r.c1 or 0, "c2": r.c2 or 0, "c3": r.c3 or 0, "c4": r.c4 or 0},
            }
            for r in q_temp.all()
        ],
        "por_fecha": [
            {
                "id": r.id_medicion,
                "fecha": str(r.fecha_medicion) if r.fecha_medicion else None,
                "temporada": r.temporada,
                "brix": safe_round(r.brix, 1),
                "firmeza": safe_round(r.firmeza, 1),
                "acidez": safe_round(r.acidez, 2),
                "peso": safe_round(r.peso, 1),
                "calibre": safe_round(r.calibre, 1),
                "n_muestra": r.n_muestra,
                "id_campo": r.id_campo,
                "campo": r.campo or "-",
                "cluster": r.cluster,
            }
            for r in q_ind.all()
        ],
    }


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------

@router.post("/bulk-import")
async def bulk_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin", "laboratorio")),
):
    """Import mediciones from Excel file."""
    import openpyxl
    from io import BytesIO
    from datetime import date as date_type

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    created = 0
    errors = []
    headers = [cell.value for cell in ws[1]]

    def _safe_int(val):
        if val is None:
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    def _safe_decimal(val):
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    clasificaciones = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sp = db.begin_nested()
        try:
            row_data = dict(zip(headers, row))
            medicion_data = MedicionCreate(
                id_posicion=_safe_int(row_data.get("id_posicion")),
                id_planta=_safe_int(row_data.get("id_planta")),
                temporada=str(row_data.get("temporada", "") or ""),
                fecha_medicion=row_data.get("fecha_medicion") or date_type.today(),
                fecha_cosecha=row_data.get("fecha_cosecha"),
                brix=_safe_decimal(row_data.get("brix")),
                acidez=_safe_decimal(row_data.get("acidez")),
                firmeza=_safe_decimal(row_data.get("firmeza")),
                calibre=_safe_decimal(row_data.get("calibre")),
                peso=_safe_decimal(row_data.get("peso")),
                color_pct=_safe_int(row_data.get("color_pct")),
                cracking_pct=_safe_int(row_data.get("cracking_pct")),
                observaciones=str(row_data.get("observaciones", "") or ""),
                # Firmeza detallada
                firmeza_punta=_safe_decimal(row_data.get("firmeza_punta")),
                firmeza_quilla=_safe_decimal(row_data.get("firmeza_quilla")),
                firmeza_hombro=_safe_decimal(row_data.get("firmeza_hombro")),
                firmeza_mejilla_1=_safe_decimal(row_data.get("firmeza_mejilla_1")),
                firmeza_mejilla_2=_safe_decimal(row_data.get("firmeza_mejilla_2")),
                # Muestra y postcosecha
                n_muestra=_safe_int(row_data.get("n_muestra")),
                periodo_almacenaje=_safe_int(row_data.get("periodo_almacenaje")),
                perimetro=_safe_decimal(row_data.get("perimetro")),
                pardeamiento=_safe_decimal(row_data.get("pardeamiento")),
                traslucidez=_safe_decimal(row_data.get("traslucidez")),
                gelificacion=_safe_decimal(row_data.get("gelificacion")),
                harinosidad=_safe_decimal(row_data.get("harinosidad")),
                color_pulpa=str(row_data["color_pulpa"]) if row_data.get("color_pulpa") else None,
                # Agronomia
                raleo_frutos=_safe_int(row_data.get("raleo_frutos")),
                rendimiento=_safe_decimal(row_data.get("rendimiento")),
                repeticion=_safe_int(row_data.get("repeticion")),
                # Color cubrimiento
                color_0_30=_safe_int(row_data.get("color_0_30")),
                color_30_50=_safe_int(row_data.get("color_30_50")),
                color_50_75=_safe_int(row_data.get("color_50_75")),
                color_75_100=_safe_int(row_data.get("color_75_100")),
                color_total=_safe_int(row_data.get("color_total")),
                # Distribucion de color
                color_verde=_safe_int(row_data.get("color_verde")),
                color_crema=_safe_int(row_data.get("color_crema")),
                color_amarillo=_safe_int(row_data.get("color_amarillo")),
                color_full=_safe_int(row_data.get("color_full")),
                color_dist_total=_safe_int(row_data.get("color_dist_total")),
                # Total frutos postcosecha
                total_frutos_pardeamiento=_safe_int(row_data.get("total_frutos_pardeamiento")),
                total_frutos_traslucidez=_safe_int(row_data.get("total_frutos_traslucidez")),
                total_frutos_gelificacion=_safe_int(row_data.get("total_frutos_gelificacion")),
                total_frutos_harinosidad=_safe_int(row_data.get("total_frutos_harinosidad")),
                # FKs directas
                id_campo=_safe_int(row_data.get("id_campo")),
                id_variedad=_safe_int(row_data.get("id_variedad")),
                id_especie=_safe_int(row_data.get("id_especie")),
                id_portainjerto=_safe_int(row_data.get("id_portainjerto")),
            )
            result = crear_medicion(db, medicion_data, usuario=user.username, auto_commit=False)
            sp.commit()
            created += 1
            if result.get("clasificacion"):
                clasificaciones.append({"row": row_idx, **result["clasificacion"]})
        except Exception as e:
            sp.rollback()
            errors.append({"row": row_idx, "error": str(e)})

    db.commit()
    return {"created": created, "errors": errors, "clasificaciones": clasificaciones}
