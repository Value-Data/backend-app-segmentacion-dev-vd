"""Migración completa: estructura real + re-vinculación de mediciones desde Excel maestro."""

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import require_role
from app.core.utils import utcnow
from app.models.sistema import Usuario
from app.models.maestras import Campo, Cuartel, Especie, Portainjerto, Pmg
from app.models.variedades import Variedad
from app.models.inventario import InventarioVivero, MovimientoInventario
from app.models.testblock import TestBlock, PosicionTestBlock, Planta, HistorialPosicion
from app.models.laboratorio import MedicionLaboratorio, ClasificacionCluster
from app.services.clustering_service import (
    clasificar_medicion as clasificar_bandas,
    determinar_regla,
    calcular_mejillas_promedio,
    calcular_punto_debil,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/seed", tags=["Seed Migración"])

_MC_SHEET = "CAROZOS"
_MC_DATA_START = 4
_DUMMY_TB_CODE = "TB-IMPORT-CAROZOS"
BATCH_SIZE = 500

LOC_NORMALIZE = {
    "Estación": "La Estación", "Parque": "El Parque", "Retorno": "El Retorno",
    "Requinoa": "Vivero Requinoa", "Plantel Madre ": "Plantel Madre",
}

ROOTSTOCK_SUFFIXES = [
    "NOGA GXN", "NOGA NEMA", "Noga GXN", "GXN", "GxN",
    "NEMA", "Nema", "NOGA", "Noga", "H41", "H43",
]

COL = {
    "especie": 0, "variedad": 1, "pmg": 2, "localidad": 3, "temporada": 4,
    "fecha_cosecha": 5, "fecha_evaluacion": 6, "periodo_almacenaje": 7,
    "raleo": 8, "planta": 9, "perimetro": 10, "rendimiento": 11, "repeticion": 12,
    "cubr_0_30": 13, "cubr_30_50": 14, "cubr_50_75": 15, "cubr_75_100": 16, "cubr_total": 17,
    "color_verde": 18, "color_crema": 19, "color_amarillo": 20, "color_full": 21, "color_total": 22,
    "color_pulpa": 23, "fruto_num": 24, "peso_g": 25,
    "firmeza_punta": 26, "firmeza_quilla": 27, "firmeza_hombro": 28,
    "firmeza_mejilla1": 29, "firmeza_mejilla2": 30,
    "brix": 31, "acidez": 32,
    "pardeamiento": 33, "pardeamiento_total": 34,
    "traslucidez": 35, "traslucidez_total": 36,
    "gelificacion": 37, "gelificacion_total": 38,
    "harinosidad": 39, "harinosidad_total": 40,
    "observaciones": 41,
}


def _safe_decimal(value, precision=2):
    if value is None:
        return None
    try:
        return round(Decimal(str(value)), precision)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _safe_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_int(value):
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _normalize_loc(loc):
    loc = loc.strip()
    return LOC_NORMALIZE.get(loc, loc)


def _extract_rootstock(var_name):
    name = str(var_name).strip()
    name = re.sub(r'\s+[Ll][Oo][Tt][Ee]\s+\d+', '', name)
    name = re.sub(r'\s+\d+F$', '', name)
    for suf in ['PACKING', 'Comercial', 'Plantel']:
        name = name.replace(f' {suf}', '')
    rootstock = 'SIN_PI'
    for rs in ROOTSTOCK_SUFFIXES:
        if f' {rs}' in name:
            rootstock = rs.upper()
            name = name.replace(f' {rs}', '').strip()
            break
    name = re.sub(r'^([A-Z])\s+(\d)', r'\1\2', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name, rootstock


def _loc_to_campo_code(loc_name):
    code = loc_name.upper().replace(" ", "-").replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    code = re.sub(r'[^A-Z0-9\-]', '', code)
    return f"CAM-{code[:15]}"


def _gen_lote_code(esp_code, var_code, pi_code, seq):
    pi_short = {
        'GXN': 'GXN', 'NEMA': 'NEMA', 'NOGA': 'NOGA',
        'NOGA GXN': 'NGX', 'NOGA NEMA': 'NNM',
        'H41': 'H41', 'H43': 'H43', 'SIN_PI': 'SPI',
    }.get(pi_code, pi_code[:4])
    return f"LOT-{esp_code}-{var_code}-{pi_short}-{seq:03d}"


@router.post("/migracion-maestro-carozos")
def migracion_maestro_carozos(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Migración completa: crea estructura real y re-importa mediciones.
    Elimina mediciones dummy de TB-IMPORT-CAROZOS y las reemplaza con links correctos.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo .xlsx/.xls")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error leyendo Excel: {str(exc)[:200]}")

    if _MC_SHEET not in wb.sheetnames:
        raise HTTPException(status_code=400, detail=f"Hoja '{_MC_SHEET}' no encontrada. Hojas: {wb.sheetnames}")

    ws = wb[_MC_SHEET]
    usuario = user.username

    # ── PASO 1: Leer y agrupar datos ─────────────────────────────────
    logger.info("Paso 1: Leyendo Excel...")

    especie_cache = {e.nombre.lower().strip(): e for e in db.query(Especie).filter(Especie.activo == True).all()}
    variedad_cache = {v.nombre.lower().strip(): v for v in db.query(Variedad).all()}
    pmg_cache = {p.nombre.lower().strip(): p for p in db.query(Pmg).filter(Pmg.activo == True).all()}
    pi_cache = {}
    for p in db.query(Portainjerto).all():
        pi_cache[p.nombre.upper()] = p
        if p.codigo:
            pi_cache[p.codigo.upper()] = p

    structure = {}
    raw_rows = []

    for row in ws.iter_rows(min_row=_MC_DATA_START, values_only=True):
        if not any(v is not None for v in row[:6]):
            continue
        especie_raw = str(row[COL["especie"]] or "").strip()
        variedad_raw = str(row[COL["variedad"]] or "").strip()
        pmg_raw = str(row[COL["pmg"]] or "").strip()
        localidad_raw = str(row[COL["localidad"]] or "").strip()
        if not especie_raw or especie_raw == "nan" or not variedad_raw or variedad_raw == "nan":
            continue

        loc_norm = _normalize_loc(localidad_raw)
        var_clean, pi_code = _extract_rootstock(variedad_raw)
        planta_num = _safe_int(row[COL["planta"]])
        rep_num = _safe_int(row[COL["repeticion"]])

        if loc_norm not in structure:
            structure[loc_norm] = {}
        key = (var_clean.lower(), pi_code)
        if key not in structure[loc_norm]:
            structure[loc_norm][key] = {
                "var_clean": var_clean, "pi_code": pi_code,
                "especie": especie_raw, "pmg": pmg_raw if pmg_raw != "nan" else None,
                "max_planta": 0, "max_rep": 0, "n_rows": 0,
            }
        info = structure[loc_norm][key]
        info["n_rows"] += 1
        if planta_num and planta_num > info["max_planta"]:
            info["max_planta"] = planta_num
        if rep_num and rep_num > info["max_rep"]:
            info["max_rep"] = rep_num
        raw_rows.append(row)

    total_lotes = 0
    total_posiciones = 0
    for loc_data in structure.values():
        for info in loc_data.values():
            info["n_posiciones"] = max(info["max_planta"], info["max_rep"], 1)
            total_lotes += 1
            total_posiciones += info["n_posiciones"]

    logger.info(f"Paso 1: {len(structure)} localidades, {total_lotes} lotes, {total_posiciones} pos, {len(raw_rows)} filas")

    # ── PASO 2: Crear campos, cuarteles, testblocks ──────────────────
    logger.info("Paso 2: Creando estructura...")
    campo_map = {}
    cuartel_map = {}
    tb_map = {}

    for loc_norm, loc_data in structure.items():
        campo_code = _loc_to_campo_code(loc_norm)
        n_var = len(loc_data)
        max_pos = max(info["n_posiciones"] for info in loc_data.values())

        campo = db.query(Campo).filter(Campo.codigo == campo_code).first()
        if not campo:
            campo = Campo(codigo=campo_code, nombre=loc_norm, ubicacion=loc_norm, activo=True)
            db.add(campo)
            db.flush()
        campo_map[loc_norm] = campo

        cuartel_code = f"CUA-{campo_code[4:]}"
        cuartel = db.query(Cuartel).filter(Cuartel.codigo == cuartel_code).first()
        if not cuartel:
            cuartel = Cuartel(id_campo=campo.id_campo, codigo=cuartel_code, nombre=f"Cuartel {loc_norm}",
                              num_hileras=n_var, pos_por_hilera=max_pos, es_testblock=True, activo=True)
            db.add(cuartel)
            db.flush()
        cuartel_map[loc_norm] = cuartel

        tb_code = f"TB-{campo_code[4:]}"
        tb = db.query(TestBlock).filter(TestBlock.codigo == tb_code).first()
        if not tb:
            tb = TestBlock(codigo=tb_code, nombre=f"TestBlock {loc_norm}", id_campo=campo.id_campo,
                           id_cuartel=cuartel.id_cuartel, num_hileras=n_var, posiciones_por_hilera=max_pos,
                           total_posiciones=sum(i["n_posiciones"] for i in loc_data.values()),
                           estado="activo", notas="Migrado desde Excel maestro CAROZOS", activo=True)
            db.add(tb)
            db.flush()
        tb_map[loc_norm] = tb

    db.flush()
    logger.info(f"Paso 2: {len(campo_map)} campos, {len(tb_map)} testblocks")

    # ── PASO 3: Crear posiciones, plantas, lotes ─────────────────────
    logger.info("Paso 3: Creando posiciones, plantas, lotes...")
    link_map = {}
    lote_seq = 0
    lotes_creados = posiciones_creadas = plantas_creadas = 0

    for loc_norm, loc_data in structure.items():
        tb = tb_map[loc_norm]
        cuartel = cuartel_map[loc_norm]
        hilera_num = 0

        for (var_key, pi_code), info in sorted(loc_data.items()):
            hilera_num += 1
            var_clean = info["var_clean"]
            n_pos = info["n_posiciones"]

            esp_obj = especie_cache.get(info["especie"].lower())
            var_obj = variedad_cache.get(var_clean.lower())
            pmg_obj = pmg_cache.get((info["pmg"] or "").lower()) if info["pmg"] else None
            pi_obj = None
            if pi_code != "SIN_PI":
                for variant in [pi_code, pi_code.title(), pi_code.replace(" ", "")]:
                    pi_obj = pi_cache.get(variant.upper())
                    if pi_obj:
                        break

            lote_seq += 1
            esp_code = esp_obj.codigo if esp_obj else "XXX"
            var_code_short = var_obj.codigo.split("-")[-1] if var_obj and var_obj.codigo else f"{lote_seq:03d}"
            codigo_lote = _gen_lote_code(esp_code, var_code_short, pi_code, lote_seq)

            lote = db.query(InventarioVivero).filter(InventarioVivero.codigo_lote == codigo_lote).first()
            if not lote:
                lote = InventarioVivero(
                    codigo_lote=codigo_lote,
                    id_variedad=var_obj.id_variedad if var_obj else None,
                    id_portainjerto=pi_obj.id_portainjerto if pi_obj else None,
                    id_especie=esp_obj.id_especie if esp_obj else None,
                    id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                    tipo_planta="MIGRACION_EXCEL",
                    cantidad_inicial=n_pos, cantidad_actual=0, cantidad_minima=0,
                    fecha_ingreso=date.today(), estado="plantado",
                    observaciones=f"Migración: {var_clean} / {pi_code} en {loc_norm}",
                )
                db.add(lote)
                db.flush()
                lotes_creados += 1

            for pos_num in range(1, n_pos + 1):
                pos_code = f"{tb.codigo}-H{hilera_num:03d}-P{pos_num:03d}"
                pos = db.query(PosicionTestBlock).filter(PosicionTestBlock.codigo_unico == pos_code).first()
                if not pos:
                    pos = PosicionTestBlock(
                        codigo_unico=pos_code, id_cuartel=cuartel.id_cuartel, id_testblock=tb.id_testblock,
                        id_variedad=var_obj.id_variedad if var_obj else None,
                        id_portainjerto=pi_obj.id_portainjerto if pi_obj else None,
                        id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                        id_lote=lote.id_inventario, hilera=hilera_num, posicion=pos_num,
                        estado="alta", fecha_alta=date.today(), fecha_plantacion=date.today(), usuario_alta=usuario,
                    )
                    db.add(pos)
                    db.flush()
                    posiciones_creadas += 1

                planta = db.query(Planta).filter(Planta.codigo == pos_code, Planta.activa == True).first()
                if not planta:
                    planta = Planta(
                        codigo=pos_code, id_posicion=pos.id_posicion,
                        id_variedad=var_obj.id_variedad if var_obj else None,
                        id_portainjerto=pi_obj.id_portainjerto if pi_obj else None,
                        id_especie=esp_obj.id_especie if esp_obj else None,
                        id_pmg=pmg_obj.id_pmg if pmg_obj else None,
                        id_lote_origen=lote.id_inventario, condicion="EN_EVALUACION",
                        activa=True, fecha_alta=date.today(), usuario_creacion=usuario,
                    )
                    db.add(planta)
                    db.flush()
                    plantas_creadas += 1

                link_map[(loc_norm, var_clean.lower(), pi_code, pos_num)] = {
                    "id_posicion": pos.id_posicion, "id_planta": planta.id_planta,
                    "id_lote": lote.id_inventario,
                    "id_variedad": var_obj.id_variedad if var_obj else None,
                    "id_especie": esp_obj.id_especie if esp_obj else None,
                    "id_portainjerto": pi_obj.id_portainjerto if pi_obj else None,
                    "id_campo": campo_map[loc_norm].id_campo,
                }

            mov = MovimientoInventario(
                id_inventario=lote.id_inventario, tipo="CARGA_INICIAL", cantidad=n_pos,
                saldo_anterior=n_pos, saldo_nuevo=0,
                motivo=f"Migración Excel a TB {tb.codigo}", referencia_destino=tb.codigo, usuario=usuario,
            )
            db.add(mov)

    db.commit()
    logger.info(f"Paso 3: {lotes_creados} lotes, {posiciones_creadas} pos, {plantas_creadas} plantas")

    # ── PASO 4: Eliminar mediciones dummy ─────────────────────────────
    logger.info("Paso 4: Eliminando mediciones dummy...")
    dummy_pos = db.query(PosicionTestBlock).filter(
        PosicionTestBlock.codigo_unico == f"{_DUMMY_TB_CODE}-H1-P1"
    ).first()

    meds_eliminadas = clasif_eliminadas = 0
    if dummy_pos:
        dummy_med_ids = [m.id_medicion for m in db.query(MedicionLaboratorio.id_medicion).filter(
            MedicionLaboratorio.id_posicion == dummy_pos.id_posicion).all()]
        if dummy_med_ids:
            # Batch delete to avoid SQL Server parameter limit
            for i in range(0, len(dummy_med_ids), 1000):
                batch = dummy_med_ids[i:i+1000]
                clasif_eliminadas += db.query(ClasificacionCluster).filter(
                    ClasificacionCluster.id_medicion.in_(batch)).delete(synchronize_session=False)
                meds_eliminadas += db.query(MedicionLaboratorio).filter(
                    MedicionLaboratorio.id_medicion.in_(batch)).delete(synchronize_session=False)
        db.commit()

    logger.info(f"Paso 4: {meds_eliminadas} mediciones, {clasif_eliminadas} clasificaciones eliminadas")

    # ── PASO 5: Re-importar mediciones ────────────────────────────────
    logger.info("Paso 5: Re-importando mediciones...")
    wb.close()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb[_MC_SHEET]

    meds_creadas = meds_clasificadas = meds_sin_link = 0
    errors = []
    batch_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=_MC_DATA_START, values_only=True), _MC_DATA_START):
        if not any(v is not None for v in row[:6]):
            continue
        especie_raw = str(row[COL["especie"]] or "").strip()
        variedad_raw = str(row[COL["variedad"]] or "").strip()
        localidad_raw = str(row[COL["localidad"]] or "").strip()
        if not especie_raw or especie_raw == "nan" or not variedad_raw or variedad_raw == "nan":
            continue

        loc_norm = _normalize_loc(localidad_raw)
        var_clean, pi_code = _extract_rootstock(variedad_raw)
        pos_num = _safe_int(row[COL["planta"]]) or _safe_int(row[COL["repeticion"]]) or 1

        link = link_map.get((loc_norm, var_clean.lower(), pi_code, pos_num))
        if not link:
            link = link_map.get((loc_norm, var_clean.lower(), pi_code, 1))
        if not link:
            meds_sin_link += 1
            if len(errors) < 50:
                errors.append({"row": row_idx, "detail": f"Sin link: {loc_norm}/{var_clean}/{pi_code}"})
            continue

        fecha_eval = _to_date(row[COL["fecha_evaluacion"]])
        fecha_cosecha = _to_date(row[COL["fecha_cosecha"]])
        fecha_medicion = fecha_eval or fecha_cosecha
        if not fecha_medicion:
            continue

        f_mej1 = _safe_float(row[COL["firmeza_mejilla1"]])
        f_mej2 = _safe_float(row[COL["firmeza_mejilla2"]])
        firmeza_mejillas = calcular_mejillas_promedio(f_mej1, f_mej2)
        f_punta = _safe_float(row[COL["firmeza_punta"]])
        f_quilla = _safe_float(row[COL["firmeza_quilla"]])
        f_hombro = _safe_float(row[COL["firmeza_hombro"]])

        color_pulpa_raw = row[COL["color_pulpa"]]
        obs_raw = row[COL["observaciones"]]
        obs_parts = []
        if obs_raw:
            obs_parts.append(str(obs_raw))
        if color_pulpa_raw:
            obs_parts.append(f"Color pulpa: {color_pulpa_raw}")

        try:
            sp = db.begin_nested()
            medicion = MedicionLaboratorio(
                id_posicion=link["id_posicion"], id_planta=link["id_planta"],
                id_variedad=link["id_variedad"], id_especie=link["id_especie"],
                id_portainjerto=link["id_portainjerto"], id_campo=link["id_campo"],
                temporada=str(row[COL["temporada"]]).strip() if row[COL["temporada"]] else None,
                fecha_medicion=fecha_medicion, fecha_cosecha=fecha_cosecha,
                brix=_safe_decimal(row[COL["brix"]], 2),
                acidez=_safe_decimal(row[COL["acidez"]], 3),
                firmeza=_safe_decimal(firmeza_mejillas, 1) if firmeza_mejillas else None,
                peso=_safe_decimal(row[COL["peso_g"]], 2),
                color_pct=_safe_int(row[COL["cubr_total"]]),
                observaciones=" | ".join(obs_parts)[:2000] if obs_parts else None,
                usuario_registro=usuario,
                firmeza_punta=_safe_decimal(f_punta, 2) if f_punta else None,
                firmeza_quilla=_safe_decimal(f_quilla, 2) if f_quilla else None,
                firmeza_hombro=_safe_decimal(f_hombro, 2) if f_hombro else None,
                firmeza_mejilla_1=_safe_decimal(f_mej1, 2) if f_mej1 else None,
                firmeza_mejilla_2=_safe_decimal(f_mej2, 2) if f_mej2 else None,
                perimetro=_safe_decimal(row[COL["perimetro"]], 2),
                n_muestra=_safe_int(row[COL["fruto_num"]]),
                periodo_almacenaje=_safe_int(row[COL["periodo_almacenaje"]]),
                repeticion=_safe_int(row[COL["repeticion"]]),
                raleo_frutos=_safe_int(row[COL["raleo"]]),
                rendimiento=_safe_decimal(row[COL["rendimiento"]], 2),
                color_0_30=_safe_int(row[COL["cubr_0_30"]]),
                color_30_50=_safe_int(row[COL["cubr_30_50"]]),
                color_50_75=_safe_int(row[COL["cubr_50_75"]]),
                color_75_100=_safe_int(row[COL["cubr_75_100"]]),
                color_total=_safe_int(row[COL["cubr_total"]]),
                color_verde=_safe_int(row[COL["color_verde"]]),
                color_crema=_safe_int(row[COL["color_crema"]]),
                color_amarillo=_safe_int(row[COL["color_amarillo"]]),
                color_full=_safe_int(row[COL["color_full"]]),
                color_dist_total=_safe_int(row[COL["color_total"]]),
                pardeamiento=_safe_decimal(row[COL["pardeamiento"]], 2),
                traslucidez=_safe_decimal(row[COL["traslucidez"]], 2),
                gelificacion=_safe_decimal(row[COL["gelificacion"]], 2),
                harinosidad=_safe_decimal(row[COL["harinosidad"]], 2),
                total_frutos_pardeamiento=_safe_int(row[COL["pardeamiento_total"]]),
                total_frutos_traslucidez=_safe_int(row[COL["traslucidez_total"]]),
                total_frutos_gelificacion=_safe_int(row[COL["gelificacion_total"]]),
                total_frutos_harinosidad=_safe_int(row[COL["harinosidad_total"]]),
                color_pulpa=str(color_pulpa_raw).strip() if color_pulpa_raw else None,
            )
            db.add(medicion)
            db.flush()
            meds_creadas += 1

            # Auto-clasificar
            punto_debil = calcular_punto_debil(f_punta, f_quilla, f_hombro)
            regla = determinar_regla(
                especie=especie_raw,
                peso_promedio=_safe_float(row[COL["peso_g"]]),
                color_pulpa=str(color_pulpa_raw).strip() if color_pulpa_raw else None,
                fecha_evaluacion=fecha_medicion,
            )
            result = clasificar_bandas(
                brix=_safe_float(row[COL["brix"]]),
                acidez=_safe_float(row[COL["acidez"]]),
                firmeza_mejillas=firmeza_mejillas,
                firmeza_punto_debil=punto_debil,
                regla=regla,
            )
            clasif = ClasificacionCluster(
                id_medicion=medicion.id_medicion,
                cluster=result["cluster"], banda_brix=result["banda_brix"],
                banda_firmeza=result["banda_firmeza"], banda_acidez=result["banda_acidez"],
                banda_calibre=result["banda_firmeza_punto"],
                score_total=Decimal(str(result["suma_bandas"])),
                metodo="migracion_v1",
            )
            db.add(clasif)
            meds_clasificadas += 1
        except Exception as exc:
            sp.rollback()
            if len(errors) < 50:
                errors.append({"row": row_idx, "error": str(exc)[:200]})
            continue

        batch_count += 1
        if batch_count >= BATCH_SIZE:
            db.commit()
            batch_count = 0

    if batch_count > 0:
        db.commit()

    logger.info(f"Paso 5: {meds_creadas} mediciones, {meds_clasificadas} clasificadas, {meds_sin_link} sin link")

    # ── PASO 6: Limpiar dummy ─────────────────────────────────────────
    dummy_cleaned = False
    if dummy_pos:
        remaining = db.query(MedicionLaboratorio).filter(
            MedicionLaboratorio.id_posicion == dummy_pos.id_posicion).count()
        if remaining == 0:
            dummy_tb = db.query(TestBlock).filter(TestBlock.codigo == _DUMMY_TB_CODE).first()
            if dummy_tb:
                dummy_tb.activo = False
            dummy_pos.estado = "baja"
            dummy_pos.motivo_baja = "Migración completada"
            db.commit()
            dummy_cleaned = True

    wb.close()
    return {
        "status": "ok",
        "estructura": {"campos": len(campo_map), "testblocks": len(tb_map),
                        "lotes": lotes_creados, "posiciones": posiciones_creadas, "plantas": plantas_creadas},
        "mediciones": {"dummy_eliminadas": meds_eliminadas, "clasificaciones_eliminadas": clasif_eliminadas,
                        "creadas": meds_creadas, "clasificadas": meds_clasificadas, "sin_link": meds_sin_link},
        "dummy_limpiado": dummy_cleaned, "errores": errors[:20],
    }
