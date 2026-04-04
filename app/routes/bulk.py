"""Bulk import/export endpoints for Excel files."""

import logging
from datetime import datetime, date
from app.core.utils import utcnow
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import inspect as sa_inspect

from app.core.database import get_db
from app.core.deps import get_current_user, require_role
from app.models.sistema import Usuario
from app.models.maestras import Campo, Especie, Portainjerto, Pmg, Vivero, Cuartel
from app.models.variedades import Variedad
from app.models.inventario import InventarioVivero
from app.models.testblock import Planta, TestBlock, PosicionTestBlock
from app.models.laboratorio import MedicionLaboratorio, ClasificacionCluster
from app.services.clustering_service import (
    clasificar_medicion as clasificar_bandas,
    determinar_regla,
    calcular_mejillas_promedio,
    calcular_punto_debil,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/bulk", tags=["Bulk Import/Export"])

# ── Template definitions ─────────────────────────────────────────────────────
TEMPLATES: dict[str, dict[str, Any]] = {
    "variedades": {
        "columns": [
            "codigo", "nombre", "id_especie", "id_pmg", "id_origen",
            "tipo", "estado", "epoca_cosecha", "vigor",
        ],
        "examples": [
            ["VAR-001", "Cherry Example", 1, 1, 1, "plantada", "prospecto", "temprana", "medio"],
            ["VAR-002", "Plum Example", 2, 2, 1, "prospecto", "en_evaluacion", "media", "alto"],
        ],
    },
    "portainjertos": {
        "columns": ["codigo", "nombre", "vigor", "tipo", "origen"],
        "examples": [
            ["PI-001", "MaxMa 14", "medio", "seleccion", "Oregeon USA"],
            ["PI-002", "Gisela 6", "bajo", "hibrido", "Giessen Alemania"],
        ],
    },
    "especies": {
        "columns": ["codigo", "nombre", "nombre_cientifico"],
        "examples": [
            ["ESP-001", "Cerezo", "Prunus avium"],
            ["ESP-002", "Ciruela", "Prunus domestica"],
        ],
    },
    "pmg": {
        "columns": [
            "nombre", "pais", "ciudad", "contacto_nombre",
            "contacto_email", "contacto_telefono",
        ],
        "examples": [
            ["Peter Stoppel - Cerasina", "Alemania", "Kressbronn", "Pablo Courbis", "info@cerezas.cl", "+56998280601"],
            ["Hortifrut", "Chile", "Santiago", "Juan Perez", "jp@hortifrut.cl", "+56912345678"],
        ],
    },
    "viveros": {
        "columns": ["codigo", "nombre", "representante", "telefono", "email"],
        "examples": [
            ["VIV-001", "Vivero Los Andes", "Juan Perez", "+56912345678", "juan@vivero.cl"],
            ["VIV-002", "Vivero Sur", "Maria Lopez", "+56987654321", "maria@viverosur.cl"],
        ],
    },
    "campos": {
        "columns": ["codigo", "nombre", "ubicacion", "comuna", "region"],
        "examples": [
            ["CAM-001", "Campo Norte", "Km 5 Ruta 5", "Los Andes", "Valparaiso"],
            ["CAM-002", "Campo Sur", "Km 10 Ruta 5", "Rancagua", "O'Higgins"],
        ],
    },
    "inventario": {
        "columns": [
            "codigo_lote", "id_variedad", "id_portainjerto", "id_especie",
            "id_vivero", "id_bodega", "tipo_planta", "cantidad_inicial",
            "cantidad_actual", "fecha_ingreso", "estado",
        ],
        "examples": [
            ["LOTE-001", 1, 1, 1, 1, 1, "Barbado", 100, 100, "2025-06-01", "disponible"],
            ["LOTE-002", 2, 2, 2, 1, 1, "Injertado", 50, 50, "2025-06-15", "disponible"],
        ],
    },
    "plantas": {
        "columns": [
            "codigo", "id_variedad", "id_portainjerto", "id_especie",
            "id_lote_origen", "condicion", "ano_plantacion",
        ],
        "examples": [
            ["PL-001", 1, 1, 1, 1, "EN_EVALUACION", 2025],
            ["PL-002", 2, 2, 2, 2, "EN_EVALUACION", 2025],
        ],
    },
    "mediciones": {
        "columns": [
            "especie", "temporada", "campo", "variedad", "portainjerto", "pmg",
            "fecha_cosecha", "fecha_evaluacion", "n_fruto", "repeticion",
            "peso_g", "perimetro_mm", "rendimiento_g",
            "firmeza_punta", "firmeza_quilla", "firmeza_hombro", "firmeza_mejilla_1", "firmeza_mejilla_2",
            "brix", "acidez",
            "color_0_30", "color_30_50", "color_50_75", "color_75_100",
            "color_pulpa",
            "periodo_almacenaje", "pardeamiento", "traslucidez", "gelificacion", "harinosidad",
            "observaciones",
        ],
        "examples": [
            [
                "Cerezo", "2024-2025", "Santa Margarita", "PRIM 2.5", "Colt", "Peter Stoppel",
                "2024-11-28", "2024-11-28", 1, 1,
                12.4, 28.5, 85.0,
                7.3, 7.0, 7.6, 9.1, 8.8,
                18.6, 0.86,
                10, 30, 40, 20,
                "Roja",
                None, None, None, None, None,
                "Fruto sano",
            ],
        ],
    },
}

# Map entity names to SQLModel classes
MODEL_MAP: dict[str, type] = {
    "variedades": Variedad,
    "portainjertos": Portainjerto,
    "especies": Especie,
    "viveros": Vivero,
    "campos": Campo,
    "pmg": Pmg,
    "inventario": InventarioVivero,
    "plantas": Planta,
}

# Columns to exclude from export (binary/large data)
EXPORT_EXCLUDE: dict[str, set[str]] = {
    "variedades": {"imagen"},
    "portainjertos": {"imagen"},
}

# Header style constants
HEADER_FILL = PatternFill(start_color="8B1A2B", end_color="8B1A2B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _workbook_to_stream(wb: openpyxl.Workbook) -> BytesIO:
    """Save workbook to a BytesIO stream."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _auto_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max(max_len + 3, 12)
        ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 50)


# ── Download template ─────────────────────────────────────────────────────────
@router.get("/template/{entity}")
def download_template(entity: str):
    """Descarga un archivo Excel de plantilla con headers y 2 filas de ejemplo."""
    template = TEMPLATES.get(entity)
    if not template:
        available = ", ".join(sorted(TEMPLATES.keys()))
        raise HTTPException(
            status_code=404,
            detail=f"Template '{entity}' no existe. Disponibles: {available}",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity

    # Header row
    for col_idx, col_name in enumerate(template["columns"], 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Example rows
    for row_idx, example in enumerate(template["examples"], 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    _auto_width(ws)

    buf = _workbook_to_stream(wb)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=template_{entity}.xlsx"},
    )


# ── Import from Excel ─────────────────────────────────────────────────────────
@router.post("/import/{entity}")
def import_data(
    entity: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Importa registros desde un archivo Excel. Inserta fila por fila para
    reportar errores individuales sin perder las filas validas."""
    model = MODEL_MAP.get(entity)
    if not model:
        available = ", ".join(sorted(MODEL_MAP.keys()))
        raise HTTPException(
            status_code=400,
            detail=f"Entidad '{entity}' no soporta importacion. Disponibles: {available}",
        )

    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo sin nombre")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {str(exc)[:200]}")

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="El archivo no tiene hojas activas")

    # Get headers from first row
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    if not headers:
        raise HTTPException(status_code=400, detail="No se encontraron columnas en la primera fila")

    # Get valid column names for the model
    mapper = sa_inspect(model)
    valid_columns = {col.key for col in mapper.mapper.column_attrs}

    created = 0
    errors: list[dict[str, Any]] = []
    total_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(v is not None for v in row):
            continue
        total_rows += 1

        try:
            # Build data dict, only including valid model columns
            data: dict[str, Any] = {}
            for i, header in enumerate(headers):
                if i < len(row) and row[i] is not None:
                    col_name = str(header).strip()
                    if col_name in valid_columns:
                        data[col_name] = row[i]

            if not data:
                errors.append({"row": row_idx, "error": "Fila sin datos validos"})
                continue

            # Set audit fields if the model has them
            if "usuario_creacion" in valid_columns:
                data["usuario_creacion"] = user.username
            if "fecha_creacion" in valid_columns:
                data["fecha_creacion"] = utcnow()

            sp = db.begin_nested()
            obj = model(**data)
            db.add(obj)
            db.flush()
            created += 1
        except Exception as exc:
            sp.rollback()
            errors.append({"row": row_idx, "error": str(exc)[:200]})

    if created > 0:
        try:
            db.commit()
        except Exception as exc:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Error al guardar los datos: {str(exc)[:200]}",
            )

    return {
        "created": created,
        "errors": errors,
        "total_rows": total_rows,
    }


# ── Export to Excel ───────────────────────────────────────────────────────────
@router.get("/export/{entity}")
def export_data(
    entity: str,
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Exporta todos los registros activos de una entidad a un archivo Excel."""
    model = MODEL_MAP.get(entity)
    if not model:
        available = ", ".join(sorted(MODEL_MAP.keys()))
        raise HTTPException(
            status_code=400,
            detail=f"Entidad '{entity}' no soporta exportacion. Disponibles: {available}",
        )

    # Query active records — check if model uses 'activo' or 'activa'
    query = db.query(model)
    if hasattr(model, "activo"):
        query = query.filter(model.activo == True)  # noqa: E712
    elif hasattr(model, "activa"):
        query = query.filter(model.activa == True)  # noqa: E712

    records = query.all()

    # Get column names, excluding binary/large fields
    mapper = sa_inspect(model)
    exclude = EXPORT_EXCLUDE.get(entity, set())
    columns = [
        col.key for col in mapper.mapper.column_attrs
        if col.key not in exclude and col.key != "password_hash"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = getattr(record, col_name, None)
            # Convert datetime/date to string for Excel compatibility
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif hasattr(value, "isoformat"):
                value = value.isoformat()
            # Skip bytes fields
            if isinstance(value, bytes):
                value = "(binario)"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

    _auto_width(ws)

    buf = _workbook_to_stream(wb)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=export_{entity}_{utcnow().strftime('%Y%m%d')}.xlsx"
        },
    )


# ── Maestro Carozos Import ───────────────────────────────────────────────────
# Constantes del Excel Maestro Carozos
_MC_HEADER_ROW = 3
_MC_DATA_START = 4
_MC_SHEET = "CAROZOS"
_MC_IMPORT_TB = "TB-IMPORT-CAROZOS"
_MC_IMPORT_CAMPO = "CAM-IMPORT"
_MC_IMPORT_CUARTEL = "CUA-IMPORT"

# Mapeo de columnas (0-indexed) en la hoja CAROZOS
_MC_COL = {
    "especie": 0, "variedad": 1, "pmg": 2, "localidad": 3, "temporada": 4,
    "fecha_cosecha": 5, "fecha_evaluacion": 6, "periodo_almacenaje": 7,
    "raleo": 8, "planta": 9, "perimetro": 10, "rendimiento": 11,
    "repeticion": 12,
    "cubr_0_30": 13, "cubr_30_50": 14, "cubr_50_75": 15, "cubr_75_100": 16,
    "cubr_total": 17,
    "color_verde": 18, "color_crema": 19, "color_amarillo": 20,
    "color_full": 21, "color_total": 22,
    "color_pulpa": 23, "fruto_num": 24, "peso_g": 25,
    "firmeza_punta": 26, "firmeza_quilla": 27, "firmeza_hombro": 28,
    "firmeza_mejilla1": 29, "firmeza_mejilla2": 30,
    "brix": 31, "acidez": 32,
    "pardeamiento": 33, "pardeamiento_total": 34,
    "traslucidez": 35, "traslucidez_total": 36,
    "gelificacion": 37, "gelificacion_total": 38,
    "harinosidad": 39, "harinosidad_total": 40,
    "observaciones": 41,
}

# Especies esperadas: nombre_lower -> (codigo, nombre, cientifico)
_MC_ESPECIES = {
    "ciruela": ("CRU", "Ciruela", "Prunus domestica"),
    "damasco": ("DAM", "Damasco", "Prunus armeniaca"),
    "durazno": ("DUR", "Durazno", "Prunus persica"),
    "nectarin": ("NEC", "Nectarin", "Prunus persica var. nucipersica"),
    "paraguayo": ("PAR", "Paraguayo", "Prunus persica var. platycarpa"),
}


def _mc_safe_decimal(value, precision: int = 2) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return round(Decimal(str(value)), precision)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _mc_safe_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _mc_safe_int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _mc_normalize(s: Optional[str]) -> str:
    return str(s).strip().lower() if s else ""


def _mc_to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _mc_ensure_infrastructure(db: Session) -> PosicionTestBlock:
    """Crea campo/cuartel/testblock/posicion dummy para mediciones importadas."""
    campo = db.query(Campo).filter(Campo.codigo == _MC_IMPORT_CAMPO).first()
    if not campo:
        campo = Campo(
            codigo=_MC_IMPORT_CAMPO, nombre="Import Masivo Carozos",
            ubicacion="Importacion desde Excel maestro", activo=True,
            usuario_creacion="import_maestro",
        )
        db.add(campo)
        db.flush()

    cuartel = db.query(Cuartel).filter(Cuartel.codigo == _MC_IMPORT_CUARTEL).first()
    if not cuartel:
        cuartel = Cuartel(
            id_campo=campo.id_campo, codigo=_MC_IMPORT_CUARTEL,
            nombre="Import Masivo", activo=True,
        )
        db.add(cuartel)
        db.flush()

    tb = db.query(TestBlock).filter(TestBlock.codigo == _MC_IMPORT_TB).first()
    if not tb:
        tb = TestBlock(
            codigo=_MC_IMPORT_TB, nombre="Import Maestro Carozos",
            id_campo=campo.id_campo, id_cuartel=cuartel.id_cuartel,
            num_hileras=1, posiciones_por_hilera=1, total_posiciones=1,
            estado="activo",
            notas="Testblock virtual para mediciones importadas desde Excel Maestro Carozos",
            activo=True,
        )
        db.add(tb)
        db.flush()

    pos_code = f"{_MC_IMPORT_TB}-H1-P1"
    pos = db.query(PosicionTestBlock).filter(
        PosicionTestBlock.codigo_unico == pos_code
    ).first()
    if not pos:
        pos = PosicionTestBlock(
            codigo_unico=pos_code, id_cuartel=cuartel.id_cuartel,
            id_testblock=tb.id_testblock, hilera=1, posicion=1,
            estado="ocupada",
            observaciones="Posicion virtual para importacion masiva",
            usuario_alta="import_maestro",
        )
        db.add(pos)
        db.flush()

    return pos


def _mc_get_or_create_especies(db: Session) -> dict[str, Especie]:
    """Carga especies existentes y crea las que falten."""
    existentes = db.query(Especie).filter(Especie.activo == True).all()  # noqa: E712
    cache: dict[str, Especie] = {_mc_normalize(e.nombre): e for e in existentes}

    for nombre_lower, (codigo, nombre, cientifico) in _MC_ESPECIES.items():
        if nombre_lower in cache:
            continue
        exists_code = db.query(Especie).filter(Especie.codigo == codigo).first()
        if exists_code:
            cache[nombre_lower] = exists_code
            continue
        esp = Especie(
            codigo=codigo, nombre=nombre, nombre_cientifico=cientifico,
            activo=True, usuario_creacion="import_maestro",
        )
        db.add(esp)
        db.flush()
        cache[nombre_lower] = esp

    return cache


def _mc_get_or_create_pmgs(db: Session, pmg_names: set[str]) -> dict[str, Pmg]:
    """Carga PMGs existentes y crea los que falten."""
    existentes = db.query(Pmg).filter(Pmg.activo == True).all()  # noqa: E712
    cache: dict[str, Pmg] = {_mc_normalize(p.nombre): p for p in existentes}

    for nombre_raw in sorted(pmg_names):
        key = _mc_normalize(nombre_raw)
        if not key or key in cache:
            continue
        codigo = nombre_raw.strip().replace(" ", "_").upper()[:20]
        exists_code = db.query(Pmg).filter(Pmg.codigo == codigo).first()
        if exists_code:
            cache[key] = exists_code
            continue
        pmg = Pmg(
            codigo=codigo, nombre=nombre_raw.strip(),
            activo=True, usuario_creacion="import_maestro",
        )
        db.add(pmg)
        db.flush()
        cache[key] = pmg

    return cache


def _mc_get_or_create_variedades(
    db: Session,
    var_set: set[tuple[str, str, str]],
    especie_cache: dict[str, Especie],
    pmg_cache: dict[str, Pmg],
) -> dict[str, Variedad]:
    """Crea variedades que no existan. Dedup por nombre case-insensitive."""
    existentes = db.query(Variedad).all()
    cache: dict[str, Variedad] = {_mc_normalize(v.nombre): v for v in existentes}

    created = 0
    for especie_raw, var_raw, pmg_raw in sorted(var_set):
        var_key = _mc_normalize(var_raw)
        if not var_key or var_key in cache:
            continue

        especie_obj = especie_cache.get(_mc_normalize(especie_raw))
        pmg_obj = pmg_cache.get(_mc_normalize(pmg_raw))
        esp_code = especie_obj.codigo if especie_obj else "XX"
        codigo = f"{esp_code}-{var_raw.strip().replace(' ', '-').upper()[:15]}"

        exists_code = db.query(Variedad).filter(Variedad.codigo == codigo).first()
        if exists_code:
            cache[var_key] = exists_code
            continue

        var = Variedad(
            codigo=codigo, nombre=var_raw.strip(),
            id_especie=especie_obj.id_especie if especie_obj else None,
            id_pmg=pmg_obj.id_pmg if pmg_obj else None,
            tipo="plantada", estado="en_evaluacion", activo=True,
            usuario_creacion="import_maestro",
        )
        db.add(var)
        db.flush()
        cache[var_key] = var
        created += 1

    return cache


@router.post("/import-maestro-carozos")
def import_maestro_carozos(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Usuario = Depends(require_role("admin")),
):
    """Importa el Excel MAESTRO CAROZOS con variedades y mediciones.

    Procesa la hoja 'CAROZOS':
    - Crea/verifica especies (Ciruela, Damasco, Durazno, Nectarin, Paraguayo)
    - Crea/verifica PMGs (programas de mejoramiento genetico)
    - Crea/verifica variedades (dedup por nombre case-insensitive)
    - Importa mediciones de laboratorio (brix, acidez, firmeza, peso)
    - Auto-clasifica cada medicion usando el motor de clustering Band-Sum

    Es idempotente: las entidades maestras no se duplican.
    Las mediciones se insertan siempre (no hay dedup por fila).
    """
    # Validar archivo
    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo sin nombre")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .xls")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo leer el archivo Excel: {str(exc)[:200]}",
        )

    if _MC_SHEET not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"Hoja '{_MC_SHEET}' no encontrada. Hojas disponibles: {wb.sheetnames}",
        )

    ws = wb[_MC_SHEET]

    # --- Paso 1: Recolectar datos unicos ---
    pmg_names: set[str] = set()
    var_set: set[tuple[str, str, str]] = set()
    total_data_rows = 0

    for row in ws.iter_rows(min_row=_MC_DATA_START, values_only=True):
        if not any(v is not None for v in row[:6]):
            continue
        total_data_rows += 1
        especie = row[_MC_COL["especie"]]
        variedad = row[_MC_COL["variedad"]]
        pmg = row[_MC_COL["pmg"]]
        if pmg:
            pmg_names.add(str(pmg).strip())
        if variedad and especie:
            var_set.add((
                str(especie).strip(),
                str(variedad).strip(),
                str(pmg).strip() if pmg else "",
            ))

    # --- Paso 2: Crear entidades maestras ---
    try:
        especie_cache = _mc_get_or_create_especies(db)
        db.flush()
        pmg_cache = _mc_get_or_create_pmgs(db, pmg_names)
        db.flush()
        variedad_cache = _mc_get_or_create_variedades(
            db, var_set, especie_cache, pmg_cache
        )
        db.flush()
        dummy_pos = _mc_ensure_infrastructure(db)
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("Error creando entidades maestras: %s", exc)
        raise HTTPException(
            status_code=500,
            detail=f"Error creando entidades maestras: {str(exc)[:300]}",
        )

    # Contar cuantas especies/pmgs/variedades se crearon vs existian
    especies_count = len(especie_cache)
    pmgs_count = len(pmg_cache)
    variedades_count = len(variedad_cache)

    # --- Paso 3: Importar mediciones ---
    # Re-leer el workbook para segunda pasada
    wb.close()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb[_MC_SHEET]

    counts = {
        "mediciones_creadas": 0, "clasificadas": 0,
        "sin_fecha": 0, "vacias": 0, "errores": 0,
    }
    errors_detail: list[dict[str, Any]] = []
    batch_count = 0
    BATCH_SIZE = 500

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=_MC_DATA_START, values_only=True), _MC_DATA_START
    ):
        if not any(v is not None for v in row[:32]):
            counts["vacias"] += 1
            continue

        especie_raw = row[_MC_COL["especie"]]
        variedad_raw = row[_MC_COL["variedad"]]
        temporada = row[_MC_COL["temporada"]]
        fecha_cosecha = _mc_to_date(row[_MC_COL["fecha_cosecha"]])
        fecha_eval = _mc_to_date(row[_MC_COL["fecha_evaluacion"]])
        fecha_medicion = fecha_eval or fecha_cosecha

        if fecha_medicion is None:
            counts["sin_fecha"] += 1
            continue

        # Valores de medicion
        brix = _mc_safe_decimal(row[_MC_COL["brix"]], 2)
        acidez = _mc_safe_decimal(row[_MC_COL["acidez"]], 3)
        peso = _mc_safe_decimal(row[_MC_COL["peso_g"]], 2)

        # Firmeza
        f_punta = _mc_safe_float(row[_MC_COL["firmeza_punta"]])
        f_quilla = _mc_safe_float(row[_MC_COL["firmeza_quilla"]])
        f_hombro = _mc_safe_float(row[_MC_COL["firmeza_hombro"]])
        f_mej1 = _mc_safe_float(row[_MC_COL["firmeza_mejilla1"]])
        f_mej2 = _mc_safe_float(row[_MC_COL["firmeza_mejilla2"]])

        firmeza_mejillas = calcular_mejillas_promedio(f_mej1, f_mej2)
        firmeza = _mc_safe_decimal(firmeza_mejillas, 1) if firmeza_mejillas else None

        cubr_total = _mc_safe_int(row[_MC_COL["cubr_total"]])

        # Construir observaciones con metadatos del Excel
        obs_parts = []
        obs_raw = row[_MC_COL["observaciones"]]
        if obs_raw:
            obs_parts.append(str(obs_raw))
        periodo = row[_MC_COL["periodo_almacenaje"]]
        if periodo:
            obs_parts.append(f"Periodo: {periodo}")
        color_pulpa_raw = row[_MC_COL["color_pulpa"]]
        if color_pulpa_raw:
            obs_parts.append(f"Color pulpa: {color_pulpa_raw}")

        # Desordenes postcosecha
        desordenes = []
        for key, label in [
            ("pardeamiento", "Pard"), ("traslucidez", "Trasl"),
            ("gelificacion", "Gelif"), ("harinosidad", "Harin"),
        ]:
            val = row[_MC_COL[key]]
            if val:
                desordenes.append(f"{label}={val}")
        if desordenes:
            obs_parts.append("Desordenes: " + ", ".join(desordenes))

        # Firmeza detallada
        firmeza_detail = []
        for val, label in [
            (f_punta, "Pta"), (f_quilla, "Qui"), (f_hombro, "Hmb"),
            (f_mej1, "Mj1"), (f_mej2, "Mj2"),
        ]:
            if val is not None:
                firmeza_detail.append(f"{label}={val}")
        if firmeza_detail:
            obs_parts.append("Firmeza: " + ", ".join(firmeza_detail))

        localidad = row[_MC_COL["localidad"]]
        if localidad:
            obs_parts.append(f"Loc: {localidad}")

        observaciones = " | ".join(obs_parts) if obs_parts else None
        if observaciones and len(observaciones) > 2000:
            observaciones = observaciones[:1997] + "..."

        try:
            sp = db.begin_nested()
            medicion = MedicionLaboratorio(
                id_posicion=dummy_pos.id_posicion,
                id_planta=None,
                temporada=str(temporada).strip() if temporada else None,
                fecha_medicion=fecha_medicion,
                fecha_cosecha=fecha_cosecha,
                brix=brix,
                acidez=acidez,
                firmeza=firmeza,
                calibre=None,
                peso=peso,
                color_pct=cubr_total,
                cracking_pct=None,
                observaciones=observaciones,
                usuario_registro=user.username,
            )
            db.add(medicion)
            db.flush()
            counts["mediciones_creadas"] += 1

            # Auto-clasificar
            especie_nombre = str(especie_raw).strip() if especie_raw else "Ciruela"
            peso_float = _mc_safe_float(row[_MC_COL["peso_g"]])
            color_pulpa_str = str(color_pulpa_raw).strip() if color_pulpa_raw else None

            if not color_pulpa_str:
                var_key = _mc_normalize(str(variedad_raw)) if variedad_raw else ""
                var_obj = variedad_cache.get(var_key)
                if var_obj and var_obj.color_pulpa:
                    color_pulpa_str = var_obj.color_pulpa

            regla = determinar_regla(
                especie=especie_nombre,
                peso_promedio=peso_float,
                color_pulpa=color_pulpa_str,
                fecha_evaluacion=fecha_medicion,
            )

            punto_debil = calcular_punto_debil(f_punta, f_quilla, f_hombro)

            result = clasificar_bandas(
                brix=_mc_safe_float(row[_MC_COL["brix"]]),
                acidez=_mc_safe_float(row[_MC_COL["acidez"]]),
                firmeza_mejillas=firmeza_mejillas if firmeza_mejillas else None,
                firmeza_punto_debil=punto_debil if punto_debil else None,
                regla=regla,
            )

            clasif = ClasificacionCluster(
                id_medicion=medicion.id_medicion,
                cluster=result["cluster"],
                banda_brix=result["banda_brix"],
                banda_firmeza=result["banda_firmeza"],
                banda_acidez=result["banda_acidez"],
                banda_calibre=result["banda_firmeza_punto"],
                score_total=Decimal(str(result["suma_bandas"])),
                metodo="import_maestro_v1",
            )
            db.add(clasif)
            counts["clasificadas"] += 1

        except Exception as exc:
            counts["errores"] += 1
            if len(errors_detail) < 20:
                errors_detail.append({"row": row_idx, "error": str(exc)[:200]})
            sp.rollback()
            continue

        batch_count += 1
        if batch_count >= BATCH_SIZE:
            db.commit()
            batch_count = 0

    # Commit final
    if batch_count > 0:
        try:
            db.commit()
        except Exception as exc:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Error en commit final: {str(exc)[:200]}",
            )

    wb.close()

    return {
        "total_filas_datos": total_data_rows,
        "especies_en_cache": especies_count,
        "pmgs_en_cache": pmgs_count,
        "variedades_en_cache": variedades_count,
        "mediciones_creadas": counts["mediciones_creadas"],
        "mediciones_clasificadas": counts["clasificadas"],
        "filas_sin_fecha": counts["sin_fecha"],
        "filas_vacias": counts["vacias"],
        "errores": counts["errores"],
        "errores_detalle": errors_detail,
    }
